import os
import io
import uuid
import math
import datetime as dt
from datetime import datetime, timedelta, date
from io import BytesIO
from collections import defaultdict

import pandas as pd
from dotenv import load_dotenv
from werkzeug.security import generate_password_hash, check_password_hash
from sqlalchemy import func, or_
from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify
import traceback as tb_module
from flask_sqlalchemy import SQLAlchemy
from flask_login import (
    LoginManager, login_user, login_required, logout_user, current_user, UserMixin
)

# ------------------ RBAC Configuration ------------------
ROLES = ['Admin', 'Management', 'CategoryManager', 'WarehouseOps', 'Viewer']

ROLE_PERMISSIONS = {
    'Admin': [
        'dashboard:view', 'sales:upload', 'sales_macro:upload', 'stock_store:upload', 'stock_cd:upload',
        'stock:query', 'distribution:generate', 'distribution:export',
        'forecast_v2:view', 'forecast_v2:run', 'runs:view', 'runs:submit', 'runs:approve', 'admin:users', 'admin:reset', 'audit:view',
        'rebalancing:view', 'rebalancing:run', 'slow_stock:view', 'slow_stock:run', 'store_health:view', 'alerts:view',
        'planner:view', 'planner:operate'
    ],
    'Management': [
        'dashboard:view', 'stock:query', 'distribution:export',
        'forecast_v2:view', 'runs:view', 'runs:approve', 'audit:view', 'rebalancing:view', 'slow_stock:view', 'store_health:view', 'alerts:view',
        'planner:view'
    ],
    'CategoryManager': [
        'dashboard:view', 'sales:upload', 'sales_macro:upload', 'distribution:generate', 'distribution:export',
        'forecast_v2:view', 'forecast_v2:run', 'runs:view', 'runs:submit', 'rebalancing:view', 'rebalancing:run',
        'slow_stock:view', 'slow_stock:run', 'store_health:view', 'alerts:view',
        'planner:view'
    ],
    'WarehouseOps': [
        'dashboard:view', 'stock_store:upload', 'stock_cd:upload', 'stock:query',
        'distribution:export', 'rebalancing:view', 'rebalancing:run', 'slow_stock:view', 'slow_stock:run', 'store_health:view', 'alerts:view',
        'planner:view', 'planner:operate'
    ],
    'Viewer': [
        'dashboard:view', 'stock:query', 'alerts:view'
    ]
}
MIN_WEEKS = 3  # mínimo de semanas de historia requeridas por SKU–Tienda

# ------------------ Stock-out Replenishment Constants ------------------
STOCKOUT_RECENT_WEEKS = 1      # Weeks to check for "no recent sales"
STOCKOUT_HIST_WEEKS = 8        # Weeks of historical data to compute average
STOCKOUT_TARGET_WOC = 1.0      # Target weeks of cover for replenishment
STOCKOUT_MAX_QTY = 3           # Maximum qty per store for stock-out replenishment
STOCKOUT_DEBUG = False         # Enable debug logging for stock-out layer


# ------------------ Pagination Helper ------------------
class Pagination:
    def __init__(self, page, per_page, total, items):
        self.page = page
        self.per_page = per_page
        self.total = total
        self.items = items
        self.pages = max(1, (total + per_page - 1) // per_page)
    
    @property
    def has_prev(self):
        return self.page > 1
    
    @property
    def has_next(self):
        return self.page < self.pages
    
    @property
    def prev_num(self):
        return self.page - 1 if self.has_prev else None
    
    @property
    def next_num(self):
        return self.page + 1 if self.has_next else None
    
    def iter_pages(self, left_edge=1, left_current=2, right_current=2, right_edge=1):
        last = 0
        for num in range(1, self.pages + 1):
            if num <= left_edge or \
               (self.page - left_current <= num <= self.page + right_current) or \
               num > self.pages - right_edge:
                if last + 1 != num:
                    yield None
                yield num
                last = num


# ------------------ Config ------------------
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
load_dotenv(os.path.join(BASE_DIR, '.env'))

app = Flask(__name__)  # ✅ primero se crea app
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'dev-secret')

db_path = os.path.join(BASE_DIR, 'app.db')  # ✅ app.db junto a app.py

app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv(
    'SQLALCHEMY_DATABASE_URI',
    f'sqlite:///{db_path}'
)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)  # ✅ si en tu proyecto aún no existe db más abajo

login_manager = LoginManager(app)
login_manager.login_view = 'login'

# ------------------ Global AJAX Error Handler ------------------
from werkzeug.exceptions import HTTPException

@app.errorhandler(Exception)
def handle_exception(e):
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    
    if isinstance(e, HTTPException):
        code = e.code
        message = e.description
    else:
        code = 500
        message = str(e) if str(e) else 'Internal Server Error'
    
    if is_ajax:
        response = {
            'ok': False,
            'message': message,
            'category': 'danger'
        }
        if app.debug:
            response['traceback'] = tb_module.format_exc()
        return jsonify(response), code
    
    if isinstance(e, HTTPException):
        return e
    
    raise e

# ------------------ Modelos ------------------


class User(db.Model, UserMixin):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)
    role = db.Column(db.String(50), nullable=False, default='Viewer')
    is_active = db.Column(db.Boolean, nullable=False, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

    def has_permission(self, permission):
        # Admin role bypasses all permission checks
        if self.role == 'Admin':
            return True
        permissions = ROLE_PERMISSIONS.get(self.role, [])
        return permission in permissions

    def get_permissions(self):
        # Admin gets all permissions
        if self.role == 'Admin':
            return list(set(p for perms in ROLE_PERMISSIONS.values() for p in perms))
        return ROLE_PERMISSIONS.get(self.role, [])


class Product(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    sku = db.Column(db.String(64), unique=True, nullable=False)
    name = db.Column(db.String(255), nullable=False)
    category = db.Column(db.String(100), nullable=True, index=True)
    eligible_for_distribution = db.Column(db.Boolean, default=True, nullable=False)
    risk_score = db.Column(db.Integer, nullable=True)
    risk_reason = db.Column(db.Text, nullable=True)
    lifecycle_status = db.Column(db.String(20), nullable=True, index=True)


class Store(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(255), unique=True, nullable=False)


class DistributionRecord(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    product_id = db.Column(db.Integer, db.ForeignKey(
        'product.id'), nullable=False)
    store_id = db.Column(db.Integer, db.ForeignKey('store.id'), nullable=False)
    quantity = db.Column(db.Integer, nullable=False)
    event_date = db.Column(db.Date, nullable=False)
    run_id = db.Column(db.String(36), nullable=True, index=True)

    product = db.relationship('Product')
    store = db.relationship('Store')


class SalesWeeklyAgg(db.Model):
    """Aggregated weekly sales for macro layer - full catalog visibility."""
    __tablename__ = 'sales_weekly_agg'
    id = db.Column(db.Integer, primary_key=True)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id'), nullable=False, index=True)
    store_id = db.Column(db.Integer, db.ForeignKey('store.id'), nullable=False, index=True)
    week_start = db.Column(db.Date, nullable=False, index=True)
    units = db.Column(db.Integer, nullable=False, default=0)
    category = db.Column(db.String(100), nullable=True, index=True)
    
    __table_args__ = (
        db.Index('ix_sales_weekly_agg_prod_store_week', 'product_id', 'store_id', 'week_start', unique=True),
        db.Index('ix_sales_weekly_agg_store_week', 'store_id', 'week_start'),
        db.Index('ix_sales_weekly_agg_cat_store_week', 'category', 'store_id', 'week_start'),
    )
    
    product = db.relationship('Product')
    store = db.relationship('Store')


class StockCD(db.Model):
    __tablename__ = 'stock_cd'
    id = db.Column(db.Integer, primary_key=True)
    as_of_date = db.Column(db.Date, nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id'), nullable=False)
    quantity = db.Column(db.Integer, nullable=False)

    product = db.relationship('Product')

class StockSnapshot(db.Model):
    __tablename__ = 'stock_snapshot'
    id = db.Column(db.Integer, primary_key=True)
    as_of_date = db.Column(db.Date, nullable=False)  # fecha del snapshot
    product_id = db.Column(db.Integer, db.ForeignKey(
        'product.id'), nullable=False)
    store_id = db.Column(db.Integer, db.ForeignKey('store.id'), nullable=False)
    quantity = db.Column(db.Integer, nullable=False)

    product = db.relationship('Product')
    store = db.relationship('Store')


RUN_WORKFLOW_STATUSES = ['DRAFT', 'PENDING_APPROVAL', 'APPROVED', 'REJECTED', 'CANCELED', 'completed']

class Run(db.Model):
    __tablename__ = 'run'
    id = db.Column(db.Integer, primary_key=True)
    run_id = db.Column(db.String(36), unique=True, nullable=False, index=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    run_type = db.Column(db.String(50), nullable=False, index=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    folio = db.Column(db.String(100), nullable=True)
    responsable = db.Column(db.String(100), nullable=True)
    categoria = db.Column(db.String(100), nullable=True)
    store_filter = db.Column(db.String(255), nullable=True)
    notes = db.Column(db.Text, nullable=True)
    status = db.Column(db.String(20), default='DRAFT', nullable=False, index=True)
    mode = db.Column(db.String(50), nullable=True)
    rows_count = db.Column(db.Integer, nullable=True)
    predictions_count = db.Column(db.Integer, nullable=True)
    is_active = db.Column(db.Boolean, default=False, nullable=False, index=True)
    
    urgency = db.Column(db.String(10), nullable=True)
    submit_note = db.Column(db.Text, nullable=True)
    submitted_by_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    submitted_at = db.Column(db.DateTime, nullable=True)
    approved_by_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    approved_at = db.Column(db.DateTime, nullable=True)
    rejected_by_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    rejected_at = db.Column(db.DateTime, nullable=True)
    reject_comment = db.Column(db.Text, nullable=True)
    planner_plan_id = db.Column(db.Integer, db.ForeignKey('distribution_plan.id'), nullable=True)

    user = db.relationship('User', foreign_keys=[user_id], backref='runs')
    submitted_by = db.relationship('User', foreign_keys=[submitted_by_user_id])
    approved_by = db.relationship('User', foreign_keys=[approved_by_user_id])
    rejected_by = db.relationship('User', foreign_keys=[rejected_by_user_id])
    planner_plan = db.relationship('DistributionPlan', foreign_keys=[planner_plan_id])

    def label(self):
        parts = []
        if self.folio:
            parts.append(f"Folio {self.folio}")
        if self.responsable:
            parts.append(f"Resp {self.responsable}")
        date_str = self.created_at.strftime('%Y-%m-%d %H:%M') if self.created_at else ''
        if parts:
            return f"{' | '.join(parts)} — {date_str}"
        return date_str or f"Run {self.run_id[:8]}"

PredictionRun = Run


class Prediction(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    product_id = db.Column(db.Integer, db.ForeignKey(
        'product.id'), nullable=False)
    store_id = db.Column(db.Integer, db.ForeignKey('store.id'), nullable=False)
    target_period_start = db.Column(db.Date, nullable=False)
    quantity = db.Column(db.Integer, nullable=False)
    model_name = db.Column(db.String(255), default='SMA_3w')
    run_id = db.Column(db.String(36), nullable=False, index=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    debug_json = db.Column(db.Text, nullable=True)

    product = db.relationship('Product')
    store = db.relationship('Store')
    
    def get_debug(self):
        if self.debug_json:
            try:
                return json.loads(self.debug_json)
            except:
                return {}
        return {}


class ForecastResult(db.Model):
    __tablename__ = 'forecast_result'
    id = db.Column(db.Integer, primary_key=True)
    run_id = db.Column(db.String(36), nullable=False, index=True)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id'), nullable=False)
    sku = db.Column(db.String(64), nullable=False)
    name = db.Column(db.String(255), nullable=True)
    demand_per_week = db.Column(db.Float, nullable=False)
    available_cd = db.Column(db.Integer, nullable=False)
    required_units = db.Column(db.Float, nullable=False)
    suggested = db.Column(db.Integer, nullable=False)
    campaign_tag = db.Column(db.String(100), nullable=True)
    buy_now = db.Column(db.Boolean, default=False, nullable=True)
    risk_level = db.Column(db.String(20), nullable=True)
    reason_code = db.Column(db.String(100), nullable=True)
    model_used = db.Column(db.String(50), nullable=True)
    lifecycle_status = db.Column(db.String(20), nullable=True)
    stock_store_total = db.Column(db.Integer, nullable=True)

    product = db.relationship('Product')


class ForecastBatchRun(db.Model):
    """Batch forecast run metadata."""
    __tablename__ = 'forecast_batch_run'
    id = db.Column(db.Integer, primary_key=True)
    batch_id = db.Column(db.String(36), unique=True, nullable=False, index=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    created_by_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    params_json = db.Column(db.Text, nullable=True)
    total_skus = db.Column(db.Integer, default=0)
    buy_now_count = db.Column(db.Integer, default=0)
    review_count = db.Column(db.Integer, default=0)
    do_not_buy_count = db.Column(db.Integer, default=0)

    user = db.relationship('User', backref='forecast_batch_runs')

    def get_params(self):
        if self.params_json:
            try:
                return json.loads(self.params_json)
            except Exception:
                return {}
        return {}


class ForecastBatchItem(db.Model):
    """Individual SKU result in a batch forecast."""
    __tablename__ = 'forecast_batch_item'
    id = db.Column(db.Integer, primary_key=True)
    batch_id = db.Column(db.String(36), nullable=False, index=True)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id'), nullable=False)
    sku = db.Column(db.String(64), nullable=False)
    product_name = db.Column(db.String(255), nullable=True)
    category = db.Column(db.String(100), nullable=True)
    lifecycle_status = db.Column(db.String(20), nullable=True)
    model_used = db.Column(db.String(50), nullable=True)
    weekly_demand = db.Column(db.Float, default=0)
    total_stock = db.Column(db.Integer, default=0)
    suggested_purchase = db.Column(db.Integer, default=0)
    decision = db.Column(db.String(20), nullable=True)
    risk_level = db.Column(db.String(20), nullable=True)
    reason_summary = db.Column(db.String(255), nullable=True)
    explain_json = db.Column(db.Text, nullable=True)

    product = db.relationship('Product')

    def get_explain(self):
        if self.explain_json:
            try:
                return json.loads(self.explain_json)
            except Exception:
                return []
        return []


class AuditLog(db.Model):
    """Audit trail for all significant actions in the system."""
    __tablename__ = 'audit_log'
    id = db.Column(db.Integer, primary_key=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False, index=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    username_snapshot = db.Column(db.String(80), nullable=True)
    role_snapshot = db.Column(db.String(50), nullable=True)
    action = db.Column(db.String(100), nullable=False, index=True)
    entity_type = db.Column(db.String(100), nullable=True)
    entity_id = db.Column(db.String(100), nullable=True)
    run_id = db.Column(db.String(36), nullable=True, index=True)
    status = db.Column(db.String(20), nullable=False, default='success')
    message = db.Column(db.String(500), nullable=True)
    metadata_json = db.Column(db.Text, nullable=True)
    ip_address = db.Column(db.String(45), nullable=True)
    user_agent = db.Column(db.String(500), nullable=True)

    user = db.relationship('User', backref='audit_logs')


class RebalanceRun(db.Model):
    """Run metadata for store-to-store rebalancing suggestions."""
    __tablename__ = 'rebalance_run'
    id = db.Column(db.Integer, primary_key=True)
    run_id = db.Column(db.String(36), unique=True, nullable=False, index=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    created_by_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    params_json = db.Column(db.Text, nullable=True)
    is_simulation = db.Column(db.Boolean, default=False, nullable=False)

    user = db.relationship('User', backref='rebalance_runs')

    def get_params(self):
        if self.params_json:
            try:
                return json.loads(self.params_json)
            except:
                return {}
        return {}


class RebalanceSuggestion(db.Model):
    """Individual rebalancing suggestion (transfer from one store to another)."""
    __tablename__ = 'rebalance_suggestion'
    id = db.Column(db.Integer, primary_key=True)
    run_id = db.Column(db.String(36), db.ForeignKey('rebalance_run.run_id'), nullable=False, index=True)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id'), nullable=False)
    from_store_id = db.Column(db.Integer, db.ForeignKey('store.id'), nullable=False)
    to_store_id = db.Column(db.Integer, db.ForeignKey('store.id'), nullable=False)
    qty = db.Column(db.Integer, nullable=False)
    sales_rate_to = db.Column(db.Float, nullable=True)
    woc_from = db.Column(db.Float, nullable=True)
    woc_to = db.Column(db.Float, nullable=True)
    score = db.Column(db.Float, nullable=True)
    reason = db.Column(db.String(255), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)

    product = db.relationship('Product')
    from_store = db.relationship('Store', foreign_keys=[from_store_id])
    to_store = db.relationship('Store', foreign_keys=[to_store_id])
    run = db.relationship('RebalanceRun', backref='suggestions')


class ManualTransferRun(db.Model):
    """Run metadata for manual transfer plans."""
    __tablename__ = 'manual_transfer_run'
    id = db.Column(db.Integer, primary_key=True)
    run_id = db.Column(db.String(36), unique=True, nullable=False, index=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    created_by_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    status = db.Column(db.String(20), nullable=False, default='draft')

    user = db.relationship('User', backref='manual_transfer_runs')


class ManualTransferItem(db.Model):
    """Individual item in a manual transfer plan."""
    __tablename__ = 'manual_transfer_item'
    id = db.Column(db.Integer, primary_key=True)
    run_id = db.Column(db.String(36), db.ForeignKey('manual_transfer_run.run_id'), nullable=False, index=True)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id'), nullable=False)
    from_store_id = db.Column(db.Integer, db.ForeignKey('store.id'), nullable=False)
    to_store_id = db.Column(db.Integer, db.ForeignKey('store.id'), nullable=False)
    qty = db.Column(db.Integer, nullable=False)
    status = db.Column(db.String(30), nullable=False, default='pending')
    notes = db.Column(db.String(255), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)

    product = db.relationship('Product')
    from_store = db.relationship('Store', foreign_keys=[from_store_id])
    to_store = db.relationship('Store', foreign_keys=[to_store_id])
    run = db.relationship('ManualTransferRun', backref='items')


class Job(db.Model):
    """Background job tracking for long-running operations."""
    __tablename__ = 'job'
    id = db.Column(db.String(36), primary_key=True)
    job_type = db.Column(db.String(50), nullable=False, index=True)
    status = db.Column(db.String(20), nullable=False, default='queued', index=True)
    progress = db.Column(db.Integer, nullable=False, default=0)
    message = db.Column(db.String(500), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    payload_json = db.Column(db.Text, nullable=True)
    result_json = db.Column(db.Text, nullable=True)

    user = db.relationship('User', backref='jobs')

    def set_payload(self, data):
        import json
        self.payload_json = json.dumps(data)

    def get_payload(self):
        import json
        if self.payload_json:
            try:
                return json.loads(self.payload_json)
            except:
                return {}
        return {}

    def set_result(self, data):
        import json
        self.result_json = json.dumps(data)

    def get_result(self):
        import json
        if self.result_json:
            try:
                return json.loads(self.result_json)
            except:
                return {}
        return {}


# ------------------ Slow Stock & Smart Reallocation Models ------------------

class SkuLifecycle(db.Model):
    """Global SKU lifecycle data: last purchase date, last sale date, and optional day counts from upload."""
    __tablename__ = 'sku_lifecycle'
    id = db.Column(db.Integer, primary_key=True)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id'), unique=True, nullable=False)
    last_purchase_date = db.Column(db.Date, nullable=True)
    last_sale_date_global = db.Column(db.Date, nullable=True)
    days_since_last_sale_global = db.Column(db.Integer, nullable=True)
    days_since_last_purchase_global = db.Column(db.Integer, nullable=True)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    product = db.relationship('Product', backref='lifecycle')


class SkuStoreLifecycle(db.Model):
    """Per SKU-store lifecycle data: last sale date at store level."""
    __tablename__ = 'sku_store_lifecycle'
    id = db.Column(db.Integer, primary_key=True)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id'), nullable=False)
    store_id = db.Column(db.Integer, db.ForeignKey('store.id'), nullable=False)
    last_sale_date_store = db.Column(db.Date, nullable=True)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    product = db.relationship('Product')
    store = db.relationship('Store')

    __table_args__ = (
        db.UniqueConstraint('product_id', 'store_id', name='uq_sku_store_lifecycle'),
    )


class SlowStockRun(db.Model):
    """Run metadata for slow stock analysis."""
    __tablename__ = 'slow_stock_run'
    id = db.Column(db.Integer, primary_key=True)
    run_id = db.Column(db.String(36), unique=True, nullable=False, index=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    created_by_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    params_json = db.Column(db.Text, nullable=True)
    dead_store_count = db.Column(db.Integer, default=0)
    slow_store_count = db.Column(db.Integer, default=0)
    transfer_count = db.Column(db.Integer, default=0)
    dead_cd_count = db.Column(db.Integer, default=0)

    user = db.relationship('User', backref='slow_stock_runs')

    def get_params(self):
        if self.params_json:
            try:
                return json.loads(self.params_json)
            except:
                return {}
        return {}


class SlowStockSuggestion(db.Model):
    """Individual slow stock transfer suggestion."""
    __tablename__ = 'slow_stock_suggestion'
    id = db.Column(db.Integer, primary_key=True)
    run_id = db.Column(db.String(36), db.ForeignKey('slow_stock_run.run_id'), nullable=False, index=True)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id'), nullable=False)
    from_store_id = db.Column(db.Integer, db.ForeignKey('store.id'), nullable=False)
    to_store_id = db.Column(db.Integer, db.ForeignKey('store.id'), nullable=False)
    qty = db.Column(db.Integer, nullable=False)
    donor_status = db.Column(db.String(20), nullable=True)  # DEAD_STORE, SLOW_STORE
    receiver_sales_rate = db.Column(db.Float, nullable=True)
    receiver_woc = db.Column(db.Float, nullable=True)
    reason = db.Column(db.String(255), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)

    product = db.relationship('Product')
    from_store = db.relationship('Store', foreign_keys=[from_store_id])
    to_store = db.relationship('Store', foreign_keys=[to_store_id])
    run = db.relationship('SlowStockRun', backref='suggestions')


# ------------------ FastPlanner Models ------------------
PLAN_STATUSES = ['APPROVED', 'IN_PROGRESS', 'PACKED', 'DISPATCHED', 'CLOSED', 'BLOCKED']
PLAN_URGENCIES = ['LOW', 'MEDIUM', 'URGENT']

class DistributionPlan(db.Model):
    """Distribution plan for warehouse execution."""
    __tablename__ = 'distribution_plan'
    id = db.Column(db.Integer, primary_key=True)
    folio = db.Column(db.String(50), unique=True, nullable=False, index=True)
    source_run_id = db.Column(db.String(36), nullable=True, index=True)
    status = db.Column(db.String(20), nullable=False, default='APPROVED', index=True)
    urgency = db.Column(db.String(10), nullable=False, default='MEDIUM')
    notes_commercial = db.Column(db.Text, nullable=True)
    notes_warehouse = db.Column(db.Text, nullable=True)
    created_by_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    approved_by_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    assigned_to_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    approved_at = db.Column(db.DateTime, nullable=True)
    started_at = db.Column(db.DateTime, nullable=True)
    closed_at = db.Column(db.DateTime, nullable=True)

    created_by = db.relationship('User', foreign_keys=[created_by_user_id])
    approved_by = db.relationship('User', foreign_keys=[approved_by_user_id])
    assigned_to = db.relationship('User', foreign_keys=[assigned_to_user_id])
    lines = db.relationship('DistributionPlanLine', backref='plan', lazy='dynamic', cascade='all, delete-orphan')
    activity_logs = db.relationship('PlanActivityLog', backref='plan', lazy='dynamic', cascade='all, delete-orphan', order_by='desc(PlanActivityLog.timestamp)')

    def total_skus(self):
        return db.session.query(db.func.count(db.func.distinct(DistributionPlanLine.product_id))).filter(DistributionPlanLine.plan_id == self.id).scalar() or 0

    def total_units(self):
        return db.session.query(db.func.sum(DistributionPlanLine.qty_planned)).filter(DistributionPlanLine.plan_id == self.id).scalar() or 0

    def total_stores(self):
        return db.session.query(db.func.count(db.func.distinct(DistributionPlanLine.store_id))).filter(DistributionPlanLine.plan_id == self.id).scalar() or 0


class DistributionPlanLine(db.Model):
    """Line item for a distribution plan."""
    __tablename__ = 'distribution_plan_line'
    id = db.Column(db.Integer, primary_key=True)
    plan_id = db.Column(db.Integer, db.ForeignKey('distribution_plan.id'), nullable=False, index=True)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id'), nullable=False)
    store_id = db.Column(db.Integer, db.ForeignKey('store.id'), nullable=False)
    qty_planned = db.Column(db.Integer, nullable=False, default=0)
    line_notes = db.Column(db.Text, nullable=True)

    product = db.relationship('Product')
    store = db.relationship('Store')


class PlanActivityLog(db.Model):
    """Activity log for distribution plan actions."""
    __tablename__ = 'plan_activity_log'
    id = db.Column(db.Integer, primary_key=True)
    plan_id = db.Column(db.Integer, db.ForeignKey('distribution_plan.id'), nullable=False, index=True)
    action = db.Column(db.String(50), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    comment = db.Column(db.Text, nullable=True)

    user = db.relationship('User')


# ------------------ Background Job Infrastructure ------------------
from concurrent.futures import ThreadPoolExecutor
from uuid import uuid4 as uuid4_gen
from sqlalchemy.orm import scoped_session, sessionmaker

job_executor = ThreadPoolExecutor(max_workers=4)

def get_isolated_session():
    """Create an isolated session for thread-safe database operations."""
    Session = scoped_session(sessionmaker(bind=db.engine))
    return Session()


def update_job_status(job_id, status=None, progress=None, message=None, result=None):
    """Update job status from within a background thread. Uses isolated session for thread safety."""
    session = None
    try:
        session = get_isolated_session()
        job = session.get(Job, job_id)
        if job:
            if status is not None:
                job.status = status
            if progress is not None:
                job.progress = progress
            if message is not None:
                job.message = message
            if result is not None:
                import json
                job.result_json = json.dumps(result, default=str, ensure_ascii=False)
            job.updated_at = datetime.utcnow()
            session.commit()
    except Exception as e:
        if session:
            session.rollback()
        app.logger.error(f"Failed to update job status {job_id}: {e}")
    finally:
        if session:
            session.close()


def create_and_run_job(job_type, task_func, payload=None, user_id=None):
    """Create a job record and run task_func in background thread.
    Returns job_id immediately so caller can redirect to status page.
    """
    job_id = str(uuid4_gen())
    job = Job(
        id=job_id,
        job_type=job_type,
        status='queued',
        progress=0,
        user_id=user_id,
        message='En cola...'
    )
    if payload:
        job.set_payload(payload)
    db.session.add(job)
    db.session.commit()
    db.session.expire_all()

    def wrapped_task():
        with app.app_context():
            task_session = None
            try:
                update_job_status(job_id, status='running', progress=5, message='Iniciando...')
                task_func(job_id, payload or {})
            except Exception as e:
                import traceback
                tb = traceback.format_exc()
                update_job_status(job_id, status='error', progress=100, message=f'Error: {str(e)[:400]}')
                app.logger.error(f"Job {job_id} failed: {e}\n{tb}")

    job_executor.submit(wrapped_task)
    return job_id


# ------------------ Login loader ------------------


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

from datetime import date  # seguramente ya lo tienes más arriba

@app.context_processor
def inject_globals():
    """
    Variables globales disponibles en todos los templates Jinja.
    """
    return {
        'date': date,
        'ROLES': ROLES,
        'ROLE_PERMISSIONS': ROLE_PERMISSIONS
    }


def require_permission(permission):
    """Decorator to require a specific permission to access a route."""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for('login'))
            if not current_user.has_permission(permission):
                flash(f'No tienes permiso para acceder a esta sección.', 'warning')
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator


# ------------------ Audit Trail Helper ------------------
import json

def log_audit(action, status="success", message="", entity_type=None, entity_id=None, run_id=None, metadata=None):
    """
    Log an audit trail entry. Safe to call - won't crash if logging fails.
    
    Args:
        action: Action identifier (e.g., "sales.upload", "distribution.run")
        status: "success" or "fail"
        message: Human-readable description
        entity_type: Type of entity affected (e.g., "DistributionRecord", "StockCD")
        entity_id: ID of specific entity (optional)
        run_id: UUID of related run (optional)
        metadata: Dict with additional context (will be JSON-serialized)
    """
    try:
        user_id = None
        username_snapshot = None
        role_snapshot = None
        ip_address = None
        user_agent = None
        
        if current_user and current_user.is_authenticated:
            user_id = current_user.id
            username_snapshot = current_user.username
            role_snapshot = current_user.role
        
        if request:
            ip_address = request.remote_addr
            user_agent = request.headers.get('User-Agent', '')[:500]
        
        metadata_json = None
        if metadata:
            try:
                metadata_json = json.dumps(metadata, default=str, ensure_ascii=False)
            except Exception:
                metadata_json = str(metadata)
        
        audit_entry = AuditLog(
            user_id=user_id,
            username_snapshot=username_snapshot,
            role_snapshot=role_snapshot,
            action=action,
            entity_type=entity_type,
            entity_id=str(entity_id) if entity_id else None,
            run_id=run_id,
            status=status,
            message=message[:500] if message else None,
            metadata_json=metadata_json,
            ip_address=ip_address,
            user_agent=user_agent
        )
        db.session.add(audit_entry)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Failed to write audit log: {e}")


# ------------------ Job Status Routes ------------------

@app.route('/jobs/<job_id>')
@login_required
def job_status_json(job_id):
    """Return job status as JSON for polling."""
    from flask import jsonify
    job = db.session.get(Job, job_id)
    if not job:
        return jsonify({'error': 'Job not found'}), 404
    return jsonify({
        'id': job.id,
        'job_type': job.job_type,
        'status': job.status,
        'progress': job.progress,
        'message': job.message,
        'created_at': job.created_at.isoformat() if job.created_at else None,
        'updated_at': job.updated_at.isoformat() if job.updated_at else None,
        'result': job.get_result() if job.status == 'done' else None
    })


@app.route('/jobs/<job_id>/view')
@login_required
def job_status_view(job_id):
    """Render job status page with auto-polling."""
    job = db.session.get(Job, job_id)
    if not job:
        flash('Trabajo no encontrado.', 'warning')
        return redirect(url_for('dashboard'))
    
    job_type_labels = {
        'upload_stock_cd': 'Carga de Stock CD',
        'upload_sales': 'Carga de Ventas',
        'redistribution': 'Redistribución entre Tiendas'
    }
    
    redirect_urls = {
        'upload_stock_cd': url_for('upload_stock_cd'),
        'upload_sales': url_for('dashboard'),
        'redistribution': url_for('rebalancing')
    }
    
    return render_template('job_status.html',
                           job=job,
                           job_label=job_type_labels.get(job.job_type, job.job_type),
                           redirect_url=redirect_urls.get(job.job_type, url_for('dashboard')))


# ------------------ Utilidades ------------------


def next_monday(ref: datetime | None = None):
    ref = ref or datetime.utcnow()
    days_ahead = (7 - ref.weekday()) % 7
    days_ahead = 7 if days_ahead == 0 else days_ahead
    return (ref + timedelta(days=days_ahead)).date()

def has_any_stock_loaded() -> bool:
    """True si existe al menos un snapshot de stock en la tabla."""
    return db.session.query(StockSnapshot.id).first() is not None


# ------------------ Simulation Mode Helpers ------------------
from flask import session

SIMULATION_EXPIRY_MINUTES = 30

def store_simulation_results(sim_type, results, kpis=None, cd_remaining=None, meta=None):
    """
    Store simulation results in Flask session.
    sim_type: 'distribution', 'forecast', 'rebalancing'
    """
    if 'simulations' not in session:
        session['simulations'] = {}
    
    session['simulations'][sim_type] = {
        'results': results,
        'kpis': kpis or {},
        'cd_remaining': cd_remaining or [],
        'meta': meta or {},
        'timestamp': datetime.utcnow().isoformat(),
        'user_id': current_user.id if current_user.is_authenticated else None
    }
    session.modified = True


def get_simulation_results(sim_type):
    """
    Retrieve simulation results from Flask session.
    Returns None if expired or not found.
    """
    if 'simulations' not in session:
        return None
    
    sim_data = session['simulations'].get(sim_type)
    if not sim_data:
        return None
    
    timestamp = datetime.fromisoformat(sim_data['timestamp'])
    if datetime.utcnow() - timestamp > timedelta(minutes=SIMULATION_EXPIRY_MINUTES):
        clear_simulation(sim_type)
        return None
    
    return sim_data


def clear_simulation(sim_type=None):
    """Clear simulation data from session."""
    if 'simulations' not in session:
        return
    
    if sim_type:
        session['simulations'].pop(sim_type, None)
    else:
        session['simulations'] = {}
    session.modified = True


def has_active_simulation(sim_type=None):
    """Check if there's an active simulation."""
    if sim_type:
        return get_simulation_results(sim_type) is not None
    return bool(session.get('simulations', {}))


# Predicción: promedio móvil 3 semanas por SKU-Tienda

from datetime import date
from collections import defaultdict

def generate_predictions(
    mode: str = "sma3_min3",
    meta: dict | None = None,
    df: pd.DataFrame | None = None,
    sales_run_id: str | None = None,
    simulate: bool = False
):
    from uuid import uuid4
    run_id = str(uuid4())
    
    meta = meta or {}
    
    if not simulate:
        prediction_run = Run(
            run_id=run_id,
            run_type='distribution',
            folio=meta.get("folio"),
            responsable=meta.get("responsable"),
            categoria=meta.get("categoria"),
            notes=meta.get("fecha_doc"),
            mode=mode,
            status='completed'
        )
        db.session.add(prediction_run)
        db.session.flush()

    # 1) Origen de datos: df pasado o histórico completo
    # Try SalesWeeklyAgg (macro layer) first when no df is passed
    use_macro_layer = False
    if df is None:
        macro_count = db.session.query(func.count(SalesWeeklyAgg.id)).scalar() or 0
        if macro_count > 0:
            use_macro_layer = True
            product_map = {p.id: p.sku for p in Product.query.all()}
            store_map = {s.id: s.name for s in Store.query.all()}
            
            agg_rows = SalesWeeklyAgg.query.all()
            data = [{
                'sku': product_map.get(r.product_id, ''),
                'store': store_map.get(r.store_id, ''),
                'quantity': r.units,
                'week_start': pd.to_datetime(r.week_start)
            } for r in agg_rows if product_map.get(r.product_id) and store_map.get(r.store_id)]
            if data:
                df = pd.DataFrame(data)
                df['date'] = df['week_start']
            else:
                use_macro_layer = False
        
        if not use_macro_layer:
            rows = DistributionRecord.query.all()
            if not rows:
                db.session.commit()
                return run_id, 0, 0, 0, 0
            data = [{
                'sku': r.product.sku,
                'store': r.store.name,
                'quantity': r.quantity,
                'date': pd.to_datetime(r.event_date)
            } for r in rows]
            df = pd.DataFrame(data)
    else:
        df = df.copy()
        df['date'] = pd.to_datetime(df['date'])

    # Normalizar columnas clave (alineado con /upload)
    df['sku'] = df['sku'].astype(str).str.strip()
    df['store'] = df['store'].astype(str).str.strip()
    df['quantity'] = pd.to_numeric(df['quantity'], errors='coerce').fillna(0).astype(int)

    # Normalizar semana (lunes como inicio)
    df['week_start'] = df['date'] - pd.to_timedelta(df['date'].dt.weekday, unit='D')

    # 2) Parámetros según modo seleccionado
    if mode == "sma3_min3":
        win = 3
        min_weeks = 3
        use_stock = True
        model_tag = "Promedio móvil 3 semanas (ajustado por stock)"
    elif mode == "sma2_min2":
        win = 2
        min_weeks = 2
        use_stock = True
        model_tag = "Promedio móvil 2 semanas (ajustado por stock)"
    elif mode == "sma1_no_min":
        win = 1
        min_weeks = 1   # mínimo 1 semana
        use_stock = True
        model_tag = "Última semana (mínimo 1 semana)"
    elif mode == "sma3_ignore_stock":
        win = 3
        min_weeks = 3
        use_stock = False
        model_tag = "Promedio móvil 3 semanas (sin ajuste de stock)"
    else:
        win = 3
        min_weeks = 3
        use_stock = True
        model_tag = "Promedio móvil 3 semanas (ajustado por stock)"

    # 2.b) Aplicar metadata al nombre del modelo
    if meta:
        extra_bits = []
        if meta.get("folio"):
            extra_bits.append(f"Folio: {meta['folio']}")
        if meta.get("responsable"):
            extra_bits.append(f"Resp: {meta['responsable']}")
        if meta.get("categoria"):
            extra_bits.append(f"Cat: {meta['categoria']}")
        if meta.get("fecha_doc"):
            extra_bits.append(f"Fecha doc: {meta['fecha_doc']}")
        if extra_bits:
            model_tag = f"{model_tag} — " + " | ".join(extra_bits)

    # 3) Define active snapshot dates ONCE (deterministic and consistent)
    active_store_stock_date = db.session.query(db.func.max(StockSnapshot.as_of_date)).scalar()
    active_cd_stock_date = db.session.query(db.func.max(StockCD.as_of_date)).scalar()
    
    # Pre-load all store stock for the active date into a lookup dict
    store_stock_lookup = {}
    if active_store_stock_date:
        stock_rows = (
            db.session.query(StockSnapshot.product_id, StockSnapshot.store_id, StockSnapshot.quantity)
            .filter(StockSnapshot.as_of_date == active_store_stock_date)
            .all()
        )
        for pid, sid, qty in stock_rows:
            store_stock_lookup[(pid, sid)] = qty or 0
    
    # Pre-load CD stock for the active date
    cd_stock_lookup = {}
    if active_cd_stock_date:
        cd_rows = StockCD.query.filter_by(as_of_date=active_cd_stock_date).all()
        for row in cd_rows:
            cd_stock_lookup[row.product_id] = row.quantity or 0
    
    # Pre-compute category sales rates per store (for BACKOFF_STOCKOUT tie-breaking)
    cat_store_sales_rate = {}
    product_category_map_legacy = {}
    all_products = Product.query.all()
    for p in all_products:
        product_category_map_legacy[p.id] = p.category
    categories_in_scope = set(c for c in product_category_map_legacy.values() if c)
    
    if categories_in_scope:
        from datetime import timedelta
        cutoff_date = date.today() - timedelta(weeks=12)
        cat_agg = (
            db.session.query(
                SalesWeeklyAgg.category,
                SalesWeeklyAgg.store_id,
                func.sum(SalesWeeklyAgg.units).label('total_units'),
                func.count(func.distinct(SalesWeeklyAgg.week_start)).label('weeks_count')
            )
            .filter(
                SalesWeeklyAgg.category.in_(categories_in_scope),
                SalesWeeklyAgg.week_start >= cutoff_date
            )
            .group_by(SalesWeeklyAgg.category, SalesWeeklyAgg.store_id)
            .all()
        )
        for cat, store_id, total_units, weeks_count in cat_agg:
            if weeks_count > 0:
                cat_store_sales_rate[(cat, store_id)] = total_units / max(weeks_count, 1)
    
    # 3b) Generar sugerencias base por SKU–Tienda
    raw_preds = []

    for (sku, store), gdf in df.groupby(['sku', 'store']):
        weekly = (
            gdf.groupby('week_start', as_index=False)['quantity']
             .sum()
             .sort_values('week_start')
        )

        # Exigir mínimo de semanas (blindado a >=1)
        effective_min_weeks = max(min_weeks, 1)
        if weekly.shape[0] < effective_min_weeks:
            continue

        # Get last N weeks for SMA calculation (primary window)
        last_weeks = weekly.tail(win)
        primary_units_sum = float(last_weeks['quantity'].sum())
        primary_weekly_rate = float(last_weeks['quantity'].mean())
        
        # Debug: capture individual week values (w0 = most recent, w1 = previous)
        week_values = last_weeks['quantity'].tolist()
        w0_units = week_values[-1] if len(week_values) >= 1 else 0
        w1_units = week_values[-2] if len(week_values) >= 2 else 0

        # Normalizar claves para buscar en BD
        sku_key = str(sku).strip()
        store_key = str(store).strip()

        # Buscar producto y tienda; si no existen, saltar
        product = Product.query.filter_by(sku=sku_key).first()
        if not product:
            continue

        store_ent = Store.query.filter_by(name=store_key).first()
        if not store_ent:
            continue

        # Ajuste por stock tienda using pre-loaded lookup (consistent date)
        if use_stock:
            stock_qty = store_stock_lookup.get((product.id, store_ent.id), 0)
        else:
            stock_qty = 0
        
        # Get CD stock for this product
        cd_qty = cd_stock_lookup.get(product.id, 0)

        # --- Stockout-aware backoff logic ---
        # If primary window has 0 sales, store is stocked out, but has proven demand in fallback window
        FALLBACK_WINDOW_WEEKS = 12
        fallback_weekly_rate = 0.0
        fallback_units_sum = 0.0
        weekly_rate_used = primary_weekly_rate
        backoff_applied = False
        
        if primary_units_sum == 0 and stock_qty == 0:
            # Compute fallback window rate (last 12 weeks)
            fallback_weeks = weekly.tail(FALLBACK_WINDOW_WEEKS)
            fallback_units_sum = float(fallback_weeks['quantity'].sum())
            if fallback_units_sum > 0:
                fallback_weekly_rate = fallback_units_sum / min(len(fallback_weeks), FALLBACK_WINDOW_WEEKS)
                weekly_rate_used = fallback_weekly_rate
                backoff_applied = True
        
        base_mean = weekly_rate_used

        suggested_before_caps = max(int(round(base_mean)) - stock_qty, 0)
        
        # Determine reason code
        if backoff_applied:
            reason_code = "BACKOFF_STOCKOUT"
        elif cd_qty == 0:
            reason_code = "CD_ZERO"
        elif stock_qty >= base_mean and suggested_before_caps == 0:
            reason_code = "STOCK_COVERS"
        elif base_mean == 0:
            reason_code = "NO_SALES_IN_WINDOW"
        elif suggested_before_caps == 0:
            reason_code = "STOCK_COVERS"
        else:
            reason_code = "OK"

        # Prioritization metrics for deterministic allocation order
        need_units = suggested_before_caps  # Primary: cover stockouts first
        sku_sales_rate_store = round(weekly_rate_used, 4)  # Secondary: higher SKU sellers first
        
        # Category rate for BACKOFF_STOCKOUT tie-breaking only
        product_category = product_category_map_legacy.get(product.id)
        cat_sales_rate_store = 0.0
        if product_category:
            cat_sales_rate_store = round(cat_store_sales_rate.get((product_category, store_ent.id), 0.0), 4)
        
        raw_preds.append({
            "sku": sku_key,
            "store": store_key,
            "product_id": product.id,
            "store_id": store_ent.id,
            "suggested": suggested_before_caps,
            "model_name": model_tag,
            "w0_units": int(w0_units),
            "w1_units": int(w1_units),
            "sma_mean": round(base_mean, 2),
            "stock_store_used": int(stock_qty),
            "cd_stock_used": int(cd_qty),
            "suggested_before_caps": suggested_before_caps,
            "reason_code": reason_code,
            "primary_units_sum": int(primary_units_sum),
            "weekly_rate_primary": round(primary_weekly_rate, 2),
            "weekly_rate_used": round(weekly_rate_used, 2),
            "fallback_units_sum": int(fallback_units_sum),
            "backoff_applied": backoff_applied,
            "need_units": need_units,
            "sku_sales_rate_store": sku_sales_rate_store,
            "cat_sales_rate_store": cat_sales_rate_store,
        })

    # 4) Aplicar límite por stock CD (priorizar tiendas con más demanda)
    # Reuse cd_stock_lookup and active_cd_stock_date from step 3
    cd_stock = cd_stock_lookup
    snapshot_date = active_cd_stock_date

    per_product = defaultdict(list)
    for r in raw_preds:
        per_product[r["product_id"]].append(r)

    final_preds = []

    # Deterministic prioritization sort key function
    def prioritization_sort_key(x):
        """
        Sorting policy (must be exact):
        1) need_units DESC (primary - cover stockouts first)
        2) sku_sales_rate_store DESC (secondary - higher SKU sellers first)
        3) cat_sales_rate_store DESC (tertiary, only when reason_code == BACKOFF_STOCKOUT)
        4) store_name ASC (final stable tie-break)
        """
        need = x.get("need_units", 0)
        sku_rate = x.get("sku_sales_rate_store", 0)
        # Category rate only used as tie-breaker when BACKOFF_STOCKOUT
        cat_rate = x.get("cat_sales_rate_store", 0) if x.get("reason_code") == "BACKOFF_STOCKOUT" else 0
        store_name = x.get("store", "")
        # Return tuple: negative for DESC, positive for ASC
        return (-need, -sku_rate, -cat_rate, store_name)

    for product_id, items in per_product.items():
        # Apply deterministic prioritization sort
        items.sort(key=prioritization_sort_key)
        
        # Assign allocation rank after sorting
        for rank, it in enumerate(items, start=1):
            it["alloc_rank"] = rank

        available = cd_stock.get(product_id, None)
        if available is None:
            # Sin stock CD cargado → dejar lo sugerido tal cual (unknown CD)
            for it in items:
                if it["reason_code"] == "OK" and it["suggested"] > 0:
                    it["reason_code"] = "CD_UNKNOWN"
                it["suggested_final"] = it["suggested"]
            final_preds.extend(items)
            continue
        
        # Track original CD for this product
        original_cd = available

        for idx, it in enumerate(items):
            want = it["suggested"]
            give = min(want, max(available, 0))
            it["suggested_final"] = give
            it["suggested"] = give
            available -= give
            
            # Update reason_code based on CD allocation
            if want > 0 and give == 0 and original_cd == 0:
                it["reason_code"] = "CD_ZERO"
            elif want > 0 and give == 0:
                it["reason_code"] = "CD_EXHAUSTED"
            elif want > 0 and give < want:
                it["reason_code"] = "CD_PARTIAL"
            
            final_preds.append(it)

            if available <= 0:
                # Lo que queda en la lista se queda en 0
                for it2 in items[idx+1:]:
                    it2["suggested_final"] = 0
                    it2["suggested"] = 0
                    if it2.get("suggested_before_caps", 0) > 0:
                        it2["reason_code"] = "CD_EXHAUSTED"
                    final_preds.append(it2)
                break

    # 4b) Stock-out Replenishment Layer (BREAK_REPLENISH)
    existing_predicted_keys = {
        (it["product_id"], it["store_id"])
        for it in final_preds
        if it["suggested"] > 0
    }
    
    base_assigned_by_product = defaultdict(int)
    for it in final_preds:
        base_assigned_by_product[it["product_id"]] += int(it["suggested"])
    
    cd_remaining_after_base = {}
    for product_id in cd_stock:
        original = cd_stock[product_id]
        assigned = base_assigned_by_product.get(product_id, 0)
        cd_remaining_after_base[product_id] = max(original - assigned, 0)
    
    today = date.today()
    recent_cutoff = today - timedelta(days=STOCKOUT_RECENT_WEEKS * 7)
    hist_cutoff = today - timedelta(days=STOCKOUT_HIST_WEEKS * 7)
    
    # Reuse active_store_stock_date from step 3
    latest_stock_date = active_store_stock_date
    
    stockout_candidates = []
    
    if latest_stock_date and snapshot_date:
        # Reuse store_stock_lookup from step 3
        store_stock_map = store_stock_lookup
        
        recent_sales_map = defaultdict(int)
        recent_q = (
            db.session.query(
                DistributionRecord.product_id,
                DistributionRecord.store_id,
                db.func.sum(DistributionRecord.quantity)
            )
            .filter(DistributionRecord.event_date >= recent_cutoff)
            .group_by(DistributionRecord.product_id, DistributionRecord.store_id)
            .all()
        )
        for pid, sid, qty in recent_q:
            recent_sales_map[(pid, sid)] = qty or 0
        
        week_expr = db.func.strftime('%Y-%W', DistributionRecord.event_date)
        hist_sales_q = (
            db.session.query(
                DistributionRecord.product_id,
                DistributionRecord.store_id,
                week_expr.label('week'),
                db.func.sum(DistributionRecord.quantity).label('weekly_qty')
            )
            .filter(DistributionRecord.event_date >= hist_cutoff)
            .filter(DistributionRecord.event_date < recent_cutoff)
            .group_by(DistributionRecord.product_id, DistributionRecord.store_id, week_expr)
            .all()
        )
        
        hist_weekly = defaultdict(list)
        for pid, sid, week, qty in hist_sales_q:
            hist_weekly[(pid, sid)].append(qty or 0)
        
        all_sku_store_pairs = set(store_stock_map.keys()) | set(hist_weekly.keys())
        
        product_lookup = {p.id: (p.sku, p.name) for p in Product.query.all()}
        store_lookup = {s.id: s.name for s in Store.query.all()}
        
        for (pid, sid) in all_sku_store_pairs:
            if (pid, sid) in existing_predicted_keys:
                continue
            
            store_stock = store_stock_map.get((pid, sid), 0)
            if store_stock > 0:
                continue
            
            recent_sales = recent_sales_map.get((pid, sid), 0)
            if recent_sales > 0:
                continue
            
            hist_weeks_list = hist_weekly.get((pid, sid), [])
            hist_total = sum(hist_weeks_list)
            if hist_total <= 0:
                continue
            
            cd_avail_for_replen = cd_remaining_after_base.get(pid, 0)
            if cd_avail_for_replen <= 0:
                continue
            
            hist_avg = hist_total / max(len(hist_weeks_list), 1)
            replen_qty = max(1, min(int(round(hist_avg * STOCKOUT_TARGET_WOC)), STOCKOUT_MAX_QTY))
            
            sku_info = product_lookup.get(pid, ('', ''))
            store_name = store_lookup.get(sid, '')
            
            stockout_candidates.append({
                "sku": sku_info[0],
                "store": store_name,
                "product_id": pid,
                "store_id": sid,
                "suggested": replen_qty,
                "hist_avg": hist_avg,
                "model_name": model_tag + " | BREAK_REPLENISH",
            })
        
        stockout_candidates.sort(key=lambda x: x["hist_avg"], reverse=True)
        
        replenish_per_product = defaultdict(list)
        for cand in stockout_candidates:
            replenish_per_product[cand["product_id"]].append(cand)
        
        replenish_added = 0
        cd_constrained = 0
        
        for product_id, cands in replenish_per_product.items():
            remaining_cd = cd_remaining_after_base.get(product_id, 0)
            
            if remaining_cd <= 0:
                for c in cands:
                    c["suggested"] = 0
                    final_preds.append(c)
                cd_constrained += len(cands)
                continue
            
            for c in cands:
                want = c["suggested"]
                give = min(want, remaining_cd)
                c["suggested"] = give
                remaining_cd -= give
                final_preds.append(c)
                
                if give > 0:
                    replenish_added += 1
                if give < want:
                    cd_constrained += 1
            
            cd_remaining_after_base[product_id] = remaining_cd
        
        if STOCKOUT_DEBUG:
            print(f"[STOCKOUT] Candidates found: {len(stockout_candidates)}")
            print(f"[STOCKOUT] Replenishment predictions added: {replenish_added}")
            print(f"[STOCKOUT] CD constrained count: {cd_constrained}")

    # 4c) Category-based Cold Start for New SKUs
    cold_start_added = 0
    cold_start_cd_allocated = 0
    
    COLD_START_MIN_FILL = 2
    COLD_START_TARGET_WOC = 1.0
    COLD_START_TOP_STORES = 20
    COLD_START_WINDOW_WEEKS = 12
    
    macro_data_available = db.session.query(func.count(SalesWeeklyAgg.id)).scalar() or 0
    
    if macro_data_available > 0:
        skus_in_df = set(df['sku'].unique()) if df is not None and 'sku' in df.columns else set()
        existing_pred_skus = {it['sku'] for it in final_preds if it.get('suggested', 0) > 0}
        
        products_with_category = (
            Product.query
            .filter(Product.category.isnot(None))
            .filter(Product.category != '')
            .all()
        )
        
        cold_start_window = date.today() - timedelta(weeks=COLD_START_WINDOW_WEEKS)
        
        category_store_rates = defaultdict(lambda: defaultdict(float))
        cat_agg_query = (
            db.session.query(
                SalesWeeklyAgg.category,
                SalesWeeklyAgg.store_id,
                func.sum(SalesWeeklyAgg.units).label('total_units'),
                func.count(func.distinct(SalesWeeklyAgg.week_start)).label('weeks')
            )
            .filter(SalesWeeklyAgg.week_start >= cold_start_window)
            .filter(SalesWeeklyAgg.category.isnot(None))
            .group_by(SalesWeeklyAgg.category, SalesWeeklyAgg.store_id)
            .all()
        )
        
        for cat, store_id, total_units, weeks in cat_agg_query:
            if weeks > 0:
                category_store_rates[cat][store_id] = (total_units or 0) / weeks
        
        store_lookup = {s.id: s.name for s in Store.query.all()}
        store_id_by_name = {s.name: s.id for s in Store.query.all()}
        
        # Reuse store_stock_lookup from step 3 for cold start
        cold_start_stock_map = store_stock_lookup
        
        cold_start_candidates = []
        
        for prod in products_with_category:
            if prod.sku in existing_pred_skus:
                continue
            
            if prod.sku in skus_in_df:
                continue
            
            category = prod.category
            if category not in category_store_rates:
                continue
            
            store_rates = category_store_rates[category]
            if not store_rates:
                continue
            
            sorted_stores = sorted(store_rates.items(), key=lambda x: x[1], reverse=True)
            eligible_stores = sorted_stores[:COLD_START_TOP_STORES]
            
            for store_id, cat_sales_rate in eligible_stores:
                if cat_sales_rate <= 0:
                    continue
                
                current_stock = cold_start_stock_map.get((prod.id, store_id), 0)
                
                need_units = max(int(math.ceil(COLD_START_TARGET_WOC * cat_sales_rate)), COLD_START_MIN_FILL)
                suggested = max(need_units - current_stock, 0)
                
                if suggested > 0:
                    cold_start_candidates.append({
                        "sku": prod.sku,
                        "store": store_lookup.get(store_id, ''),
                        "product_id": prod.id,
                        "store_id": store_id,
                        "suggested": suggested,
                        "model_name": f"COLD_START_CATEGORY ({category})",
                        "category": category,
                        "cat_sales_rate": cat_sales_rate
                    })
        
        cold_start_candidates.sort(key=lambda x: x['cat_sales_rate'], reverse=True)
        
        cold_start_per_product = defaultdict(list)
        for cand in cold_start_candidates:
            cold_start_per_product[cand["product_id"]].append(cand)
        
        for product_id, cands in cold_start_per_product.items():
            remaining_cd = cd_remaining_after_base.get(product_id, 0)
            
            if remaining_cd <= 0:
                continue
            
            for c in cands:
                want = c["suggested"]
                give = min(want, remaining_cd)
                if give > 0:
                    c["suggested"] = give
                    remaining_cd -= give
                    final_preds.append(c)
                    cold_start_added += 1
                    cold_start_cd_allocated += give
                
                if remaining_cd <= 0:
                    break
            
            cd_remaining_after_base[product_id] = remaining_cd

    # 5) Guardar predicciones y registrar asignaciones
    target_week = next_monday()
    preds_bulk = []
    assigned_by_product = defaultdict(int)

    for it in final_preds:
        assigned_qty = int(it["suggested"])
        assigned_by_product[it["product_id"]] += assigned_qty

    if simulate:
        sim_results = []
        product_cache = {p.id: p for p in Product.query.all()}
        store_cache = {s.id: s for s in Store.query.all()}
        
        for it in final_preds:
            prod = product_cache.get(it["product_id"])
            store_obj = store_cache.get(it["store_id"])
            sim_results.append({
                'sku': it['sku'],
                'product_name': prod.name if prod else '',
                'store': it['store'],
                'quantity': int(it['suggested']),
                'model_name': it.get('model_name', ''),
                'target_week': target_week.isoformat()
            })
        
        cd_remaining = []
        if snapshot_date:
            for cd_row in cd_stock_rows:
                assigned = assigned_by_product.get(cd_row.product_id, 0)
                remaining = max(cd_row.quantity - assigned, 0)
                prod = product_cache.get(cd_row.product_id)
                cd_remaining.append({
                    'sku': prod.sku if prod else '',
                    'product_name': prod.name if prod else '',
                    'original': cd_row.quantity,
                    'assigned': assigned,
                    'remaining': remaining
                })
        
        kpis = {
            'total_units': sum(r['quantity'] for r in sim_results),
            'num_skus': len(set(r['sku'] for r in sim_results)),
            'num_stores': len(set(r['store'] for r in sim_results)),
            'num_predictions': len(sim_results)
        }
        
        return run_id, len(final_preds), sim_results, cd_remaining, kpis

    for it in final_preds:
        assigned_qty = int(it["suggested"])
        
        debug_data = {
            'w0_units': it.get('w0_units'),
            'w1_units': it.get('w1_units'),
            'sma_mean': it.get('sma_mean'),
            'stock_store_used': it.get('stock_store_used'),
            'cd_stock_used': it.get('cd_stock_used'),
            'suggested_before_caps': it.get('suggested_before_caps'),
            'suggested_final': it.get('suggested_final', assigned_qty),
            'reason_code': it.get('reason_code', 'OK'),
        }
        debug_json_str = json.dumps(debug_data)

        existing = Prediction.query.filter_by(
            product_id=it["product_id"],
            store_id=it["store_id"],
            target_period_start=target_week,
            run_id=run_id
        ).first()

        if existing:
            existing.quantity = assigned_qty
            existing.model_name = it["model_name"]
            existing.debug_json = debug_json_str
        else:
            preds_bulk.append(Prediction(
                product_id=it["product_id"],
                store_id=it["store_id"],
                target_period_start=target_week,
                quantity=assigned_qty,
                model_name=it["model_name"],
                run_id=run_id,
                debug_json=debug_json_str
            ))

    if preds_bulk:
        db.session.add_all(preds_bulk)

    # 6) Descontar del stock CD lo que realmente asignamos
    if snapshot_date:
        for product_id, assigned_total in assigned_by_product.items():
            cd_row = StockCD.query.filter_by(as_of_date=snapshot_date, product_id=product_id).first()
            if cd_row:
                cd_row.quantity = max(cd_row.quantity - assigned_total, 0)

    db.session.commit()

    try:
        g.latest_run_id = run_id
    except Exception:
        pass

    return run_id, len(final_preds), None, None, None


def generate_predictions_from_macro(
    scope_skus: list,
    mode: str = "sma3_min3",
    meta: dict | None = None,
    user_id: int | None = None
):
    """
    Generate distribution predictions from SalesWeeklyAgg (macro layer)
    constrained to the specified scope_skus only.
    
    CRITICAL: This function NEVER uses macro sales as the SKU universe.
    It ONLY generates predictions for SKUs in scope_skus.
    """
    from uuid import uuid4
    
    if not scope_skus:
        raise ValueError("scope_skus cannot be empty - no distribution will be generated")
    
    run_id = str(uuid4())
    meta = meta or {}
    
    prediction_run = Run(
        run_id=run_id,
        run_type='distribution',
        user_id=user_id,
        folio=meta.get("folio"),
        responsable=meta.get("responsable"),
        categoria=meta.get("categoria"),
        notes=meta.get("fecha_doc"),
        mode=mode,
        status='completed'
    )
    db.session.add(prediction_run)
    db.session.flush()
    
    # weeks_target: horizon for target units calculation
    # SMA1 -> 1 week, SMA2 -> 2 weeks, SMA3 -> 2 weeks (conservative default)
    if mode == "sma3_min3":
        win = 3
        min_weeks = 3
        weeks_target = 2
        use_stock = True
        model_tag = "Promedio móvil 3 semanas (macro)"
    elif mode == "sma2_min2":
        win = 2
        min_weeks = 2
        weeks_target = 2
        use_stock = True
        model_tag = "Promedio móvil 2 semanas (macro)"
    elif mode == "sma1_no_min":
        win = 1
        min_weeks = 1
        weeks_target = 1
        use_stock = True
        model_tag = "Última semana (macro)"
    elif mode == "sma3_ignore_stock":
        win = 3
        min_weeks = 3
        weeks_target = 2
        use_stock = False
        model_tag = "Promedio móvil 3 semanas sin ajuste (macro)"
    else:
        win = 3
        min_weeks = 3
        weeks_target = 2
        use_stock = True
        model_tag = "Promedio móvil 3 semanas (macro)"
    
    if meta:
        extra_bits = []
        if meta.get("folio"):
            extra_bits.append(f"Folio: {meta['folio']}")
        if meta.get("responsable"):
            extra_bits.append(f"Resp: {meta['responsable']}")
        if meta.get("categoria"):
            extra_bits.append(f"Cat: {meta['categoria']}")
        if meta.get("fecha_doc"):
            extra_bits.append(f"Fecha doc: {meta['fecha_doc']}")
        if extra_bits:
            model_tag = f"{model_tag} — " + " | ".join(extra_bits)
    
    scope_skus_set = set(str(s).strip() for s in scope_skus)
    
    if not scope_skus_set:
        db.session.rollback()
        raise ValueError("ABORT: Distribution request file contains no SKUs. Scope cannot be empty.")
    
    MAX_WOC_CAP = 2.0
    MAX_UNITS_PER_STORE_EVENT = 10
    TOTAL_UNITS_KILL_SWITCH = 50000
    MIN_TOP10_ALLOC = 1
    TOP10_RANK_CUTOFF = 10
    
    # Targeted debug logging - only for specific SKU-store if env vars set
    DEBUG_SKU = os.environ.get('DEBUG_SKU', '').strip()
    DEBUG_STORE = os.environ.get('DEBUG_STORE', '').strip()
    debug_target_active = bool(DEBUG_SKU and DEBUG_STORE)
    if debug_target_active:
        print(f"[DISTRIBUTION DEBUG] Targeted debug active for SKU={DEBUG_SKU}, Store={DEBUG_STORE}")
    
    product_map = {}
    for sku in scope_skus_set:
        prod = Product.query.filter_by(sku=sku).first()
        if prod:
            product_map[sku] = prod
    
    if not product_map:
        db.session.commit()
        return run_id, 0, 0
    
    product_ids = [p.id for p in product_map.values()]
    store_map = {s.id: s.name for s in Store.query.all()}
    store_id_map = {s.name: s.id for s in Store.query.all()}
    
    agg_rows = (
        SalesWeeklyAgg.query
        .filter(SalesWeeklyAgg.product_id.in_(product_ids))
        .all()
    )
    
    sku_by_pid = {p.id: p.sku for p in product_map.values()}
    
    data = []
    for r in agg_rows:
        sku_val = sku_by_pid.get(r.product_id)
        store_name = store_map.get(r.store_id)
        if sku_val and store_name:
            data.append({
                'sku': sku_val,
                'store': store_name,
                'product_id': r.product_id,
                'store_id': r.store_id,
                'quantity': r.units,
                'week_start': r.week_start
            })
    
    raw_preds = []
    skus_with_sales_estimation = set()
    
    # Pre-compute category sales rates per store (for BACKOFF_STOCKOUT tie-breaking)
    # Map: (category, store_id) -> avg_weekly_rate
    cat_store_sales_rate = {}
    product_category_map = {p.id: p.category for p in product_map.values()}
    categories_in_scope = set(c for c in product_category_map.values() if c)
    
    if categories_in_scope:
        from datetime import timedelta
        cutoff_date = date.today() - timedelta(weeks=12)
        cat_agg = (
            db.session.query(
                SalesWeeklyAgg.category,
                SalesWeeklyAgg.store_id,
                func.sum(SalesWeeklyAgg.units).label('total_units'),
                func.count(func.distinct(SalesWeeklyAgg.week_start)).label('weeks_count')
            )
            .filter(
                SalesWeeklyAgg.category.in_(categories_in_scope),
                SalesWeeklyAgg.week_start >= cutoff_date
            )
            .group_by(SalesWeeklyAgg.category, SalesWeeklyAgg.store_id)
            .all()
        )
        for cat, store_id, total_units, weeks_count in cat_agg:
            if weeks_count > 0:
                cat_store_sales_rate[(cat, store_id)] = total_units / max(weeks_count, 1)
    
    if data:
        df = pd.DataFrame(data)
        df['week_start'] = pd.to_datetime(df['week_start'])
        
        for (sku, store), gdf in df.groupby(['sku', 'store']):
            weekly = (
                gdf.groupby('week_start', as_index=False)['quantity']
                .sum()
                .sort_values('week_start')
            )
            
            effective_min_weeks = max(min_weeks, 1)
            if weekly.shape[0] < effective_min_weeks:
                continue
            
            # Get last N weeks for SMA calculation (primary window)
            last_weeks = weekly.tail(win)
            primary_units_sum = float(last_weeks['quantity'].sum())
            primary_weekly_rate = float(last_weeks['quantity'].mean())
            
            sku_key = str(sku).strip()
            store_key = str(store).strip()
            
            product = product_map.get(sku_key)
            if not product:
                continue
            
            store_id = store_id_map.get(store_key)
            if not store_id:
                continue
            
            if use_stock:
                latest_stock = (
                    db.session.query(StockSnapshot)
                    .filter(
                        StockSnapshot.product_id == product.id,
                        StockSnapshot.store_id == store_id
                    )
                    .order_by(StockSnapshot.as_of_date.desc())
                    .first()
                )
                stock_qty = latest_stock.quantity if latest_stock else 0
            else:
                stock_qty = 0
            
            # --- Stockout-aware backoff logic ---
            # If primary window has 0 sales, store is stocked out, but has proven demand in fallback window
            FALLBACK_WINDOW_WEEKS = 12
            fallback_weekly_rate = 0.0
            fallback_units_sum = 0.0
            weekly_rate_used = primary_weekly_rate
            backoff_applied = False
            
            if primary_units_sum == 0 and stock_qty == 0:
                # Compute fallback window rate (last 12 weeks)
                fallback_weeks = weekly.tail(FALLBACK_WINDOW_WEEKS)
                fallback_units_sum = float(fallback_weeks['quantity'].sum())
                if fallback_units_sum > 0:
                    fallback_weekly_rate = fallback_units_sum / min(len(fallback_weeks), FALLBACK_WINDOW_WEEKS)
                    weekly_rate_used = fallback_weekly_rate
                    backoff_applied = True
            
            weekly_rate = weekly_rate_used
            
            import math
            # Horizon-based target: target_units = ceil(weekly_rate * weeks_target)
            target_units = math.ceil(weekly_rate * weeks_target) if weekly_rate > 0 else 0
            suggested_raw = max(target_units - stock_qty, 0)
            max_alloc_woc = math.ceil(MAX_WOC_CAP * weekly_rate)
            suggested = min(suggested_raw, max_alloc_woc, MAX_UNITS_PER_STORE_EVENT)
            
            # Determine reason code for caps and zero cases
            reason_code = "OK"
            if backoff_applied:
                reason_code = "BACKOFF_STOCKOUT"
            elif suggested_raw == 0 and weekly_rate > 0:
                # STOCK_COVERS: stock already covers the horizon target
                reason_code = "STOCK_COVERS"
            elif weekly_rate == 0:
                reason_code = "NO_SALES_IN_WINDOW"
            elif suggested < suggested_raw:
                if suggested == MAX_UNITS_PER_STORE_EVENT and suggested_raw > MAX_UNITS_PER_STORE_EVENT:
                    reason_code = "CAPPED_EVENT"
                elif suggested == max_alloc_woc and suggested_raw > max_alloc_woc:
                    reason_code = "CAPPED_WOC"
            
            # Get week values for debug (last N weeks)
            week_values = last_weeks['quantity'].tolist()
            w0_units = int(week_values[-1]) if len(week_values) >= 1 else 0
            w1_units = int(week_values[-2]) if len(week_values) >= 2 else 0
            
            # Compute total sales in window for ranking
            total_sales_in_window = float(weekly.tail(win)['quantity'].sum())
            
            # Prioritization metrics for deterministic allocation order
            need_units = suggested_raw  # Primary: cover stockouts first
            sku_sales_rate_store = round(weekly_rate, 4)  # Secondary: higher SKU sellers first
            
            # Category rate for BACKOFF_STOCKOUT tie-breaking only
            product_category = product_category_map.get(product.id)
            cat_sales_rate_store = 0.0
            if product_category:
                cat_sales_rate_store = round(cat_store_sales_rate.get((product_category, store_id), 0.0), 4)
            
            raw_preds.append({
                "sku": sku_key,
                "store": store_key,
                "product_id": product.id,
                "store_id": store_id,
                "suggested": suggested,
                "suggested_before_caps": suggested_raw,
                "weekly_rate": weekly_rate,
                "target_units": target_units,
                "weeks_target": weeks_target,
                "model_name": model_tag,
                "w0_units": w0_units,
                "w1_units": w1_units,
                "sma_mean": round(weekly_rate, 2),
                "stock_store_used": int(stock_qty),
                "reason_code": reason_code,
                "total_sales_in_window": total_sales_in_window,
                "primary_units_sum": int(primary_units_sum),
                "weekly_rate_primary": round(primary_weekly_rate, 2),
                "weekly_rate_used": round(weekly_rate_used, 2),
                "fallback_units_sum": int(fallback_units_sum),
                "backoff_applied": backoff_applied,
                "need_units": need_units,
                "sku_sales_rate_store": sku_sales_rate_store,
                "cat_sales_rate_store": cat_sales_rate_store,
            })
            skus_with_sales_estimation.add(sku_key)
    
    # --- Top-10 Minimum Guaranteed Allocation ---
    # Compute rank per SKU across stores by total sales in window
    from collections import defaultdict
    preds_by_product = defaultdict(list)
    for pred in raw_preds:
        preds_by_product[pred["product_id"]].append(pred)
    
    for product_id, preds_list in preds_by_product.items():
        # Sort by total_sales_in_window descending to compute rank
        sorted_by_sales = sorted(preds_list, key=lambda x: x.get("total_sales_in_window", 0), reverse=True)
        for rank, pred in enumerate(sorted_by_sales, start=1):
            pred["store_rank"] = rank
            
            # Apply top-10 minimum guarantee rule
            if (rank <= TOP10_RANK_CUTOFF and
                pred.get("suggested", 0) == 0 and
                pred.get("reason_code") == "STOCK_COVERS" and
                pred.get("w0_units", 0) + pred.get("w1_units", 0) > 0 and  # recent sales
                pred.get("stock_store_used", 0) < pred.get("sma_mean", 0)):  # low stock
                
                # Apply minimum allocation (will be validated against CD later)
                pred["suggested"] = MIN_TOP10_ALLOC
                pred["reason_code"] = "TOP10_MIN_GUARANTEE"
    
    snapshot_date = db.session.query(db.func.max(StockCD.as_of_date)).scalar()
    
    if snapshot_date:
        cd_stock_rows = StockCD.query.filter_by(as_of_date=snapshot_date).filter(
            StockCD.product_id.in_(product_ids)
        ).all()
        cd_stock = {row.product_id: row.quantity for row in cd_stock_rows}
    else:
        cd_stock = {}
    
    per_product = defaultdict(list)
    for r in raw_preds:
        r["cd_stock_used"] = cd_stock.get(r["product_id"], 0)
        per_product[r["product_id"]].append(r)
    
    final_preds = []
    
    # Deterministic prioritization sort key function
    def prioritization_sort_key(x):
        """
        Sorting policy (must be exact):
        1) need_units DESC (primary - cover stockouts first)
        2) sku_sales_rate_store DESC (secondary - higher SKU sellers first)
        3) cat_sales_rate_store DESC (tertiary, only when reason_code == BACKOFF_STOCKOUT)
        4) store_name ASC (final stable tie-break)
        """
        need = x.get("need_units", 0)
        sku_rate = x.get("sku_sales_rate_store", 0)
        # Category rate only used as tie-breaker when BACKOFF_STOCKOUT
        cat_rate = x.get("cat_sales_rate_store", 0) if x.get("reason_code") == "BACKOFF_STOCKOUT" else 0
        store_name = x.get("store", "")
        # Return tuple: negative for DESC, positive for ASC
        return (-need, -sku_rate, -cat_rate, store_name)
    
    for product_id, items in per_product.items():
        # Apply deterministic prioritization sort
        items.sort(key=prioritization_sort_key)
        
        # Assign allocation rank after sorting
        for rank, it in enumerate(items, start=1):
            it["alloc_rank"] = rank
        
        available = cd_stock.get(product_id, None)
        original_cd = available if available is not None else 0
        
        if available is None:
            for it in items:
                it["suggested_final"] = it["suggested"]
                if it["reason_code"] == "OK" and it["suggested"] > 0:
                    it["reason_code"] = "CD_UNKNOWN"
            final_preds.extend(items)
            continue
        
        for idx, it in enumerate(items):
            want = it["suggested"]
            give = min(want, max(available, 0))
            it["suggested_final"] = give
            it["suggested"] = give
            available -= give
            
            # Update reason_code based on CD allocation
            if want > 0 and give == 0 and original_cd == 0:
                it["reason_code"] = "CD_ZERO"
            elif want > 0 and give == 0:
                it["reason_code"] = "CD_EXHAUSTED"
            elif want > 0 and give < want:
                it["reason_code"] = "CD_PARTIAL"
            
            final_preds.append(it)
            
            if available <= 0:
                for it2 in items[idx+1:]:
                    it2["suggested_final"] = 0
                    it2["suggested"] = 0
                    if it2.get("suggested_before_caps", 0) > 0:
                        it2["reason_code"] = "CD_EXHAUSTED"
                    final_preds.append(it2)
                break
    
    # --- Targeted per-row debug logging ---
    if debug_target_active:
        for it in final_preds:
            if it.get("sku") == DEBUG_SKU and it.get("store") == DEBUG_STORE:
                print(f"[DISTRIBUTION ROW DEBUG] sku={it.get('sku')} store={it.get('store')} "
                      f"store_id={it.get('store_id')} product_id={it.get('product_id')} "
                      f"w0_units={it.get('w0_units')} w1_units={it.get('w1_units')} "
                      f"sma_mean={it.get('sma_mean')} weeks_target={it.get('weeks_target')} target_units={it.get('target_units')} "
                      f"stock_store_used={it.get('stock_store_used')} cd_stock_used={it.get('cd_stock_used')} "
                      f"suggested_before_caps={it.get('suggested_before_caps')} suggested_final={it.get('suggested_final')} "
                      f"store_rank={it.get('store_rank')} reason={it.get('reason_code')}")
                break
        else:
            print(f"[DISTRIBUTION ROW DEBUG] No matching row found for sku={DEBUG_SKU} store={DEBUG_STORE}")
    
    skus_with_predictions = {it['sku'] for it in final_preds if it.get('suggested', 0) > 0}
    skus_without_macro_sales = scope_skus_set - skus_with_sales_estimation
    
    latest_stock_date_for_sku_check = db.session.query(db.func.max(StockSnapshot.as_of_date)).scalar()
    skus_with_any_store_stock = set()
    if latest_stock_date_for_sku_check:
        store_stock_by_sku = (
            db.session.query(StockSnapshot.product_id, func.sum(StockSnapshot.quantity))
            .filter(StockSnapshot.as_of_date == latest_stock_date_for_sku_check)
            .filter(StockSnapshot.product_id.in_(product_ids))
            .group_by(StockSnapshot.product_id)
            .all()
        )
        pid_to_sku = {p.id: sku for sku, p in product_map.items()}
        for pid, total_stock in store_stock_by_sku:
            if total_stock and total_stock > 0:
                sku_val = pid_to_sku.get(pid)
                if sku_val:
                    skus_with_any_store_stock.add(sku_val)
    
    skus_eligible_for_cold_start = skus_without_macro_sales - skus_with_any_store_stock
    
    COLD_START_MIN_FILL = 2
    COLD_START_TOP_STORES = 20
    COLD_START_WINDOW_WEEKS = 12
    MAX_CAT_WOC = 6.0
    
    cold_start_window = date.today() - timedelta(weeks=COLD_START_WINDOW_WEEKS)
    
    cold_start_candidates = []
    cold_start_no_category = []
    
    if skus_eligible_for_cold_start:
        import math
        
        category_store_rates = defaultdict(lambda: defaultdict(float))
        cat_agg_query = (
            db.session.query(
                SalesWeeklyAgg.category,
                SalesWeeklyAgg.store_id,
                func.sum(SalesWeeklyAgg.units).label('total_units'),
                func.count(func.distinct(SalesWeeklyAgg.week_start)).label('weeks')
            )
            .filter(SalesWeeklyAgg.week_start >= cold_start_window)
            .filter(SalesWeeklyAgg.category.isnot(None))
            .group_by(SalesWeeklyAgg.category, SalesWeeklyAgg.store_id)
            .all()
        )
        
        for cat, store_id, total_units, weeks in cat_agg_query:
            if weeks > 0:
                category_store_rates[cat][store_id] = (total_units or 0) / weeks
        
        latest_stock_date = db.session.query(db.func.max(StockSnapshot.as_of_date)).scalar()
        store_stock_map = {}
        if latest_stock_date:
            stock_snap_q = (
                db.session.query(StockSnapshot.product_id, StockSnapshot.store_id, StockSnapshot.quantity)
                .filter(StockSnapshot.as_of_date == latest_stock_date)
                .all()
            )
            for pid, sid, qty in stock_snap_q:
                store_stock_map[(pid, sid)] = qty or 0
        
        category_store_stock = defaultdict(lambda: defaultdict(float))
        if latest_stock_date:
            cat_stock_q = (
                db.session.query(
                    Product.category,
                    StockSnapshot.store_id,
                    func.sum(StockSnapshot.quantity)
                )
                .join(Product, StockSnapshot.product_id == Product.id)
                .filter(StockSnapshot.as_of_date == latest_stock_date)
                .filter(Product.category.isnot(None))
                .group_by(Product.category, StockSnapshot.store_id)
                .all()
            )
            for cat, store_id, total_cat_stock in cat_stock_q:
                category_store_stock[cat][store_id] = total_cat_stock or 0
        
        cd_remaining_after_base = {}
        base_assigned_by_product = defaultdict(int)
        for it in final_preds:
            base_assigned_by_product[it["product_id"]] += int(it["suggested"])
        
        for product_id in cd_stock:
            original = cd_stock[product_id]
            assigned = base_assigned_by_product.get(product_id, 0)
            cd_remaining_after_base[product_id] = max(original - assigned, 0)
        
        for sku in skus_eligible_for_cold_start:
            prod = product_map.get(sku)
            if not prod:
                continue
            
            category = prod.category
            if not category or category.strip() == '':
                cold_start_no_category.append({
                    "sku": sku,
                    "product_id": prod.id,
                    "reason": "NO_CATEGORY"
                })
                continue
            
            cat_rates = category_store_rates.get(category, {})
            if not cat_rates:
                cold_start_no_category.append({
                    "sku": sku,
                    "product_id": prod.id,
                    "reason": "NO_CATEGORY_SALES"
                })
                continue
            
            cat_stock_by_store = category_store_stock.get(category, {})
            
            eligible_stores = []
            for store_id, cat_sales_rate in cat_rates.items():
                if cat_sales_rate <= 0:
                    continue
                store_stock = store_stock_map.get((prod.id, store_id), 0)
                if store_stock > 0:
                    continue
                cat_stock = cat_stock_by_store.get(store_id, 0)
                cat_woc = cat_stock / max(cat_sales_rate, 1)
                if cat_woc > MAX_CAT_WOC:
                    continue
                eligible_stores.append((store_id, cat_sales_rate))
            
            top_stores = sorted(eligible_stores, key=lambda x: x[1], reverse=True)[:COLD_START_TOP_STORES]
            
            if not top_stores:
                continue
            
            available_cd = cd_remaining_after_base.get(prod.id, cd_stock.get(prod.id))
            if available_cd is None or available_cd <= 0:
                continue
            
            budget = int(available_cd)
            n_stores = len(top_stores)
            
            store_list = []
            for store_id, cat_sales_rate in top_stores:
                rate = cat_sales_rate
                max_alloc_woc = math.ceil(MAX_WOC_CAP * rate) if rate > 0 else COLD_START_MIN_FILL
                max_cap = min(max_alloc_woc, MAX_UNITS_PER_STORE_EVENT)
                store_list.append({
                    'store_id': store_id,
                    'cat_sales_rate': cat_sales_rate,
                    'max_cap': max_cap,
                    'allocated': 0
                })
            
            for s in store_list:
                if budget <= 0:
                    break
                give = min(COLD_START_MIN_FILL, s['max_cap'], budget)
                s['allocated'] = give
                budget -= give
            
            if budget > 0:
                for s in store_list:
                    if budget <= 0:
                        break
                    room = s['max_cap'] - s['allocated']
                    if room > 0:
                        give = min(room, budget)
                        s['allocated'] += give
                        budget -= give
            
            for _ in range(n_stores):
                changed = False
                for i in range(n_stores - 1):
                    hi = store_list[i]
                    lo = store_list[i + 1]
                    if hi['allocated'] < lo['allocated']:
                        diff = lo['allocated'] - hi['allocated']
                        room_hi = hi['max_cap'] - hi['allocated']
                        if room_hi > 0:
                            transfer = min(diff, room_hi, lo['allocated'] - COLD_START_MIN_FILL) if lo['allocated'] > COLD_START_MIN_FILL else 0
                            if transfer > 0:
                                hi['allocated'] += transfer
                                lo['allocated'] -= transfer
                                changed = True
                        if hi['allocated'] < lo['allocated'] and hi['allocated'] >= hi['max_cap']:
                            lo['allocated'] = hi['allocated']
                            changed = True
                if not changed:
                    break
            
            total_allocated = sum(s['allocated'] for s in store_list)
            for s in store_list:
                store_name = store_map.get(s['store_id'])
                if not store_name:
                    continue
                suggested = int(s['allocated'])
                if suggested > 0:
                    cold_start_candidates.append({
                        "sku": sku,
                        "store": store_name,
                        "product_id": prod.id,
                        "store_id": s['store_id'],
                        "suggested": suggested,
                        "model_name": f"COLD_START_CATEGORY ({category})",
                        "category": category,
                        "cat_sales_rate": s['cat_sales_rate']
                    })
            
            cd_remaining_after_base[prod.id] = max(int(available_cd) - total_allocated, 0)
        
        for cand in cold_start_candidates:
            final_preds.append(cand)
    
    total_units = sum(int(it["suggested"]) for it in final_preds)
    cold_start_used = len([p for p in final_preds if 'COLD_START' in p.get('model_name', '')])
    sales_based_used = len(final_preds) - cold_start_used
    capped_rows = len([p for p in final_preds if p.get('suggested', 0) == MAX_UNITS_PER_STORE_EVENT])
    max_single_row = max((p.get('suggested', 0) for p in final_preds), default=0)
    
    print(f"[DISTRIBUTION DEBUG] === RUN SAFETY CHECK ===")
    print(f"[DISTRIBUTION DEBUG] Scope SKUs: {len(scope_skus_set)}")
    print(f"[DISTRIBUTION DEBUG] SKUs with macro sales data: {len(skus_with_sales_estimation)}")
    print(f"[DISTRIBUTION DEBUG] SKUs without macro sales: {len(skus_without_macro_sales)}")
    print(f"[DISTRIBUTION DEBUG] SKUs with any store stock: {len(skus_with_any_store_stock)}")
    print(f"[DISTRIBUTION DEBUG] SKUs eligible for cold start (no sales AND no stock): {len(skus_eligible_for_cold_start)}")
    print(f"[DISTRIBUTION DEBUG] SKUs skipped (no category): {len(cold_start_no_category)}")
    print(f"[DISTRIBUTION DEBUG] Final predictions: {len(final_preds)} ({sales_based_used} sales-based, {cold_start_used} cold start)")
    print(f"[DISTRIBUTION DEBUG] Total units: {total_units}")
    print(f"[DISTRIBUTION DEBUG] Capped rows (hit MAX_UNITS limit): {capped_rows}")
    print(f"[DISTRIBUTION DEBUG] Max single row qty: {max_single_row}")
    
    if total_units > TOTAL_UNITS_KILL_SWITCH:
        db.session.rollback()
        print(f"[DISTRIBUTION KILL-SWITCH] ABORT: total_units ({total_units}) > {TOTAL_UNITS_KILL_SWITCH}")
        raise ValueError(f"ABORT: Run exceeds safety limit. Total units ({total_units}) > {TOTAL_UNITS_KILL_SWITCH}. Distribution not saved.")
    
    if max_single_row > MAX_UNITS_PER_STORE_EVENT:
        db.session.rollback()
        print(f"[DISTRIBUTION KILL-SWITCH] ABORT: max_single_row ({max_single_row}) > {MAX_UNITS_PER_STORE_EVENT}")
        raise ValueError(f"ABORT: Run exceeds per-store safety limit. Max row ({max_single_row}) > {MAX_UNITS_PER_STORE_EVENT}. Distribution not saved.")
    
    target_week = next_monday()
    preds_bulk = []
    assigned_by_product = defaultdict(int)
    
    for it in final_preds:
        assigned_qty = int(it["suggested"])
        assigned_by_product[it["product_id"]] += assigned_qty
    
    for it in final_preds:
        assigned_qty = int(it["suggested"])
        
        debug_data = {
            'w0_units': it.get('w0_units'),
            'w1_units': it.get('w1_units'),
            'sma_mean': it.get('sma_mean'),
            'weeks_target': it.get('weeks_target'),
            'target_units': it.get('target_units'),
            'stock_store_used': it.get('stock_store_used'),
            'cd_stock_used': it.get('cd_stock_used'),
            'suggested_before_caps': it.get('suggested_before_caps'),
            'suggested_final': it.get('suggested_final', assigned_qty),
            'reason_code': it.get('reason_code', 'OK'),
            'store_rank': it.get('store_rank'),
            'total_sales_in_window': it.get('total_sales_in_window'),
        }
        debug_json_str = json.dumps(debug_data)
        
        existing = Prediction.query.filter_by(
            product_id=it["product_id"],
            store_id=it["store_id"],
            target_period_start=target_week,
            run_id=run_id
        ).first()
        
        if existing:
            existing.quantity = assigned_qty
            existing.model_name = it["model_name"]
            existing.debug_json = debug_json_str
        else:
            preds_bulk.append(Prediction(
                product_id=it["product_id"],
                store_id=it["store_id"],
                target_period_start=target_week,
                quantity=assigned_qty,
                model_name=it["model_name"],
                run_id=run_id,
                debug_json=debug_json_str
            ))
    
    if preds_bulk:
        db.session.add_all(preds_bulk)
    
    if snapshot_date:
        for product_id, assigned_total in assigned_by_product.items():
            cd_row = StockCD.query.filter_by(as_of_date=snapshot_date, product_id=product_id).first()
            if cd_row:
                cd_row.quantity = max(cd_row.quantity - assigned_total, 0)
    
    db.session.commit()
    
    print(f"[DISTRIBUTION DEBUG] === RUN COMPLETED SUCCESSFULLY ===")
    
    return run_id, len(final_preds), total_units


from flask import g
from datetime import date

@app.context_processor
def inject_sidebar_counts():
    """
    Inyecta sidebar_counts en TODAS las plantillas que extienden base.html,
    así evitamos errores de 'sidebar_counts is undefined'.
    """
    try:
        # Total registros de ventas cargadas
        sales_count = DistributionRecord.query.count()

        # SKUs distintos con ventas
        sku_sales = db.session.query(DistributionRecord.product_id).distinct().count()

        # SKUs con stock en tiendas
        store_stock_skus = db.session.query(StockSnapshot.product_id).distinct().count()

        # SKUs con stock en CD (solo > 0)
        cd_stock_skus = (
            db.session.query(StockCD.product_id)
            .filter(StockCD.quantity > 0)
            .distinct()
            .count()
        )

        # Predicciones de la última semana sugerida
        latest_week = db.session.query(db.func.max(Prediction.target_period_start)).scalar()
        if latest_week:
            predictions_count = (
                Prediction.query
                .filter(Prediction.target_period_start == latest_week)
                .count()
            )
        else:
            predictions_count = 0

        pending_approvals_count = 0
        if current_user.is_authenticated and current_user.has_permission('runs:approve'):
            pending_approvals_count = Run.query.filter_by(status='PENDING_APPROVAL').count()

        counts = {
            "dashboard": sales_count,
            "upload": sku_sales,
            "stock_store": store_stock_skus,
            "stock_cd": cd_stock_skus,
            "predictions": predictions_count,
            "pending_approvals": pending_approvals_count,
        }

    except Exception:
        counts = {
            "dashboard": 0,
            "upload": 0,
            "stock_store": 0,
            "stock_cd": 0,
            "predictions": 0,
            "pending_approvals": 0,
        }

    return dict(sidebar_counts=counts)

# ------------------ Rutas ------------------

@app.route('/login', methods=['GET', 'POST'])
def login():
    # si ya está logeado, directo al dashboard
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        username = (request.form.get('username') or "").strip().lower()
        password = request.form.get('password') or ""

        user = User.query.filter_by(username=username).first()
        if user and check_password_hash(user.password_hash, password):
            if not user.is_active:
                flash('Tu cuenta está desactivada. Contacta al administrador.', 'danger')
                return render_template('login.html')
            login_user(user)
            return redirect(url_for('dashboard'))

        flash('Credenciales inválidas.', 'danger')

    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/')
def index():
    # raíz redirige al login o dashboard según estado
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

from datetime import date, timedelta
from sqlalchemy import func

from datetime import date, timedelta
from sqlalchemy import func

from datetime import date, timedelta
from sqlalchemy import func

from datetime import date, timedelta
from collections import defaultdict

from datetime import date, timedelta
from sqlalchemy import func

from datetime import date, timedelta
from sqlalchemy import func

from datetime import date
from sqlalchemy import func, desc

@app.route('/dashboard', methods=['GET'])
@login_required
@require_permission('dashboard:view')
def dashboard():
    # --- Filtros globales ---
    store_filter = (request.args.get('store') or '').strip()
    folio_filter = (request.args.get('folio') or '').strip()
    responsable_filter = (request.args.get('responsable') or '').strip()
    run_id_filter = (request.args.get('run_id') or '').strip()

    # --- Pagination params per table ---
    pred_page = request.args.get('pred_page', 1, type=int)
    pred_size = request.args.get('pred_size', 10, type=int)
    pred_search = (request.args.get('pred_search') or '').strip()

    # Validate page sizes
    pred_size = pred_size if pred_size in [10, 25, 50] else 10

    # --- Fechas / helpers ---
    today = date.today()

    # --- Load available distribution runs (limit 50, newest first) ---
    runs = (
        PredictionRun.query
        .filter(PredictionRun.run_id.isnot(None))
        .filter(PredictionRun.run_id != "")
        .filter(PredictionRun.run_type == 'distribution')
        .order_by(PredictionRun.created_at.desc())
        .limit(50)
        .all()
    )

    # --- Determine selected_run_id (validate it exists) ---
    valid_run_ids = {r.run_id for r in runs}
    if run_id_filter and run_id_filter in valid_run_ids:
        selected_run_id = run_id_filter
    else:
        active_run = next((r for r in runs if r.is_active), None)
        if active_run:
            selected_run_id = active_run.run_id
        elif runs:
            selected_run_id = runs[0].run_id
        else:
            selected_run_id = None

    # --- Última semana de predicción (del run seleccionado) ---
    if selected_run_id:
        latest_week = (
            db.session.query(func.max(Prediction.target_period_start))
            .filter(Prediction.run_id == selected_run_id)
            .scalar()
        )
    else:
        latest_week = None

    # --- Query base de predicciones + joins ---
    pred_q = (
        db.session.query(Prediction, Product, Store)
        .join(Product, Prediction.product_id == Product.id)
        .join(Store, Prediction.store_id == Store.id)
    )

    # Always filter by selected run_id
    if selected_run_id:
        pred_q = pred_q.filter(Prediction.run_id == selected_run_id)
    
    # Only show rows with quantity > 0 (operational filter)
    pred_q = pred_q.filter(Prediction.quantity > 0)

    # Filtro tienda
    if store_filter:
        pred_q = pred_q.filter(Store.name == store_filter)

    # Filtros folio / responsable (filter within the selected run)
    if folio_filter:
        pred_q = pred_q.filter(Prediction.model_name.ilike(f"%Folio:%{folio_filter}%"))
    if responsable_filter:
        pred_q = pred_q.filter(Prediction.model_name.ilike(f"%Resp:%{responsable_filter}%"))

    # Search filter for predictions
    if pred_search:
        pred_q = pred_q.filter(
            or_(
                Product.sku.ilike(f"%{pred_search}%"),
                Product.name.ilike(f"%{pred_search}%"),
                Store.name.ilike(f"%{pred_search}%")
            )
        )

    pred_total = pred_q.count()
    predictions = (
        pred_q
        .order_by(Product.sku.asc(), Store.name.asc())
        .offset((pred_page - 1) * pred_size)
        .limit(pred_size)
        .all()
    )
    pred_pagination = Pagination(pred_page, pred_size, pred_total, predictions)

    # --- KPI (from full filtered set, not just current page) ---
    kpi_q = (
        db.session.query(
            func.coalesce(func.sum(Prediction.quantity), 0),
            func.count(func.distinct(Prediction.product_id)),
            func.count(func.distinct(Prediction.store_id))
        )
        .filter(Prediction.run_id == selected_run_id)
        .filter(Prediction.quantity > 0) if selected_run_id else None
    )
    if kpi_q and store_filter:
        kpi_q = kpi_q.join(Store, Prediction.store_id == Store.id).filter(Store.name == store_filter)
    kpi_result = kpi_q.first() if kpi_q else (0, 0, 0)
    kpi_units_suggested = int(kpi_result[0] or 0) if kpi_result else 0
    kpi_skus_distintos = int(kpi_result[1] or 0) if kpi_result else 0
    kpi_tiendas_alcanzadas = int(kpi_result[2] or 0) if kpi_result else 0

    # Stock CD hoy (total)
    kpi_stock_cd_total = (
        db.session.query(func.coalesce(func.sum(StockCD.quantity), 0))
        .filter(StockCD.as_of_date == today)
        .scalar()
    )

    # --- Executive KPIs (lightweight aggregates) ---
    kpi_total_skus = Product.query.count()
    
    kpi_stores_with_stockout = 0
    kpi_pct_skus_no_movement = 0.0
    try:
        MOVEMENT_DAYS = 30
        cutoff_movement = today - timedelta(days=MOVEMENT_DAYS)
        
        skus_with_movement = (
            db.session.query(func.count(func.distinct(SalesWeeklyAgg.product_id)))
            .filter(SalesWeeklyAgg.week_start >= cutoff_movement)
            .filter(SalesWeeklyAgg.units > 0)
            .scalar() or 0
        )
        if kpi_total_skus > 0:
            kpi_pct_skus_no_movement = round(((kpi_total_skus - skus_with_movement) / kpi_total_skus) * 100, 1)
        
        kpi_stores_with_stockout = (
            db.session.query(func.count(func.distinct(StockStore.store_id)))
            .filter(StockStore.as_of_date == today)
            .filter(StockStore.quantity == 0)
            .scalar() or 0
        )
    except Exception:
        pass

    # --- Lista de tiendas para el select ---
    stores = Store.query.order_by(Store.name.asc()).all()
    
    # --- CD Stock Health Distribution (lightweight aggregation) ---
    cd_health_data = {'healthy': 0, 'slow': 0, 'dead': 0, 'allocated': 0}
    try:
        DEAD_DAYS = 90
        SLOW_DAYS = 30
        cutoff_dead = today - timedelta(days=DEAD_DAYS)
        cutoff_slow = today - timedelta(days=SLOW_DAYS)
        
        last_sale_subq = (
            db.session.query(
                SalesWeeklyAgg.product_id,
                func.max(SalesWeeklyAgg.week_start).label('last_sale')
            )
            .filter(SalesWeeklyAgg.units > 0)
            .group_by(SalesWeeklyAgg.product_id)
            .subquery()
        )
        
        cd_with_sales = (
            db.session.query(
                StockCD.product_id,
                StockCD.quantity,
                last_sale_subq.c.last_sale
            )
            .outerjoin(last_sale_subq, StockCD.product_id == last_sale_subq.c.product_id)
            .filter(StockCD.as_of_date == today)
            .filter(StockCD.quantity > 0)
            .all()
        )
        
        allocated_subq = (
            db.session.query(
                DistributionPlanLine.product_id,
                func.sum(DistributionPlanLine.qty_planned).label('allocated_qty')
            )
            .join(DistributionPlan, DistributionPlanLine.plan_id == DistributionPlan.id)
            .filter(DistributionPlan.status.in_(['APPROVED', 'IN_PROGRESS', 'PACKED']))
            .group_by(DistributionPlanLine.product_id)
            .all()
        )
        allocated_map = {r.product_id: int(r.allocated_qty or 0) for r in allocated_subq}
        
        for row in cd_with_sales:
            qty = int(row.quantity or 0)
            last_sale = row.last_sale
            alloc = min(allocated_map.get(row.product_id, 0), qty)
            remaining = qty - alloc
            cd_health_data['allocated'] += alloc
            
            if remaining > 0:
                if last_sale is None or last_sale < cutoff_dead:
                    cd_health_data['dead'] += remaining
                elif last_sale < cutoff_slow:
                    cd_health_data['slow'] += remaining
                else:
                    cd_health_data['healthy'] += remaining
    except Exception:
        pass
    
    # --- Lightweight alerts summary for dashboard (cached) ---
    try:
        alerts_summary = get_alerts_summary(store_filter=store_filter)
    except Exception:
        alerts_summary = {'high_count': 0, 'medium_count': 0, 'low_count': 0, 'total_count': 0}

    return render_template(
        "dashboard.html",
        stores=stores,
        selected_store=store_filter,
        folio=folio_filter,
        responsable=responsable_filter,

        latest_week=latest_week,

        predictions=predictions,
        pred_pagination=pred_pagination,
        pred_search=pred_search,
        pred_size=pred_size,



        kpi_units_suggested=kpi_units_suggested,
        kpi_skus_distintos=kpi_skus_distintos,
        kpi_tiendas_alcanzadas=kpi_tiendas_alcanzadas,
        kpi_stock_cd_total=int(kpi_stock_cd_total or 0),
        kpi_total_skus=kpi_total_skus,
        kpi_pct_skus_no_movement=kpi_pct_skus_no_movement,
        kpi_stores_with_stockout=kpi_stores_with_stockout,

        cd_health_data=cd_health_data,

        active_simulations={
            sim_type: get_simulation_results(sim_type)
            for sim_type in ['distribution', 'rebalancing', 'forecast']
            if get_simulation_results(sim_type)
        },

        # Alerts summary (lightweight)
        alerts_summary=alerts_summary,

        # (opcional) por si después quieres mostrar “corrida actual”
        runs=runs,
        selected_run_id=selected_run_id,
    )


# ─────────────────────────────────────────────────────────────────────────────
# DASHBOARD BY CATEGORY - Macro operational view per category
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/dashboard_category', methods=['GET'])
@login_required
@require_permission('dashboard:view')
def dashboard_category():
    """
    Dashboard by Category - Macro operational visibility based on recent sales and stock.
    Uses SalesWeeklyAgg for performance (indexed queries).
    """
    from sqlalchemy import distinct as sql_distinct
    
    # --- Get filter params ---
    selected_category = request.args.get('category', '').strip()
    window_days = int(request.args.get('window', 30))
    selected_store = request.args.get('store', '').strip()
    
    # Validate window
    if window_days not in [30, 60, 90]:
        window_days = 30
    
    # Calculate date window
    today = date.today()
    window_start = today - timedelta(days=window_days)
    
    # --- Get distinct categories (non-null, non-empty) ---
    categories = db.session.query(sql_distinct(Product.category)).filter(
        Product.category.isnot(None),
        Product.category != ''
    ).order_by(Product.category).all()
    categories = [c[0] for c in categories if c[0]]
    
    # --- Get stores for filter ---
    stores = Store.query.order_by(Store.name.asc()).all()
    
    # --- Default KPIs ---
    kpi_sales_units = 0
    kpi_avg_weekly_demand = 0.0
    kpi_cd_stock = 0
    kpi_stores_stock = 0
    
    # Chart data
    top_stores_chart = []
    stock_comparison = {'cd': 0, 'stores': 0}
    
    # Top 10 SKUs
    top_skus = []
    
    # Health proportions
    health_data = {'dead_slow': 0, 'healthy': 0}
    
    if selected_category:
        # --- Optimized queries using SalesWeeklyAgg.category index (no product_ids list) ---
        
        # Resolve store filter once
        store_id_filter = None
        if selected_store:
            store_obj = Store.query.filter_by(name=selected_store).first()
            if store_obj:
                store_id_filter = store_obj.id
        
        # --- KPI 1: Sales in window (using SalesWeeklyAgg.category index) ---
        sales_query = db.session.query(func.sum(SalesWeeklyAgg.units)).filter(
            SalesWeeklyAgg.category == selected_category,
            SalesWeeklyAgg.week_start >= window_start
        )
        if store_id_filter:
            sales_query = sales_query.filter(SalesWeeklyAgg.store_id == store_id_filter)
        kpi_sales_units = sales_query.scalar() or 0
        
        # --- KPI 2: Avg weekly demand ---
        weeks_in_window = max(window_days / 7, 1)
        kpi_avg_weekly_demand = round(kpi_sales_units / weeks_in_window, 1)
        
        # --- KPI 3: CD Stock (join Product for category filter) ---
        cd_stock_query = db.session.query(func.sum(StockCD.quantity)).join(
            Product, StockCD.product_id == Product.id
        ).filter(Product.category == selected_category)
        kpi_cd_stock = cd_stock_query.scalar() or 0
        
        # --- KPI 4: Stores Stock (join Product for category filter) ---
        stores_stock_query = db.session.query(func.sum(StockSnapshot.quantity)).join(
            Product, StockSnapshot.product_id == Product.id
        ).filter(Product.category == selected_category)
        if store_id_filter:
            stores_stock_query = stores_stock_query.filter(StockSnapshot.store_id == store_id_filter)
        kpi_stores_stock = stores_stock_query.scalar() or 0
        
        # --- Chart 1: Top 10 stores by category sales (using category index) ---
        top_stores_query = db.session.query(
            Store.name,
            func.sum(SalesWeeklyAgg.units).label('total_units')
        ).join(Store, SalesWeeklyAgg.store_id == Store.id).filter(
            SalesWeeklyAgg.category == selected_category,
            SalesWeeklyAgg.week_start >= window_start
        ).group_by(Store.name).order_by(func.sum(SalesWeeklyAgg.units).desc()).limit(10).all()
        
        top_stores_chart = [{'store': r[0], 'units': int(r[1] or 0)} for r in top_stores_query]
        
        # --- Chart 2: Stock comparison ---
        stock_comparison = {'cd': kpi_cd_stock, 'stores': kpi_stores_stock}
        
        # --- Top 10 SKUs by sales (using category index) ---
        top_skus_query = db.session.query(
            Product.sku,
            Product.name,
            func.sum(SalesWeeklyAgg.units).label('total_units')
        ).join(Product, SalesWeeklyAgg.product_id == Product.id).filter(
            SalesWeeklyAgg.category == selected_category,
            SalesWeeklyAgg.week_start >= window_start
        ).group_by(Product.sku, Product.name).order_by(func.sum(SalesWeeklyAgg.units).desc()).limit(10).all()
        
        top_skus = [{'sku': r[0], 'name': r[1], 'units': int(r[2] or 0)} for r in top_skus_query]
        
        # --- Health proportions (count by category in SQL) ---
        recent_window = today - timedelta(days=60)
        total_products = db.session.query(func.count(Product.id)).filter(
            Product.category == selected_category
        ).scalar() or 0
        
        active_products = db.session.query(func.count(sql_distinct(SalesWeeklyAgg.product_id))).filter(
            SalesWeeklyAgg.category == selected_category,
            SalesWeeklyAgg.week_start >= recent_window,
            SalesWeeklyAgg.units > 0
        ).scalar() or 0
        
        health_data['healthy'] = active_products
        health_data['dead_slow'] = max(total_products - active_products, 0)
    
    return render_template(
        'dashboard_category.html',
        categories=categories,
        selected_category=selected_category,
        stores=stores,
        selected_store=selected_store,
        window_days=window_days,
        
        kpi_sales_units=kpi_sales_units,
        kpi_avg_weekly_demand=kpi_avg_weekly_demand,
        kpi_cd_stock=kpi_cd_stock,
        kpi_stores_stock=kpi_stores_stock,
        
        top_stores_chart=top_stores_chart,
        stock_comparison=stock_comparison,
        top_skus=top_skus,
        health_data=health_data,
    )


@app.route('/dashboard_category/no_movement')
@login_required
@require_permission('dashboard:view')
def category_no_movement():
    """
    Paginated list of SKUs with no movement (stock but no sales in window).
    Optimized: Uses category index directly, no product_ids materialization.
    """
    category = request.args.get('category', '')
    window_days = int(request.args.get('window_days', 30))
    store_name = request.args.get('store', '')
    page = int(request.args.get('page', 1))
    per_page = 10
    
    if not category:
        flash('Debe seleccionar una categoría.', 'warning')
        return redirect(url_for('dashboard_category'))
    
    today = date.today()
    window_start = today - timedelta(days=window_days)
    
    store_id_filter = None
    if store_name:
        store_obj = Store.query.filter_by(name=store_name).first()
        if store_obj:
            store_id_filter = store_obj.id
    
    sales_subq = db.session.query(
        SalesWeeklyAgg.product_id,
        func.sum(SalesWeeklyAgg.units).label('total_sales')
    ).filter(
        SalesWeeklyAgg.category == category,
        SalesWeeklyAgg.week_start >= window_start
    )
    if store_id_filter:
        sales_subq = sales_subq.filter(SalesWeeklyAgg.store_id == store_id_filter)
    sales_subq = sales_subq.group_by(SalesWeeklyAgg.product_id).subquery()
    
    last_sale_subq = db.session.query(
        SalesWeeklyAgg.product_id,
        func.max(SalesWeeklyAgg.week_start).label('last_sale_date')
    ).filter(
        SalesWeeklyAgg.category == category,
        SalesWeeklyAgg.units > 0
    )
    if store_id_filter:
        last_sale_subq = last_sale_subq.filter(SalesWeeklyAgg.store_id == store_id_filter)
    last_sale_subq = last_sale_subq.group_by(SalesWeeklyAgg.product_id).subquery()
    
    cd_stock_subq = db.session.query(
        StockCD.product_id,
        func.sum(StockCD.quantity).label('stock_cd')
    ).join(Product, StockCD.product_id == Product.id).filter(
        Product.category == category
    ).group_by(StockCD.product_id).subquery()
    
    store_stock_subq = db.session.query(
        StockSnapshot.product_id,
        func.sum(StockSnapshot.quantity).label('stock_stores')
    ).join(Product, StockSnapshot.product_id == Product.id).filter(
        Product.category == category
    )
    if store_id_filter:
        store_stock_subq = store_stock_subq.filter(StockSnapshot.store_id == store_id_filter)
    store_stock_subq = store_stock_subq.group_by(StockSnapshot.product_id).subquery()
    
    main_query = db.session.query(
        Product.id,
        Product.sku,
        Product.name,
        Product.eligible_for_distribution,
        Product.risk_score,
        func.coalesce(sales_subq.c.total_sales, 0).label('total_sales'),
        func.coalesce(cd_stock_subq.c.stock_cd, 0).label('stock_cd'),
        func.coalesce(store_stock_subq.c.stock_stores, 0).label('stock_stores'),
        last_sale_subq.c.last_sale_date.label('last_sale_date')
    ).outerjoin(sales_subq, Product.id == sales_subq.c.product_id
    ).outerjoin(cd_stock_subq, Product.id == cd_stock_subq.c.product_id
    ).outerjoin(store_stock_subq, Product.id == store_stock_subq.c.product_id
    ).outerjoin(last_sale_subq, Product.id == last_sale_subq.c.product_id
    ).filter(
        Product.category == category,
        func.coalesce(sales_subq.c.total_sales, 0) == 0,
        (func.coalesce(cd_stock_subq.c.stock_cd, 0) + func.coalesce(store_stock_subq.c.stock_stores, 0)) > 0
    )
    
    total_count = main_query.count()
    total_pages = max((total_count + per_page - 1) // per_page, 1)
    page = max(1, min(page, total_pages))
    
    results = main_query.order_by(
        (func.coalesce(cd_stock_subq.c.stock_cd, 0) + func.coalesce(store_stock_subq.c.stock_stores, 0)).desc(),
        Product.sku
    ).offset((page - 1) * per_page).limit(per_page).all()
    
    items = []
    for r in results:
        total_stock = int(r.stock_cd or 0) + int(r.stock_stores or 0)
        last_sale = r.last_sale_date.strftime('%Y-%m-%d') if r.last_sale_date else None
        risk_band = None
        if r.risk_score is not None:
            if r.risk_score >= 70:
                risk_band = 'HIGH'
            elif r.risk_score >= 40:
                risk_band = 'MEDIUM'
            else:
                risk_band = 'LOW'
        items.append({
            'id': r.id,
            'sku': r.sku,
            'name': r.name,
            'category': category,
            'stock_cd': int(r.stock_cd or 0),
            'stock_stores': int(r.stock_stores or 0),
            'total_stock': total_stock,
            'last_sale': last_sale,
            'flagged': not r.eligible_for_distribution,
            'risk_score': r.risk_score,
            'risk_band': risk_band
        })
    
    return render_template(
        'category_no_movement.html',
        category=category,
        window_days=window_days,
        store_name=store_name,
        items=items,
        total_count=total_count,
        page=page,
        total_pages=total_pages
    )


@app.route('/dashboard_category/no_movement/export')
@login_required
@require_permission('dashboard:view')
def category_no_movement_export():
    """
    Export full list of no-movement SKUs to Excel.
    """
    category = request.args.get('category', '')
    window_days = int(request.args.get('window_days', 30))
    store_name = request.args.get('store', '')
    
    if not category:
        flash('Debe seleccionar una categoría.', 'warning')
        return redirect(url_for('dashboard_category'))
    
    today = date.today()
    window_start = today - timedelta(days=window_days)
    
    store_id_filter = None
    if store_name:
        store_obj = Store.query.filter_by(name=store_name).first()
        if store_obj:
            store_id_filter = store_obj.id
    
    sales_subq = db.session.query(
        SalesWeeklyAgg.product_id,
        func.sum(SalesWeeklyAgg.units).label('total_sales')
    ).filter(
        SalesWeeklyAgg.category == category,
        SalesWeeklyAgg.week_start >= window_start
    )
    if store_id_filter:
        sales_subq = sales_subq.filter(SalesWeeklyAgg.store_id == store_id_filter)
    sales_subq = sales_subq.group_by(SalesWeeklyAgg.product_id).subquery()
    
    last_sale_subq = db.session.query(
        SalesWeeklyAgg.product_id,
        func.max(SalesWeeklyAgg.week_start).label('last_sale_date')
    ).filter(
        SalesWeeklyAgg.category == category,
        SalesWeeklyAgg.units > 0
    )
    if store_id_filter:
        last_sale_subq = last_sale_subq.filter(SalesWeeklyAgg.store_id == store_id_filter)
    last_sale_subq = last_sale_subq.group_by(SalesWeeklyAgg.product_id).subquery()
    
    cd_stock_subq = db.session.query(
        StockCD.product_id,
        func.sum(StockCD.quantity).label('stock_cd')
    ).join(Product, StockCD.product_id == Product.id).filter(
        Product.category == category
    ).group_by(StockCD.product_id).subquery()
    
    store_stock_subq = db.session.query(
        StockSnapshot.product_id,
        func.sum(StockSnapshot.quantity).label('stock_stores')
    ).join(Product, StockSnapshot.product_id == Product.id).filter(
        Product.category == category
    )
    if store_id_filter:
        store_stock_subq = store_stock_subq.filter(StockSnapshot.store_id == store_id_filter)
    store_stock_subq = store_stock_subq.group_by(StockSnapshot.product_id).subquery()
    
    results = db.session.query(
        Product.sku,
        Product.name,
        func.coalesce(cd_stock_subq.c.stock_cd, 0).label('stock_cd'),
        func.coalesce(store_stock_subq.c.stock_stores, 0).label('stock_stores'),
        last_sale_subq.c.last_sale_date.label('last_sale_date')
    ).outerjoin(sales_subq, Product.id == sales_subq.c.product_id
    ).outerjoin(cd_stock_subq, Product.id == cd_stock_subq.c.product_id
    ).outerjoin(store_stock_subq, Product.id == store_stock_subq.c.product_id
    ).outerjoin(last_sale_subq, Product.id == last_sale_subq.c.product_id
    ).filter(
        Product.category == category,
        func.coalesce(sales_subq.c.total_sales, 0) == 0,
        (func.coalesce(cd_stock_subq.c.stock_cd, 0) + func.coalesce(store_stock_subq.c.stock_stores, 0)) > 0
    ).order_by(
        (func.coalesce(cd_stock_subq.c.stock_cd, 0) + func.coalesce(store_stock_subq.c.stock_stores, 0)).desc(),
        Product.sku
    ).all()
    
    rows = []
    for r in results:
        total_stock = int(r.stock_cd or 0) + int(r.stock_stores or 0)
        last_sale = r.last_sale_date.strftime('%Y-%m-%d') if r.last_sale_date else ''
        rows.append({
            'SKU': r.sku,
            'Producto': r.name,
            'Categoría': category,
            'Stock CD': int(r.stock_cd or 0),
            'Stock Tiendas': int(r.stock_stores or 0),
            'Stock Total': total_stock,
            'Última Venta': last_sale
        })
    
    df = pd.DataFrame(rows)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sin Movimiento')
    output.seek(0)
    
    filename = f"sin_movimiento_{category}_{window_days}d.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@app.route('/purchase_forecast', methods=['GET', 'POST'])
@login_required
@require_permission('forecast_v2:view')
def purchase_forecast():
    """Legacy route - redirect to V2."""
    return redirect(url_for('purchase_forecast_v2'))

@app.route('/purchase_forecast_v2', methods=['GET'])
@login_required
@require_permission('forecast_v2:view')
def purchase_forecast_v2():
    """
    Purchase Forecast V2 - Interactive forecast page with Plotly charts.
    Uses /api/forecast_v2 for data.
    """
    stores = Store.query.order_by(Store.name.asc()).all()
    products = Product.query.order_by(Product.sku.asc()).limit(500).all()
    categories = db.session.query(Product.category).filter(
        Product.category.isnot(None)
    ).distinct().order_by(Product.category.asc()).all()
    categories = [c[0] for c in categories if c[0]]

    return render_template(
        'purchase_forecast_v2.html',
        stores=stores,
        products=products,
        categories=categories,
    )


@app.route('/purchase_forecast_v2_legacy', methods=['GET', 'POST'])
@login_required
@require_permission('forecast_v2:view')
def purchase_forecast_v2_legacy():
    """
    Purchase Forecast V2 - Lifecycle-aware forecasting with alerts integration.
    
    Uses ONLY SalesWeeklyAgg as single source of truth for demand calculation.
    Integrates with alerts table and slow stock analysis for risk assessment.
    
    Lifecycle model selection:
    - ACTIVE: SMA from user-selected method (sma2, sma3, last_week)
    - SLOW: Long SMA (8-12 weeks) with demand penalty
    - DEAD: Forecast = 0, no purchase recommended
    - NEW: Category-based cold start model
    """
    from uuid import uuid4
    today = date.today()
    stores = Store.query.order_by(Store.name.asc()).all()
    
    MIN_FILL = 2
    MAX_UNITS_PER_SKU = 10000
    SLOW_DEMAND_PENALTY = 0.5
    SLOW_LOOKBACK_WEEKS = 12

    if request.method == 'POST':
        try:
            lead_time_days = int(request.form.get('lead_time_days', 14))
        except ValueError:
            lead_time_days = 14
        try:
            coverage_weeks = float(request.form.get('coverage_weeks', 4))
        except ValueError:
            coverage_weeks = 4.0
        try:
            safety_weeks = float(request.form.get('safety_weeks', 1))
        except ValueError:
            safety_weeks = 1.0
        try:
            min_weeks_history = int(request.form.get('min_weeks_history', 1))
        except ValueError:
            min_weeks_history = 1

        demand_method = request.form.get('demand_method', 'sma3').strip()
        store_filter = request.form.get('store_filter', '').strip()
        campaign_tag = request.form.get('campaign_tag', '').strip()
    else:
        lead_time_days = 14
        coverage_weeks = 4.0
        safety_weeks = 1.0
        min_weeks_history = 1
        demand_method = 'sma3'
        store_filter = ''
        campaign_tag = ''

    lead_time_days = max(lead_time_days, 0)
    coverage_weeks = max(coverage_weeks, 0.5)
    safety_weeks = max(safety_weeks, 0)
    min_weeks_history = max(min_weeks_history, 1)

    if demand_method == 'sma3':
        lookback_days = 21
        weeks_divisor = 3.0
        model_label = 'SMA3'
    elif demand_method == 'sma2':
        lookback_days = 14
        weeks_divisor = 2.0
        model_label = 'SMA2'
    else:
        lookback_days = 7
        weeks_divisor = 1.0
        model_label = 'LAST_WEEK'

    cutoff = today - timedelta(days=lookback_days)
    cutoff_slow = today - timedelta(days=SLOW_LOOKBACK_WEEKS * 7)
    lead_time_weeks = lead_time_days / 7.0
    total_weeks_needed = coverage_weeks + safety_weeks + lead_time_weeks

    store_id_filter = None
    if store_filter:
        store_obj = Store.query.filter_by(name=store_filter).first()
        if store_obj:
            store_id_filter = store_obj.id

    sales_30d_map, sales_90d_map, sales_ever_map = get_lifecycle_sales_maps(store_id_filter)

    all_products = Product.query.all()
    product_map = {p.id: p for p in all_products}
    all_product_ids = set(product_map.keys())

    lifecycle_map = classify_sku_lifecycle(sales_30d_map, sales_90d_map, sales_ever_map, all_product_ids)

    for pid, status in lifecycle_map.items():
        product = product_map.get(pid)
        if product and product.lifecycle_status != status:
            product.lifecycle_status = status
    
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()

    sales_q = (
        db.session.query(
            Product.id.label('product_id'),
            Product.sku.label('sku'),
            Product.name.label('name'),
            Product.category.label('category'),
            db.func.sum(SalesWeeklyAgg.units).label('qty_period'),
            db.func.count(db.func.distinct(SalesWeeklyAgg.week_start)).label('weeks_with_sales')
        )
        .join(Product, SalesWeeklyAgg.product_id == Product.id)
        .filter(SalesWeeklyAgg.week_start >= cutoff)
    )
    if store_id_filter:
        sales_q = sales_q.filter(SalesWeeklyAgg.store_id == store_id_filter)

    sales_rows_active = (
        sales_q
        .group_by(Product.id, Product.sku, Product.name, Product.category)
        .having(db.func.count(db.func.distinct(SalesWeeklyAgg.week_start)) >= min_weeks_history)
        .all()
    )
    active_sales_map = {r.product_id: r for r in sales_rows_active}

    sales_q_slow = (
        db.session.query(
            Product.id.label('product_id'),
            Product.sku.label('sku'),
            Product.name.label('name'),
            Product.category.label('category'),
            db.func.sum(SalesWeeklyAgg.units).label('qty_period'),
            db.func.count(db.func.distinct(SalesWeeklyAgg.week_start)).label('weeks_with_sales')
        )
        .join(Product, SalesWeeklyAgg.product_id == Product.id)
        .filter(SalesWeeklyAgg.week_start >= cutoff_slow)
    )
    if store_id_filter:
        sales_q_slow = sales_q_slow.filter(SalesWeeklyAgg.store_id == store_id_filter)

    sales_rows_slow = (
        sales_q_slow
        .group_by(Product.id, Product.sku, Product.name, Product.category)
        .all()
    )
    slow_sales_map = {r.product_id: r for r in sales_rows_slow}

    latest_cd_date = db.session.query(db.func.max(StockCD.as_of_date)).scalar() or today
    cd_rows = (
        db.session.query(
            StockCD.product_id,
            db.func.sum(StockCD.quantity).label('cd_qty')
        )
        .filter(StockCD.as_of_date == latest_cd_date)
        .group_by(StockCD.product_id)
        .all()
    )
    cd_stock_map = {r.product_id: int(r.cd_qty) for r in cd_rows}

    latest_store_date = db.session.query(db.func.max(StockSnapshot.as_of_date)).scalar() or today
    store_stock_rows = (
        db.session.query(
            StockSnapshot.product_id,
            db.func.sum(StockSnapshot.quantity).label('store_qty')
        )
        .filter(StockSnapshot.as_of_date == latest_store_date)
        .group_by(StockSnapshot.product_id)
        .all()
    )
    store_stock_map = {r.product_id: int(r.store_qty) for r in store_stock_rows}

    alerts_list = compute_alerts()
    alerts_by_sku = {}
    for alert in alerts_list:
        sku = alert.get('sku')
        if sku not in alerts_by_sku:
            alerts_by_sku[sku] = []
        alerts_by_sku[sku].append(alert)

    slow_stock_skus = set()
    slow_stock_run = SlowStockRun.query.order_by(SlowStockRun.created_at.desc()).first()
    if slow_stock_run:
        slow_suggestions = SlowStockSuggestion.query.filter_by(run_id=slow_stock_run.run_id).all()
        for s in slow_suggestions:
            product = product_map.get(s.product_id)
            if product:
                slow_stock_skus.add(product.sku)

    category_demand_cache = {}

    result = []

    for pid, lifecycle in lifecycle_map.items():
        product = product_map.get(pid)
        if not product:
            continue

        demand_per_week = 0.0
        model_used = model_label
        reason_codes = []
        buy_now = False
        risk_level = 'LOW'

        if lifecycle == 'DEAD':
            model_used = 'DEAD_BLOCK'
            reason_codes.append('NO_RECENT_SALES')
            risk_level = 'HIGH'
            suggested = 0
        elif lifecycle == 'NEW':
            if product.category:
                if product.category not in category_demand_cache:
                    category_demand_cache[product.category] = get_category_avg_demand(
                        product.category, store_id_filter, weeks=4
                    )
                category_avg = category_demand_cache[product.category]
                demand_per_week = category_avg
                model_used = 'CATEGORY_COLD_START'
                reason_codes.append('CATEGORY_COLD_START')
            else:
                demand_per_week = 0.0
                model_used = 'NO_CATEGORY'
                reason_codes.append('NEW_NO_CATEGORY')
            risk_level = 'MEDIUM'
        elif lifecycle == 'SLOW':
            sales_data = slow_sales_map.get(pid)
            if sales_data and sales_data.weeks_with_sales > 0:
                raw_demand = float(sales_data.qty_period) / float(SLOW_LOOKBACK_WEEKS)
                demand_per_week = raw_demand * SLOW_DEMAND_PENALTY
                model_used = f'SMA{SLOW_LOOKBACK_WEEKS}_PENALTY'
                reason_codes.append('SLOW_DEMAND_REDUCED')
            else:
                demand_per_week = 0.0
            risk_level = 'MEDIUM'
        else:
            sales_data = active_sales_map.get(pid)
            if sales_data:
                demand_per_week = float(sales_data.qty_period) / weeks_divisor if weeks_divisor > 0 else 0.0
                model_used = model_label
                reason_codes.append('RECENT_SALES_HIGH' if demand_per_week > 5 else 'RECENT_SALES_OK')

        stock_cd = cd_stock_map.get(pid, 0)
        stock_store = store_stock_map.get(pid, 0)
        total_stock = stock_cd + stock_store

        forecast_demand = demand_per_week * total_weeks_needed
        purchase_qty = max(0, int(round(forecast_demand)) - total_stock)

        if lifecycle == 'DEAD':
            purchase_qty = 0

        sku_alerts = alerts_by_sku.get(product.sku, [])
        for alert in sku_alerts:
            alert_type = alert.get('type', '')
            alert_severity = alert.get('severity', 'LOW')

            if alert_type == 'PROJECTED_STOCKOUT':
                buy_now = True
                if alert_severity == 'HIGH':
                    risk_level = 'HIGH'
                    reason_codes.append('PROJECTED_STOCKOUT_14D')
                else:
                    if risk_level != 'HIGH':
                        risk_level = 'MEDIUM'
                    reason_codes.append('PROJECTED_STOCKOUT_28D')

            elif alert_type == 'BROKEN_STOCK':
                buy_now = True
                if alert_severity == 'HIGH':
                    risk_level = 'HIGH'
                    reason_codes.append('BROKEN_STOCK_URGENT')
                else:
                    if risk_level == 'LOW':
                        risk_level = 'MEDIUM'
                    reason_codes.append('BROKEN_STOCK_RECENT')

            elif alert_type == 'OVERSTOCK':
                purchase_qty = 0
                reason_codes.append('OVERSTOCK_BLOCK')
                risk_level = 'LOW'

            elif alert_type == 'SILENT_SKU':
                reason_codes.append('SILENT_SKU_WARNING')

        if product.sku in slow_stock_skus:
            purchase_qty = int(purchase_qty * 0.5)
            if risk_level == 'LOW':
                risk_level = 'MEDIUM'
            reason_codes.append('SLOW_STOCK_PENALTY')

        if purchase_qty > 0 and purchase_qty < MIN_FILL:
            purchase_qty = MIN_FILL
        purchase_qty = min(purchase_qty, MAX_UNITS_PER_SKU)

        if lifecycle == 'DEAD':
            purchase_qty = 0

        if not reason_codes:
            reason_codes.append('STANDARD_FORECAST')
        reason_code = ';'.join(reason_codes[:3])

        result.append({
            "product_id": pid,
            "sku": product.sku,
            "name": product.name,
            "category": product.category or '',
            "demand_per_week": round(demand_per_week, 2),
            "available_cd": stock_cd,
            "stock_store_total": stock_store,
            "required_units": round(forecast_demand, 2),
            "suggested": purchase_qty,
            "campaign_tag": campaign_tag,
            "buy_now": buy_now,
            "risk_level": risk_level,
            "reason_code": reason_code,
            "model_used": model_used,
            "lifecycle_status": lifecycle,
        })

    result = [r for r in result if r['suggested'] > 0 or r['buy_now']]
    result.sort(key=lambda x: (0 if x['buy_now'] else 1, -x['suggested']))

    has_cd_stock = bool(cd_rows)
    has_sales = bool(sales_rows_active)

    forecast_run_id = None
    if request.method == 'POST' and result:
        forecast_run_id = str(uuid4())
        forecast_run = Run(
            run_id=forecast_run_id,
            run_type='forecast_v2',
            folio=campaign_tag or None,
            store_filter=store_filter or None,
            notes=f"method={demand_method}, lead={lead_time_days}d, coverage={coverage_weeks}w, safety={safety_weeks}w, lifecycle_aware=true",
            status='completed',
            mode=demand_method
        )
        db.session.add(forecast_run)

        for r in result:
            fr = ForecastResult(
                run_id=forecast_run_id,
                product_id=r['product_id'],
                sku=str(r['sku']),
                name=r['name'],
                demand_per_week=r['demand_per_week'],
                available_cd=r['available_cd'],
                required_units=r['required_units'],
                suggested=r['suggested'],
                campaign_tag=campaign_tag or None,
                buy_now=r['buy_now'],
                risk_level=r['risk_level'],
                reason_code=r['reason_code'],
                model_used=r['model_used'],
                lifecycle_status=r['lifecycle_status'],
                stock_store_total=r['stock_store_total']
            )
            db.session.add(fr)

        db.session.commit()

    result_limited = result[:50]

    return render_template(
        'purchase_forecast_v2.html',
        rows=result_limited,
        total_skus=len(result),
        stores=stores,
        lead_time_days=lead_time_days,
        coverage_weeks=coverage_weeks,
        safety_weeks=safety_weeks,
        min_weeks_history=min_weeks_history,
        demand_method=demand_method,
        store_filter=store_filter,
        campaign_tag=campaign_tag,
        today=today,
        cd_snapshot_date=latest_cd_date,
        has_cd_stock=has_cd_stock,
        has_sales=has_sales,
        forecast_run_id=forecast_run_id,
    )

@app.route('/export_purchase_forecast_v2', methods=['POST'])
@login_required
@require_permission('forecast_v2:run')
def export_purchase_forecast_v2():
    """Export Purchase Forecast V2 results to Excel with lifecycle-aware logic."""
    from datetime import datetime
    today = date.today()
    
    MIN_FILL = 2
    MAX_UNITS_PER_SKU = 10000
    SLOW_DEMAND_PENALTY = 0.5
    SLOW_LOOKBACK_WEEKS = 12

    try:
        lead_time_days = int(request.form.get('lead_time_days', 14))
    except ValueError:
        lead_time_days = 14
    try:
        coverage_weeks = float(request.form.get('coverage_weeks', 4))
    except ValueError:
        coverage_weeks = 4.0
    try:
        safety_weeks = float(request.form.get('safety_weeks', 1))
    except ValueError:
        safety_weeks = 1.0
    try:
        min_weeks_history = int(request.form.get('min_weeks_history', 1))
    except ValueError:
        min_weeks_history = 1

    demand_method = request.form.get('demand_method', 'sma3').strip()
    store_filter = request.form.get('store_filter', '').strip()
    campaign_tag = request.form.get('campaign_tag', '').strip()

    lead_time_days = max(lead_time_days, 0)
    coverage_weeks = max(coverage_weeks, 0.5)
    safety_weeks = max(safety_weeks, 0)
    min_weeks_history = max(min_weeks_history, 1)

    if demand_method == 'sma3':
        lookback_days = 21
        weeks_divisor = 3.0
        model_label = 'SMA3'
    elif demand_method == 'sma2':
        lookback_days = 14
        weeks_divisor = 2.0
        model_label = 'SMA2'
    else:
        lookback_days = 7
        weeks_divisor = 1.0
        model_label = 'LAST_WEEK'

    cutoff = today - timedelta(days=lookback_days)
    cutoff_slow = today - timedelta(days=SLOW_LOOKBACK_WEEKS * 7)
    lead_time_weeks = lead_time_days / 7.0
    total_weeks_needed = coverage_weeks + safety_weeks + lead_time_weeks

    store_id_filter = None
    if store_filter:
        store_obj = Store.query.filter_by(name=store_filter).first()
        if store_obj:
            store_id_filter = store_obj.id

    sales_30d_map, sales_90d_map, sales_ever_map = get_lifecycle_sales_maps(store_id_filter)

    all_products = Product.query.all()
    product_map = {p.id: p for p in all_products}
    all_product_ids = set(product_map.keys())

    lifecycle_map = classify_sku_lifecycle(sales_30d_map, sales_90d_map, sales_ever_map, all_product_ids)

    sales_q = (
        db.session.query(
            Product.id.label('product_id'),
            Product.sku.label('sku'),
            Product.name.label('name'),
            Product.category.label('category'),
            db.func.sum(SalesWeeklyAgg.units).label('qty_period'),
            db.func.count(db.func.distinct(SalesWeeklyAgg.week_start)).label('weeks_with_sales')
        )
        .join(Product, SalesWeeklyAgg.product_id == Product.id)
        .filter(SalesWeeklyAgg.week_start >= cutoff)
    )
    if store_id_filter:
        sales_q = sales_q.filter(SalesWeeklyAgg.store_id == store_id_filter)
    sales_rows_active = (
        sales_q
        .group_by(Product.id, Product.sku, Product.name, Product.category)
        .having(db.func.count(db.func.distinct(SalesWeeklyAgg.week_start)) >= min_weeks_history)
        .all()
    )
    active_sales_map = {r.product_id: r for r in sales_rows_active}

    sales_q_slow = (
        db.session.query(
            Product.id.label('product_id'),
            db.func.sum(SalesWeeklyAgg.units).label('qty_period'),
        )
        .join(Product, SalesWeeklyAgg.product_id == Product.id)
        .filter(SalesWeeklyAgg.week_start >= cutoff_slow)
    )
    if store_id_filter:
        sales_q_slow = sales_q_slow.filter(SalesWeeklyAgg.store_id == store_id_filter)
    sales_rows_slow = sales_q_slow.group_by(Product.id).all()
    slow_sales_map = {r.product_id: r for r in sales_rows_slow}

    latest_cd_date = db.session.query(db.func.max(StockCD.as_of_date)).scalar() or today
    cd_rows = (
        db.session.query(StockCD.product_id, db.func.sum(StockCD.quantity).label('cd_qty'))
        .filter(StockCD.as_of_date == latest_cd_date)
        .group_by(StockCD.product_id)
        .all()
    )
    cd_stock_map = {r.product_id: int(r.cd_qty) for r in cd_rows}

    latest_store_date = db.session.query(db.func.max(StockSnapshot.as_of_date)).scalar() or today
    store_stock_rows = (
        db.session.query(StockSnapshot.product_id, db.func.sum(StockSnapshot.quantity).label('store_qty'))
        .filter(StockSnapshot.as_of_date == latest_store_date)
        .group_by(StockSnapshot.product_id)
        .all()
    )
    store_stock_map = {r.product_id: int(r.store_qty) for r in store_stock_rows}

    alerts_list = compute_alerts()
    alerts_by_sku = {}
    for alert in alerts_list:
        sku = alert.get('sku')
        if sku not in alerts_by_sku:
            alerts_by_sku[sku] = []
        alerts_by_sku[sku].append(alert)

    slow_stock_skus = set()
    slow_stock_run = SlowStockRun.query.order_by(SlowStockRun.created_at.desc()).first()
    if slow_stock_run:
        slow_suggestions = SlowStockSuggestion.query.filter_by(run_id=slow_stock_run.run_id).all()
        for s in slow_suggestions:
            product = product_map.get(s.product_id)
            if product:
                slow_stock_skus.add(product.sku)

    category_demand_cache = {}

    rows = []
    for pid, lifecycle in lifecycle_map.items():
        product = product_map.get(pid)
        if not product:
            continue

        demand_per_week = 0.0
        model_used = model_label
        reason_codes = []
        buy_now = False
        risk_level = 'LOW'

        if lifecycle == 'DEAD':
            model_used = 'DEAD_BLOCK'
            reason_codes.append('NO_RECENT_SALES')
            risk_level = 'HIGH'
            suggested = 0
        elif lifecycle == 'NEW':
            if product.category:
                if product.category not in category_demand_cache:
                    category_demand_cache[product.category] = get_category_avg_demand(product.category, store_id_filter, weeks=4)
                category_avg = category_demand_cache[product.category]
                demand_per_week = category_avg
                model_used = 'CATEGORY_COLD_START'
                reason_codes.append('CATEGORY_COLD_START')
            else:
                demand_per_week = 0.0
                model_used = 'NO_CATEGORY'
                reason_codes.append('NEW_NO_CATEGORY')
            risk_level = 'MEDIUM'
        elif lifecycle == 'SLOW':
            sales_data = slow_sales_map.get(pid)
            if sales_data and sales_data.qty_period > 0:
                raw_demand = float(sales_data.qty_period) / float(SLOW_LOOKBACK_WEEKS)
                demand_per_week = raw_demand * SLOW_DEMAND_PENALTY
                model_used = f'SMA{SLOW_LOOKBACK_WEEKS}_PENALTY'
                reason_codes.append('SLOW_DEMAND_REDUCED')
            else:
                demand_per_week = 0.0
            risk_level = 'MEDIUM'
        else:
            sales_data = active_sales_map.get(pid)
            if sales_data:
                demand_per_week = float(sales_data.qty_period) / weeks_divisor if weeks_divisor > 0 else 0.0
                model_used = model_label
                reason_codes.append('RECENT_SALES_HIGH' if demand_per_week > 5 else 'RECENT_SALES_OK')

        stock_cd = cd_stock_map.get(pid, 0)
        stock_store = store_stock_map.get(pid, 0)
        total_stock = stock_cd + stock_store

        forecast_demand = demand_per_week * total_weeks_needed
        purchase_qty = max(0, int(round(forecast_demand)) - total_stock)

        if lifecycle == 'DEAD':
            purchase_qty = 0

        sku_alerts = alerts_by_sku.get(product.sku, [])
        for alert in sku_alerts:
            alert_type = alert.get('type', '')
            alert_severity = alert.get('severity', 'LOW')
            if alert_type == 'PROJECTED_STOCKOUT':
                buy_now = True
                if alert_severity == 'HIGH':
                    risk_level = 'HIGH'
                    reason_codes.append('PROJECTED_STOCKOUT_14D')
                else:
                    if risk_level != 'HIGH':
                        risk_level = 'MEDIUM'
                    reason_codes.append('PROJECTED_STOCKOUT_28D')
            elif alert_type == 'BROKEN_STOCK':
                if alert_severity == 'HIGH':
                    buy_now = True
                    risk_level = 'HIGH'
                    reason_codes.append('BROKEN_STOCK_URGENT')
            elif alert_type == 'OVERSTOCK':
                purchase_qty = 0
                reason_codes.append('OVERSTOCK_BLOCK')
                risk_level = 'LOW'

        if product.sku in slow_stock_skus:
            purchase_qty = int(purchase_qty * 0.5)
            if risk_level == 'LOW':
                risk_level = 'MEDIUM'
            reason_codes.append('SLOW_STOCK_PENALTY')

        if purchase_qty > 0 and purchase_qty < MIN_FILL:
            purchase_qty = MIN_FILL
        purchase_qty = min(purchase_qty, MAX_UNITS_PER_SKU)

        if lifecycle == 'DEAD':
            purchase_qty = 0

        if not reason_codes:
            reason_codes.append('STANDARD_FORECAST')
        reason_code = ';'.join(reason_codes[:3])

        if purchase_qty > 0 or buy_now:
            rows.append({
                "SKU": str(product.sku),
                "Producto": product.name,
                "Categoria": product.category or '',
                "Lifecycle": lifecycle,
                "Demanda/Semana": round(demand_per_week, 2),
                "Stock CD": stock_cd,
                "Stock Tiendas": stock_store,
                "Stock Total": total_stock,
                "Requerido": round(forecast_demand, 2),
                "Sugerido Compra": purchase_qty,
                "Comprar Ahora": "SI" if buy_now else "NO",
                "Riesgo": risk_level,
                "Razon": reason_code,
                "Modelo": model_used,
                "Campana": campaign_tag or "",
                "Lead Time Dias": lead_time_days,
                "Cobertura Sem": coverage_weeks,
                "Safety Sem": safety_weeks,
            })

    rows.sort(key=lambda x: (0 if x["Comprar Ahora"] == "SI" else 1, -x["Sugerido Compra"]))

    if not rows:
        rows.append({
            "SKU": "",
            "Producto": "Sin datos",
            "Categoria": "",
            "Lifecycle": "",
            "Demanda/Semana": 0,
            "Stock CD": 0,
            "Stock Tiendas": 0,
            "Stock Total": 0,
            "Requerido": 0,
            "Sugerido Compra": 0,
            "Comprar Ahora": "NO",
            "Riesgo": "",
            "Razon": "",
            "Modelo": "",
            "Campana": campaign_tag or "",
            "Lead Time Dias": lead_time_days,
            "Cobertura Sem": coverage_weeks,
            "Safety Sem": safety_weeks,
        })

    df = pd.DataFrame(rows)
    df["SKU"] = df["SKU"].astype(str)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Forecast Compra V2")
        ws = writer.sheets["Forecast Compra V2"]
        for cell in ws["A"]:
            cell.number_format = "@"

    output.seek(0)
    ts = datetime.now().strftime('%Y%m%d_%H%M')
    tag_part = f"_{campaign_tag.replace(' ', '_')}" if campaign_tag else ""
    fname = f"forecast_compra_v2_{ts}{tag_part}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route('/api/forecast_v2/chart_data', methods=['GET'])
@login_required
@require_permission('forecast_v2:view')
def api_forecast_v2_chart_data():
    """
    API endpoint returning forecast + sales history data as JSON for Plotly charts.
    Query params: sku (optional), store_filter (optional), demand_method, lead_time_days, coverage_weeks, safety_weeks
    """
    from flask import jsonify
    today = date.today()

    sku_filter = request.args.get('sku', '').strip()
    store_filter = request.args.get('store_filter', '').strip()
    demand_method = request.args.get('demand_method', 'sma3').strip()

    try:
        lead_time_days = int(request.args.get('lead_time_days', 14))
    except ValueError:
        lead_time_days = 14
    try:
        coverage_weeks = float(request.args.get('coverage_weeks', 4))
    except ValueError:
        coverage_weeks = 4.0
    try:
        safety_weeks = float(request.args.get('safety_weeks', 1))
    except ValueError:
        safety_weeks = 1.0

    lead_time_days = max(lead_time_days, 0)
    coverage_weeks = max(coverage_weeks, 0.5)
    safety_weeks = max(safety_weeks, 0)

    if demand_method == 'sma3':
        lookback_days = 21
        weeks_divisor = 3.0
    elif demand_method == 'sma2':
        lookback_days = 14
        weeks_divisor = 2.0
    else:
        lookback_days = 7
        weeks_divisor = 1.0

    lead_time_weeks = lead_time_days / 7.0
    total_weeks_needed = coverage_weeks + safety_weeks + lead_time_weeks

    history_weeks = 12
    history_cutoff = today - timedelta(days=history_weeks * 7)

    week_expr = db.func.strftime('%Y-%W', DistributionRecord.event_date)

    history_q = (
        db.session.query(
            Product.sku.label('sku'),
            Product.name.label('name'),
            week_expr.label('week'),
            db.func.sum(DistributionRecord.quantity).label('qty')
        )
        .join(Product, DistributionRecord.product_id == Product.id)
        .join(Store, DistributionRecord.store_id == Store.id)
        .filter(DistributionRecord.event_date >= history_cutoff)
    )

    if sku_filter:
        history_q = history_q.filter(Product.sku == sku_filter)
    if store_filter:
        history_q = history_q.filter(Store.name == store_filter)

    history_rows = (
        history_q
        .group_by(Product.sku, Product.name, week_expr)
        .order_by(Product.sku, week_expr)
        .all()
    )

    history_by_sku = {}
    for r in history_rows:
        sku_str = str(r.sku)
        if sku_str not in history_by_sku:
            history_by_sku[sku_str] = {'name': r.name, 'weeks': [], 'quantities': []}
        history_by_sku[sku_str]['weeks'].append(r.week)
        history_by_sku[sku_str]['quantities'].append(int(r.qty))

    cutoff = today - timedelta(days=lookback_days)
    sales_q = (
        db.session.query(
            Product.id.label('product_id'),
            Product.sku.label('sku'),
            Product.name.label('name'),
            db.func.sum(DistributionRecord.quantity).label('qty_period'),
        )
        .join(Product, DistributionRecord.product_id == Product.id)
        .join(Store, DistributionRecord.store_id == Store.id)
        .filter(DistributionRecord.event_date >= cutoff)
    )

    if sku_filter:
        sales_q = sales_q.filter(Product.sku == sku_filter)
    if store_filter:
        sales_q = sales_q.filter(Store.name == store_filter)

    sales_rows = (
        sales_q
        .group_by(Product.id, Product.sku, Product.name)
        .all()
    )

    latest_cd_date = db.session.query(db.func.max(StockCD.as_of_date)).scalar() or today
    cd_rows = (
        db.session.query(
            StockCD.product_id,
            db.func.sum(StockCD.quantity).label('cd_qty')
        )
        .filter(StockCD.as_of_date == latest_cd_date)
        .group_by(StockCD.product_id)
        .all()
    )
    cd_stock_map = {r.product_id: int(r.cd_qty) for r in cd_rows}

    forecast_data = []
    for r in sales_rows:
        demand_per_week = float(r.qty_period) / weeks_divisor if weeks_divisor > 0 else 0.0
        required_units = demand_per_week * total_weeks_needed
        available_units = cd_stock_map.get(r.product_id, 0)
        suggested = max(int(round(required_units)) - available_units, 0)

        sku_str = str(r.sku)
        hist = history_by_sku.get(sku_str, {'weeks': [], 'quantities': []})

        forecast_data.append({
            "sku": sku_str,
            "name": r.name,
            "demand_per_week": round(demand_per_week, 2),
            "available_cd": available_units,
            "required_units": round(required_units, 2),
            "suggested": suggested,
            "history_weeks": hist['weeks'],
            "history_quantities": hist['quantities'],
        })

    forecast_data.sort(key=lambda x: x["suggested"], reverse=True)

    all_skus = [{"sku": str(p.sku), "name": p.name} for p in Product.query.order_by(Product.sku).all()]

    total_suggested = sum(f['suggested'] for f in forecast_data)
    total_demand = sum(f['demand_per_week'] for f in forecast_data)
    total_stock_cd = sum(f['available_cd'] for f in forecast_data)
    skus_with_purchase = sum(1 for f in forecast_data if f['suggested'] > 0)

    if sku_filter:
        forecast_response = [f for f in forecast_data if f['sku'] == sku_filter]
        if not forecast_response:
            hist = history_by_sku.get(sku_filter, {'weeks': [], 'quantities': []})
            product = Product.query.filter_by(sku=sku_filter).first()
            if product:
                forecast_response = [{
                    "sku": sku_filter,
                    "name": product.name,
                    "demand_per_week": 0,
                    "available_cd": cd_stock_map.get(product.id, 0),
                    "required_units": 0,
                    "suggested": 0,
                    "history_weeks": hist['weeks'],
                    "history_quantities": hist['quantities'],
                }]
    else:
        forecast_response = forecast_data[:100]

    return jsonify({
        "forecast": forecast_response,
        "all_skus": all_skus,
        "kpis": {
            "total_suggested": total_suggested,
            "total_demand_per_week": round(total_demand, 2),
            "total_stock_cd": total_stock_cd,
            "skus_with_purchase": skus_with_purchase,
            "total_skus": len(forecast_data),
        },
        "params": {
            "demand_method": demand_method,
            "lead_time_days": lead_time_days,
            "coverage_weeks": coverage_weeks,
            "safety_weeks": safety_weeks,
            "store_filter": store_filter,
            "sku_filter": sku_filter,
        }
    })


@app.route('/api/forecast_v2', methods=['GET'])
@login_required
@require_permission('forecast_v2:view')
def api_forecast_v2():
    """
    Detailed forecast API for a specific SKU.
    Query params:
      - sku (required): SKU code
      - store (optional): store name, empty = all stores
      - horizon_weeks (default 8): forecast horizon
      - history_weeks (default 12): weeks of history to return
      - lead_time_weeks (default 4): lead time in weeks
      - safety_pct (default 0.10): safety stock percentage
    """
    import math
    from flask import jsonify
    today = date.today()

    sku_param = request.args.get('sku', '').strip()
    if not sku_param:
        return jsonify({"error": "sku parameter is required"}), 400

    store_param = request.args.get('store', '').strip()

    try:
        horizon_weeks = int(request.args.get('horizon_weeks', 8))
    except ValueError:
        horizon_weeks = 8
    try:
        history_weeks = int(request.args.get('history_weeks', 12))
    except ValueError:
        history_weeks = 12
    try:
        lead_time_weeks = float(request.args.get('lead_time_weeks', 4))
    except ValueError:
        lead_time_weeks = 4.0
    try:
        safety_pct = float(request.args.get('safety_pct', 0.10))
    except ValueError:
        safety_pct = 0.10

    horizon_weeks = max(horizon_weeks, 1)
    history_weeks = max(history_weeks, 1)
    lead_time_weeks = max(lead_time_weeks, 0)
    safety_pct = max(safety_pct, 0)

    product = Product.query.filter_by(sku=sku_param).first()
    if not product:
        log_audit(
            action='forecast_v2.run',
            status='fail',
            message=f"SKU not found: {sku_param}",
            metadata={'sku': sku_param, 'store': store_param or 'ALL'}
        )
        return jsonify({"error": f"SKU '{sku_param}' not found"}), 404

    history_cutoff = today - timedelta(days=history_weeks * 7)

    history_q = (
        db.session.query(
            SalesWeeklyAgg.week_start,
            db.func.sum(SalesWeeklyAgg.units).label('units')
        )
        .filter(SalesWeeklyAgg.product_id == product.id)
        .filter(SalesWeeklyAgg.week_start >= history_cutoff)
    )

    if store_param:
        store_obj = Store.query.filter_by(name=store_param).first()
        if store_obj:
            history_q = history_q.filter(SalesWeeklyAgg.store_id == store_obj.id)

    history_rows = (
        history_q
        .group_by(SalesWeeklyAgg.week_start)
        .order_by(SalesWeeklyAgg.week_start)
        .all()
    )

    history_data = [{"week": str(r.week_start), "units": int(r.units)} for r in history_rows]

    if len(history_data) >= 4:
        last4_units = [h['units'] for h in history_data[-4:]]
        avg_last4 = sum(last4_units) / 4.0
    elif history_data:
        avg_last4 = sum(h['units'] for h in history_data) / len(history_data)
    else:
        avg_last4 = 0

    if not history_data:
        log_audit(
            action='forecast_v2.run',
            status='success',
            message=f"No history data for SKU {sku_param}",
            metadata={
                'sku': sku_param,
                'store': store_param or 'ALL',
                'horizon_weeks': horizon_weeks,
                'lead_time_weeks': lead_time_weeks,
                'safety_pct': safety_pct,
                'kpis': {'avg_last4': 0, 'forecast_total': 0, 'stock_cd': 0, 'suggested_purchase': 0, 'weeks_of_cover': 0}
            }
        )
        return jsonify({
            "sku": str(sku_param),
            "store": store_param if store_param else "ALL",
            "history": [],
            "forecast": [],
            "bands": {"lower": [], "upper": []},
            "kpis": {
                "avg_last4": 0,
                "forecast_total": 0,
                "stock_cd": 0,
                "suggested_purchase": 0,
                "weeks_of_cover": 0
            }
        })

    forecast_data = []
    current_week = today - timedelta(days=today.weekday())
    for i in range(horizon_weeks):
        week_start = current_week + timedelta(weeks=i)
        forecast_units = round(avg_last4)
        forecast_data.append({"week": week_start.strftime('%Y-%m-%d'), "units": forecast_units})

    if len(history_data) >= 4:
        last4_units = [h['units'] for h in history_data[-4:]]
        std_dev = (sum((x - avg_last4) ** 2 for x in last4_units) / 4) ** 0.5
    else:
        std_dev = avg_last4 * 0.2

    bands_lower = [max(0, round(avg_last4 - 1.5 * std_dev)) for _ in range(horizon_weeks)]
    bands_upper = [round(avg_last4 + 1.5 * std_dev) for _ in range(horizon_weeks)]

    latest_cd_date = db.session.query(db.func.max(StockCD.as_of_date)).scalar() or today
    cd_stock_row = (
        db.session.query(db.func.sum(StockCD.quantity).label('qty'))
        .filter(StockCD.product_id == product.id)
        .filter(StockCD.as_of_date == latest_cd_date)
        .first()
    )
    stock_cd = int(cd_stock_row.qty) if cd_stock_row and cd_stock_row.qty else 0

    latest_store_date = db.session.query(db.func.max(StockSnapshot.as_of_date)).scalar() or today
    store_stock_q = (
        db.session.query(db.func.sum(StockSnapshot.quantity).label('qty'))
        .filter(StockSnapshot.product_id == product.id)
        .filter(StockSnapshot.as_of_date == latest_store_date)
    )
    if store_param:
        store_obj = Store.query.filter_by(name=store_param).first()
        if store_obj:
            store_stock_q = store_stock_q.filter(StockSnapshot.store_id == store_obj.id)
    store_stock_row = store_stock_q.first()
    stock_stores = int(store_stock_row.qty) if store_stock_row and store_stock_row.qty else 0
    
    stock_total = stock_cd + stock_stores

    forecast_total = sum(f['units'] for f in forecast_data)
    demand_over_lead_time = avg_last4 * lead_time_weeks
    safety_units = math.ceil(demand_over_lead_time * safety_pct)
    raw_purchase_qty = max(0, math.ceil(demand_over_lead_time + safety_units - stock_total))

    weeks_of_cover = round(stock_total / max(avg_last4, 1), 1)

    kpis = {
        "avg_last4": round(avg_last4, 2),
        "forecast_total": forecast_total,
        "stock_cd": stock_total,
        "stock_stores": stock_stores,
        "stock_cd_only": stock_cd,
        "suggested_purchase": raw_purchase_qty,
        "weeks_of_cover": weeks_of_cover
    }

    store_id_filter = None
    if store_param:
        store_obj = Store.query.filter_by(name=store_param).first()
        if store_obj:
            store_id_filter = store_obj.id

    sales_30d_map, sales_90d_map, sales_ever_map = get_lifecycle_sales_maps(store_id_filter)
    lifecycle_map = classify_sku_lifecycle(sales_30d_map, sales_90d_map, sales_ever_map, {product.id})
    lifecycle_status = lifecycle_map.get(product.id, 'UNKNOWN')
    
    all_alerts = compute_alerts()
    sku_alerts = [a for a in all_alerts if a.get('sku', '').lower() == sku_param.lower()]
    
    slow_stock_flagged = not product.eligible_for_distribution if hasattr(product, 'eligible_for_distribution') else False
    
    # Use SET for deduplicated reason codes
    reason_codes_set = set()
    risk_level = 'LOW'
    
    # Detect broken_stock_flag from alerts
    broken_stock_flag = False
    for alert in sku_alerts:
        alert_type = alert.get('type', '')
        alert_severity = alert.get('severity', 'LOW')
        if alert_type == 'PROJECTED_STOCKOUT':
            risk_level = 'HIGH' if alert_severity == 'HIGH' else 'MEDIUM'
            reason_codes_set.add('PROJECTED_STOCKOUT')
        elif alert_type == 'BROKEN_STOCK':
            broken_stock_flag = True
            if alert_severity == 'HIGH':
                risk_level = 'HIGH'
            reason_codes_set.add('BROKEN_STOCK')
        elif alert_type == 'OVERSTOCK':
            reason_codes_set.add('OVERSTOCK')
        elif alert_type == 'SILENT_SKU':
            reason_codes_set.add('SILENT_SKU')
    
    if slow_stock_flagged:
        reason_codes_set.add('SLOW_STOCK_FLAG')
        if risk_level == 'LOW':
            risk_level = 'MEDIUM'
    
    if lifecycle_status == 'DEAD':
        reason_codes_set.add('DEAD_SKU')
    elif lifecycle_status == 'SLOW':
        reason_codes_set.add('SLOW_SKU')
    
    # Decision engine with proper precedence (per spec)
    MAX_WOC = 6.0  # Max weeks of cover threshold for overstock
    
    # Redefine overstock_flag to avoid contradiction: only if NOT broken stock
    overstock_flag = (weeks_of_cover > MAX_WOC) and (not broken_stock_flag)
    
    # Decision precedence policy:
    # A) If overstock_flag == True -> DO_NOT_BUY
    # B) Else if broken_stock_flag and raw_purchase_qty == 0 -> REVIEW (redistribute first)
    # C) Else if raw_purchase_qty > 0 -> BUY_NOW or REVIEW based on coverage
    # D) Else -> DO_NOT_BUY based on lifecycle/other rules
    
    if overstock_flag:
        decision = 'DO_NOT_BUY'
        final_purchase_qty = 0
        reason_codes_set.add('OVERSTOCK_BLOCK')
    elif lifecycle_status == 'DEAD':
        decision = 'DO_NOT_BUY'
        final_purchase_qty = 0
    elif broken_stock_flag and raw_purchase_qty == 0:
        # Stores have stockouts but total stock covers demand -> redistribute
        decision = 'REVIEW'
        final_purchase_qty = 0
        reason_codes_set.add('BROKEN_STOCK_REBALANCE')
    elif raw_purchase_qty > 0:
        # Need to purchase
        if broken_stock_flag:
            reason_codes_set.add('BROKEN_STOCK')
        reason_codes_set.add('DEMAND_EXCEEDS_STOCK')
        if weeks_of_cover < lead_time_weeks:
            decision = 'BUY_NOW'
        else:
            decision = 'BUY_NOW' if risk_level == 'HIGH' else 'REVIEW'
        final_purchase_qty = raw_purchase_qty
    else:
        decision = 'DO_NOT_BUY'
        final_purchase_qty = 0
    
    # Consistency guards
    if decision == 'DO_NOT_BUY':
        final_purchase_qty = 0
    elif final_purchase_qty > 0 and decision == 'DO_NOT_BUY':
        # Force REVIEW if we have units to buy
        decision = 'REVIEW'
    
    # Convert set to list for output
    reason_codes = list(reason_codes_set)
    buy_now = (decision == 'BUY_NOW')
    
    kpis['suggested_purchase'] = final_purchase_qty
    
    model_used = 'SMA4'
    if lifecycle_status == 'SLOW':
        model_used = 'SMA12_PENALTY'
    elif lifecycle_status == 'NEW':
        model_used = 'CATEGORY_AVG'
    
    explain_bullets = generate_forecast_explanation(
        avg_weekly_demand=avg_last4,
        horizon_demand=forecast_total,
        total_stock=stock_total,
        weeks_of_cover=weeks_of_cover,
        lead_time_weeks=lead_time_weeks,
        safety_pct=safety_pct,
        lifecycle_status=lifecycle_status,
        alerts=sku_alerts,
        slow_stock_flagged=slow_stock_flagged,
        suggested_purchase=final_purchase_qty
    )

    log_audit(
        action='forecast_v2.run',
        status='success',
        message=f"Forecast for SKU {sku_param}",
        metadata={
            'sku': sku_param,
            'store': store_param or 'ALL',
            'horizon_weeks': horizon_weeks,
            'lead_time_weeks': lead_time_weeks,
            'safety_pct': safety_pct,
            'kpis': kpis
        }
    )

    return jsonify({
        "sku": str(sku_param),
        "product_name": product.name,
        "category": product.category or '',
        "store": store_param if store_param else "ALL",
        "history": history_data,
        "forecast": forecast_data,
        "bands": {
            "lower": bands_lower,
            "upper": bands_upper
        },
        "kpis": kpis,
        "lifecycle_status": lifecycle_status,
        "model_used": model_used,
        "buy_now": buy_now,
        "risk_level": risk_level,
        "decision": decision,
        "reason_code": '; '.join(reason_codes) if reason_codes else 'STANDARD',
        "reasons": reason_codes,
        "alerts": [{"type": a.get('type'), "severity": a.get('severity')} for a in sku_alerts],
        "slow_stock_flagged": slow_stock_flagged,
        "broken_stock_flag": broken_stock_flag,
        "explain_bullets": explain_bullets
    })


def generate_forecast_explanation(avg_weekly_demand, horizon_demand, total_stock, weeks_of_cover,
                                   lead_time_weeks, safety_pct, lifecycle_status, alerts,
                                   slow_stock_flagged, suggested_purchase):
    """Generate business-readable explanation bullets for forecast recommendation."""
    bullets = []
    
    bullets.append(f"Demanda semanal promedio: {avg_weekly_demand:.1f} unidades")
    bullets.append(f"Stock actual: {total_stock} u (~{weeks_of_cover:.1f} semanas de cobertura)")
    
    coverage_gap = lead_time_weeks + (lead_time_weeks * safety_pct) - weeks_of_cover
    if coverage_gap > 0:
        bullets.append(f"Lead time: {lead_time_weeks:.0f} semanas → riesgo de stockout durante lead time")
    else:
        bullets.append(f"Lead time: {lead_time_weeks:.0f} semanas → cobertura suficiente")
    
    for alert in alerts:
        alert_type = alert.get('type', '')
        if alert_type == 'PROJECTED_STOCKOUT':
            bullets.append("Alerta: Stockout proyectado detectado")
        elif alert_type == 'OVERSTOCK':
            bullets.append("Alerta: Sobrestock activo - compra no recomendada")
        elif alert_type == 'BROKEN_STOCK':
            bullets.append("Alerta: Quiebre de stock detectado")
        elif alert_type == 'SILENT_SKU':
            bullets.append("Alerta: SKU sin movimiento reciente")
    
    if slow_stock_flagged:
        bullets.append("SKU flaggeado como stock lento - revisar antes de comprar")
    
    if lifecycle_status == 'DEAD':
        bullets.append("Ciclo de vida: MUERTO (90+ días sin ventas) - compra bloqueada")
    elif lifecycle_status == 'SLOW':
        bullets.append("Ciclo de vida: LENTO (31-90 días) - penalización 50% aplicada")
    elif lifecycle_status == 'NEW':
        bullets.append("Ciclo de vida: NUEVO - usando promedio de categoría")
    elif lifecycle_status == 'ACTIVE':
        bullets.append("Ciclo de vida: ACTIVO - ventas en últimos 30 días")
    
    if suggested_purchase > 0:
        bullets.append(f"Compra sugerida: {suggested_purchase} unidades")
    else:
        bullets.append("Compra sugerida: 0 unidades (stock suficiente)")
    
    return bullets[:6]


@app.route('/forecast/export', methods=['GET'])
@login_required
@require_permission('forecast_v2:run')
def export_single_sku_forecast():
    """Export single SKU forecast to Excel."""
    import math
    today = date.today()
    
    sku_param = request.args.get('sku', '').strip()
    if not sku_param:
        flash('SKU requerido para exportar', 'warning')
        return redirect(url_for('purchase_forecast_v2'))
    
    store_param = request.args.get('store', '').strip()
    try:
        horizon_weeks = int(request.args.get('horizon_weeks', 8))
    except ValueError:
        horizon_weeks = 8
    try:
        lead_time_weeks = float(request.args.get('lead_time_weeks', 4))
    except ValueError:
        lead_time_weeks = 4.0
    try:
        safety_pct = float(request.args.get('safety_pct', 0.10))
    except ValueError:
        safety_pct = 0.10
    
    product = Product.query.filter_by(sku=sku_param).first()
    if not product:
        flash(f'SKU {sku_param} no encontrado', 'warning')
        return redirect(url_for('purchase_forecast_v2'))
    
    store_id_filter = None
    if store_param:
        store_obj = Store.query.filter_by(name=store_param).first()
        if store_obj:
            store_id_filter = store_obj.id
    
    history_weeks = 12
    history_cutoff = today - timedelta(days=history_weeks * 7)
    
    history_q = (
        db.session.query(db.func.sum(SalesWeeklyAgg.units).label('units'))
        .filter(SalesWeeklyAgg.product_id == product.id)
        .filter(SalesWeeklyAgg.week_start >= history_cutoff)
    )
    if store_id_filter:
        history_q = history_q.filter(SalesWeeklyAgg.store_id == store_id_filter)
    
    total_units = history_q.scalar() or 0
    avg_weekly = total_units / history_weeks if history_weeks > 0 else 0
    
    latest_cd_date = db.session.query(db.func.max(StockCD.as_of_date)).scalar() or today
    cd_stock = db.session.query(db.func.sum(StockCD.quantity)).filter(
        StockCD.product_id == product.id, StockCD.as_of_date == latest_cd_date
    ).scalar() or 0
    
    latest_store_date = db.session.query(db.func.max(StockSnapshot.as_of_date)).scalar() or today
    store_stock_q = db.session.query(db.func.sum(StockSnapshot.quantity)).filter(
        StockSnapshot.product_id == product.id, StockSnapshot.as_of_date == latest_store_date
    )
    if store_id_filter:
        store_stock_q = store_stock_q.filter(StockSnapshot.store_id == store_id_filter)
    store_stock = store_stock_q.scalar() or 0
    
    total_stock = int(cd_stock) + int(store_stock)
    
    sales_30d_map, sales_90d_map, sales_ever_map = get_lifecycle_sales_maps(store_id_filter)
    lifecycle_map = classify_sku_lifecycle(sales_30d_map, sales_90d_map, sales_ever_map, {product.id})
    lifecycle_status = lifecycle_map.get(product.id, 'UNKNOWN')
    
    all_alerts = compute_alerts()
    sku_alerts = [a for a in all_alerts if a.get('sku', '').lower() == sku_param.lower()]
    
    slow_stock_flagged = not product.eligible_for_distribution if hasattr(product, 'eligible_for_distribution') else False
    
    buy_now = False
    risk_level = 'LOW'
    reasons = []
    
    for alert in sku_alerts:
        alert_type = alert.get('type', '')
        if alert_type in ['PROJECTED_STOCKOUT', 'BROKEN_STOCK']:
            buy_now = True
            risk_level = 'HIGH'
            reasons.append(alert_type)
        elif alert_type == 'OVERSTOCK':
            reasons.append('OVERSTOCK')
    
    if slow_stock_flagged:
        reasons.append('SLOW_STOCK')
    if lifecycle_status in ['DEAD', 'SLOW']:
        reasons.append(lifecycle_status)
    
    horizon_demand = avg_weekly * horizon_weeks
    demand_lead_time = avg_weekly * lead_time_weeks
    safety_units = math.ceil(demand_lead_time * safety_pct)
    suggested_purchase = max(0, math.ceil(demand_lead_time + safety_units - total_stock))
    
    if lifecycle_status == 'DEAD' or 'OVERSTOCK' in reasons:
        suggested_purchase = 0
    
    model_used = 'SMA4'
    if lifecycle_status == 'SLOW':
        model_used = 'SMA12_PENALTY'
    elif lifecycle_status == 'NEW':
        model_used = 'CATEGORY_AVG'
    
    data = [{
        'SKU': str(sku_param),
        'Producto': product.name,
        'Categoría': product.category or '',
        'Ciclo de Vida': lifecycle_status,
        'Modelo Usado': model_used,
        'Demanda Semanal': round(avg_weekly, 2),
        'Demanda Horizonte': round(horizon_demand, 0),
        'Stock CD': int(cd_stock),
        'Stock Tiendas': int(store_stock),
        'Stock Total': total_stock,
        'Compra Sugerida': suggested_purchase,
        'Comprar Ahora': 'Sí' if buy_now else 'No',
        'Nivel Riesgo': risk_level,
        'Razones': '; '.join(reasons) if reasons else 'STANDARD',
        'Horizonte (sem)': horizon_weeks,
        'Lead Time (sem)': lead_time_weeks,
        'Safety %': f"{safety_pct*100:.0f}%",
        'Tienda': store_param or 'Todas',
        'Fecha Export': today.strftime('%Y-%m-%d')
    }]
    
    df = pd.DataFrame(data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Forecast')
    output.seek(0)
    
    filename = f"forecast_{sku_param}_{today.strftime('%Y%m%d')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/forecast/batch', methods=['POST'])
@login_required
@require_permission('forecast_v2:run')
def api_forecast_batch():
    """Run batch forecast for multiple SKUs."""
    import math
    today = date.today()
    
    try:
        horizon_weeks = int(request.form.get('horizon_weeks', 8))
    except ValueError:
        horizon_weeks = 8
    try:
        lead_time_weeks = float(request.form.get('lead_time_weeks', 4))
    except ValueError:
        lead_time_weeks = 4.0
    try:
        safety_pct = float(request.form.get('safety_pct', 0.10))
    except ValueError:
        safety_pct = 0.10
    
    sku_list = []
    
    if 'file' in request.files and request.files['file'].filename:
        file = request.files['file']
        try:
            if file.filename.endswith('.csv'):
                df = pd.read_csv(file)
            else:
                df = pd.read_excel(file)
            
            sku_col = None
            for col in df.columns:
                if col.lower() in ['sku', 'sku_code', 'codigo', 'código']:
                    sku_col = col
                    break
            if sku_col is None and len(df.columns) > 0:
                sku_col = df.columns[0]
            
            if sku_col:
                sku_list = [str(s).strip() for s in df[sku_col].dropna().tolist() if str(s).strip()]
        except Exception as e:
            return jsonify({"error": f"Error procesando archivo: {str(e)}"}), 400
    else:
        sku_text = request.form.get('sku_list', '').strip()
        if sku_text:
            sku_list = [s.strip() for s in sku_text.replace(',', '\n').split('\n') if s.strip()]
    
    if not sku_list:
        return jsonify({"error": "No SKUs provided"}), 400
    
    sku_list = list(dict.fromkeys(sku_list))[:500]
    
    products = Product.query.filter(Product.sku.in_(sku_list)).all()
    product_map = {p.sku: p for p in products}
    product_id_map = {p.id: p for p in products}
    found_skus = set(product_map.keys())
    
    sales_30d_map, sales_90d_map, sales_ever_map = get_lifecycle_sales_maps(None)
    all_product_ids = set(p.id for p in products)
    lifecycle_map = classify_sku_lifecycle(sales_30d_map, sales_90d_map, sales_ever_map, all_product_ids)
    
    all_alerts = compute_alerts()
    alerts_by_sku = {}
    for a in all_alerts:
        sku = a.get('sku', '')
        if sku not in alerts_by_sku:
            alerts_by_sku[sku] = []
        alerts_by_sku[sku].append(a)
    
    history_weeks = 12
    history_cutoff = today - timedelta(days=history_weeks * 7)
    
    sales_rows = (
        db.session.query(
            SalesWeeklyAgg.product_id,
            db.func.sum(SalesWeeklyAgg.units).label('total_units')
        )
        .filter(SalesWeeklyAgg.product_id.in_(all_product_ids))
        .filter(SalesWeeklyAgg.week_start >= history_cutoff)
        .group_by(SalesWeeklyAgg.product_id)
        .all()
    )
    sales_map = {r.product_id: r.total_units / history_weeks for r in sales_rows}
    
    latest_cd_date = db.session.query(db.func.max(StockCD.as_of_date)).scalar() or today
    cd_rows = (
        db.session.query(StockCD.product_id, db.func.sum(StockCD.quantity).label('qty'))
        .filter(StockCD.product_id.in_(all_product_ids))
        .filter(StockCD.as_of_date == latest_cd_date)
        .group_by(StockCD.product_id)
        .all()
    )
    cd_stock_map = {r.product_id: int(r.qty) for r in cd_rows}
    
    latest_store_date = db.session.query(db.func.max(StockSnapshot.as_of_date)).scalar() or today
    store_rows = (
        db.session.query(StockSnapshot.product_id, db.func.sum(StockSnapshot.quantity).label('qty'))
        .filter(StockSnapshot.product_id.in_(all_product_ids))
        .filter(StockSnapshot.as_of_date == latest_store_date)
        .group_by(StockSnapshot.product_id)
        .all()
    )
    store_stock_map = {r.product_id: int(r.qty) for r in store_rows}
    
    batch_id = str(uuid.uuid4())
    results = []
    buy_now_count = 0
    review_count = 0
    do_not_buy_count = 0
    broken_stock_count = 0
    
    for sku in sku_list:
        if sku not in found_skus:
            continue
        
        product = product_map[sku]
        pid = product.id
        
        avg_weekly = sales_map.get(pid, 0)
        cd_stock = cd_stock_map.get(pid, 0)
        store_stock = store_stock_map.get(pid, 0)
        total_stock = cd_stock + store_stock
        
        lifecycle_status = lifecycle_map.get(pid, 'UNKNOWN')
        sku_alerts = alerts_by_sku.get(sku, [])
        slow_stock_flagged = not product.eligible_for_distribution if hasattr(product, 'eligible_for_distribution') else False
        
        # Use SET for deduplicated reason codes
        reasons_set = set()
        risk_level = 'LOW'
        
        # Detect broken_stock_flag from alerts
        broken_stock_flag = False
        for alert in sku_alerts:
            alert_type = alert.get('type', '')
            alert_severity = alert.get('severity', 'LOW')
            if alert_type == 'PROJECTED_STOCKOUT':
                risk_level = 'HIGH' if alert_severity == 'HIGH' else 'MEDIUM'
                reasons_set.add('PROJECTED_STOCKOUT')
            elif alert_type == 'BROKEN_STOCK':
                broken_stock_flag = True
                if alert_severity == 'HIGH':
                    risk_level = 'HIGH'
                reasons_set.add('BROKEN_STOCK')
            elif alert_type == 'OVERSTOCK':
                reasons_set.add('OVERSTOCK')
        
        if slow_stock_flagged:
            reasons_set.add('SLOW_STOCK')
            if risk_level == 'LOW':
                risk_level = 'MEDIUM'
        
        if lifecycle_status == 'DEAD':
            reasons_set.add('DEAD')
        elif lifecycle_status == 'SLOW':
            reasons_set.add('SLOW')
        
        weeks_of_cover = total_stock / max(avg_weekly, 0.1)
        demand_lead_time = avg_weekly * lead_time_weeks
        safety_units = math.ceil(demand_lead_time * safety_pct)
        raw_purchase_qty = max(0, math.ceil(demand_lead_time + safety_units - total_stock))
        
        # Decision engine with proper precedence (per spec)
        MAX_WOC = 6.0  # Max weeks of cover threshold for overstock
        
        # Redefine overstock_flag to avoid contradiction: only if NOT broken stock
        overstock_flag = (weeks_of_cover > MAX_WOC) and (not broken_stock_flag)
        
        # Decision precedence policy:
        # A) If overstock_flag == True -> DO_NOT_BUY
        # B) Else if broken_stock_flag and raw_purchase_qty == 0 -> REVIEW (redistribute first)
        # C) Else if raw_purchase_qty > 0 -> BUY_NOW or REVIEW based on coverage
        # D) Else -> DO_NOT_BUY based on lifecycle/other rules
        
        if overstock_flag:
            decision = 'DO_NOT_BUY'
            final_purchase_qty = 0
            reasons_set.add('OVERSTOCK_BLOCK')
            do_not_buy_count += 1
        elif lifecycle_status == 'DEAD':
            decision = 'DO_NOT_BUY'
            final_purchase_qty = 0
            do_not_buy_count += 1
        elif broken_stock_flag and raw_purchase_qty == 0:
            # Stores have stockouts but total stock covers demand -> redistribute
            decision = 'REVIEW'
            final_purchase_qty = 0
            reasons_set.add('BROKEN_STOCK_REBALANCE')
            review_count += 1
        elif raw_purchase_qty > 0:
            # Need to purchase
            if broken_stock_flag:
                reasons_set.add('BROKEN_STOCK')
            reasons_set.add('DEMAND_EXCEEDS_STOCK')
            if weeks_of_cover < lead_time_weeks:
                decision = 'BUY_NOW'
                buy_now_count += 1
            else:
                decision = 'BUY_NOW' if risk_level == 'HIGH' else 'REVIEW'
                if decision == 'BUY_NOW':
                    buy_now_count += 1
                else:
                    review_count += 1
            final_purchase_qty = raw_purchase_qty
        else:
            decision = 'DO_NOT_BUY'
            final_purchase_qty = 0
            do_not_buy_count += 1
        
        # Consistency guards
        if decision == 'DO_NOT_BUY':
            final_purchase_qty = 0
        elif final_purchase_qty > 0 and decision == 'DO_NOT_BUY':
            # Force REVIEW if we have units to buy
            decision = 'REVIEW'
        
        # Convert set to list
        reasons = list(reasons_set)
        
        # Track broken stock count
        if broken_stock_flag:
            broken_stock_count += 1
        
        model_used = 'SMA4'
        if lifecycle_status == 'SLOW':
            model_used = 'SMA12_PENALTY'
        elif lifecycle_status == 'NEW':
            model_used = 'CATEGORY_AVG'
        
        explain_bullets = generate_forecast_explanation(
            avg_weekly_demand=avg_weekly,
            horizon_demand=avg_weekly * horizon_weeks,
            total_stock=total_stock,
            weeks_of_cover=weeks_of_cover,
            lead_time_weeks=lead_time_weeks,
            safety_pct=safety_pct,
            lifecycle_status=lifecycle_status,
            alerts=sku_alerts,
            slow_stock_flagged=slow_stock_flagged,
            suggested_purchase=final_purchase_qty
        )
        
        item = ForecastBatchItem(
            batch_id=batch_id,
            product_id=pid,
            sku=str(sku),
            product_name=product.name,
            category=product.category or '',
            lifecycle_status=lifecycle_status,
            model_used=model_used,
            weekly_demand=round(avg_weekly, 2),
            total_stock=total_stock,
            suggested_purchase=final_purchase_qty,
            decision=decision,
            risk_level=risk_level,
            reason_summary='; '.join(reasons[:3]) if reasons else 'STANDARD',
            explain_json=json.dumps(explain_bullets)
        )
        db.session.add(item)
        
        results.append({
            'sku': str(sku),
            'product_name': product.name,
            'category': product.category or '',
            'lifecycle_status': lifecycle_status,
            'model_used': model_used,
            'weekly_demand': round(avg_weekly, 2),
            'total_stock': total_stock,
            'suggested_purchase': final_purchase_qty,
            'decision': decision,
            'risk_level': risk_level,
            'reason_summary': '; '.join(reasons) if reasons else 'STANDARD',
            'reasons': reasons,
            'broken_stock_flag': broken_stock_flag,
            'explain_bullets': explain_bullets
        })
    
    batch_run = ForecastBatchRun(
        batch_id=batch_id,
        created_by_user_id=current_user.id,
        params_json=json.dumps({
            'horizon_weeks': horizon_weeks,
            'lead_time_weeks': lead_time_weeks,
            'safety_pct': safety_pct
        }),
        total_skus=len(results),
        buy_now_count=buy_now_count,
        review_count=review_count,
        do_not_buy_count=do_not_buy_count
    )
    db.session.add(batch_run)
    db.session.commit()
    
    log_audit(
        action='forecast_v2.batch',
        status='success',
        message=f'Batch forecast completed: {len(results)} SKUs',
        run_id=batch_id,
        metadata={
            'total_skus': len(results),
            'buy_now': buy_now_count,
            'review': review_count,
            'do_not_buy': do_not_buy_count
        }
    )
    
    return jsonify({
        'batch_id': batch_id,
        'total_skus': len(results),
        'buy_now_count': buy_now_count,
        'review_count': review_count,
        'do_not_buy_count': do_not_buy_count,
        'broken_stock_count': broken_stock_count,
        'not_found': len(sku_list) - len(results),
        'results': results
    })


@app.route('/api/forecast/batch/<batch_id>', methods=['GET'])
@login_required
@require_permission('forecast_v2:view')
def api_get_batch_forecast(batch_id):
    """Get batch forecast results with pagination."""
    page = request.args.get('page', 1, type=int)
    per_page = 10
    
    batch_run = ForecastBatchRun.query.filter_by(batch_id=batch_id).first()
    if not batch_run:
        return jsonify({"error": "Batch not found"}), 404
    
    items_q = ForecastBatchItem.query.filter_by(batch_id=batch_id).order_by(ForecastBatchItem.id)
    total = items_q.count()
    items = items_q.offset((page - 1) * per_page).limit(per_page).all()
    
    return jsonify({
        'batch_id': batch_id,
        'created_at': batch_run.created_at.isoformat(),
        'total_skus': batch_run.total_skus,
        'buy_now_count': batch_run.buy_now_count,
        'review_count': batch_run.review_count,
        'do_not_buy_count': batch_run.do_not_buy_count,
        'page': page,
        'per_page': per_page,
        'total_pages': (total + per_page - 1) // per_page,
        'results': [{
            'sku': item.sku,
            'product_name': item.product_name,
            'category': item.category,
            'lifecycle_status': item.lifecycle_status,
            'model_used': item.model_used,
            'weekly_demand': item.weekly_demand,
            'total_stock': item.total_stock,
            'suggested_purchase': item.suggested_purchase,
            'decision': item.decision,
            'risk_level': item.risk_level,
            'reason_summary': item.reason_summary,
            'explain_bullets': item.get_explain()
        } for item in items]
    })


@app.route('/forecast/batch/export/<batch_id>', methods=['GET'])
@login_required
@require_permission('forecast_v2:run')
def export_batch_forecast(batch_id):
    """Export batch forecast to Excel."""
    today = date.today()
    
    batch_run = ForecastBatchRun.query.filter_by(batch_id=batch_id).first()
    if not batch_run:
        flash('Batch no encontrado', 'warning')
        return redirect(url_for('purchase_forecast_v2'))
    
    items = ForecastBatchItem.query.filter_by(batch_id=batch_id).order_by(ForecastBatchItem.id).all()
    
    data = []
    for item in items:
        data.append({
            'SKU': str(item.sku),
            'Producto': item.product_name,
            'Categoría': item.category,
            'Ciclo de Vida': item.lifecycle_status,
            'Modelo Usado': item.model_used,
            'Demanda Semanal': item.weekly_demand,
            'Stock Total': item.total_stock,
            'Compra Sugerida': item.suggested_purchase,
            'Decisión': item.decision,
            'Nivel Riesgo': item.risk_level,
            'Razones': item.reason_summary
        })
    
    df = pd.DataFrame(data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Batch Forecast')
    output.seek(0)
    
    filename = f"batch_forecast_{batch_id[:8]}_{today.strftime('%Y%m%d')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/forecast/category', methods=['GET'])
@login_required
@require_permission('forecast_v2:view')
def api_forecast_category():
    """
    Category Forecast - Aggregated demand forecast at category level.
    Uses SalesWeeklyAgg as single source of truth.
    Returns executive-level metrics and decision.
    """
    category = request.args.get('category', '').strip()
    if not category:
        return jsonify({"error": "Category is required"}), 400
    
    try:
        horizon_weeks = int(request.args.get('horizon_weeks', 8))
    except ValueError:
        horizon_weeks = 8
    try:
        lead_time_weeks = float(request.args.get('lead_time_weeks', 4))
    except ValueError:
        lead_time_weeks = 4.0
    try:
        safety_pct = float(request.args.get('safety_pct', 0.10))
    except ValueError:
        safety_pct = 0.10
    
    store_filter = request.args.get('store', '').strip()
    store_id = None
    if store_filter:
        store_obj = Store.query.filter_by(name=store_filter).first()
        if store_obj:
            store_id = store_obj.id
    
    today = date.today()
    
    # Calculate weekly demand from SalesWeeklyAgg (last 8 weeks)
    lookback_start = today - timedelta(weeks=8)
    
    demand_q = db.session.query(
        func.sum(SalesWeeklyAgg.units).label('total_units'),
        func.count(func.distinct(SalesWeeklyAgg.week_start)).label('weeks_count')
    ).filter(
        SalesWeeklyAgg.category == category,
        SalesWeeklyAgg.week_start >= lookback_start
    )
    if store_id:
        demand_q = demand_q.filter(SalesWeeklyAgg.store_id == store_id)
    
    demand_result = demand_q.first()
    total_units_sold = demand_result.total_units or 0
    weeks_count = demand_result.weeks_count or 1
    
    weekly_demand = round(total_units_sold / max(weeks_count, 1), 1)
    horizon_demand = round(weekly_demand * horizon_weeks)
    
    # Get current stock levels (CD and Stores)
    cd_stock_q = db.session.query(
        func.sum(StockCD.quantity)
    ).join(Product, StockCD.product_id == Product.id).filter(
        Product.category == category
    )
    stock_cd = cd_stock_q.scalar() or 0
    
    store_stock_q = db.session.query(
        func.sum(StockSnapshot.quantity)
    ).join(Product, StockSnapshot.product_id == Product.id).filter(
        Product.category == category
    )
    if store_id:
        store_stock_q = store_stock_q.filter(StockSnapshot.store_id == store_id)
    stock_stores = store_stock_q.scalar() or 0
    
    total_stock = stock_cd + stock_stores
    
    # Calculate coverage and deficit
    safety_stock = round(weekly_demand * lead_time_weeks * safety_pct)
    lead_time_demand = round(weekly_demand * lead_time_weeks)
    
    # Deficit = horizon_demand + safety_stock - total_stock (if positive)
    raw_deficit = horizon_demand + safety_stock - total_stock
    deficit = max(0, raw_deficit)
    
    # Coverage in weeks
    coverage_weeks = round(total_stock / weekly_demand, 1) if weekly_demand > 0 else 999
    
    # Decision logic
    decision = "DO_NOT_BUY"
    risk_level = "LOW"
    reason_code = "ADEQUATE_STOCK"
    
    if deficit > 0:
        if coverage_weeks < lead_time_weeks:
            decision = "BUY_NOW"
            risk_level = "HIGH"
            reason_code = "COVERAGE_BELOW_LEAD_TIME"
        elif coverage_weeks < (lead_time_weeks + horizon_weeks * 0.5):
            decision = "REVIEW"
            risk_level = "MEDIUM"
            reason_code = "COVERAGE_PARTIAL"
        else:
            decision = "DO_NOT_BUY"
            risk_level = "LOW"
            reason_code = "ADEQUATE_COVERAGE"
    
    # Suggested action split (redistribute vs purchase)
    redistribute_potential = 0
    purchase_required = deficit
    
    if stock_stores > 0 and stock_cd < deficit:
        # Check store imbalance - some stores may have excess
        store_stats = db.session.query(
            func.max(StockSnapshot.quantity).label('max_qty'),
            func.min(StockSnapshot.quantity).label('min_qty'),
            func.avg(StockSnapshot.quantity).label('avg_qty')
        ).join(Product, StockSnapshot.product_id == Product.id).filter(
            Product.category == category,
            StockSnapshot.quantity > 0
        ).first()
        
        if store_stats and store_stats.max_qty and store_stats.avg_qty:
            imbalance = (store_stats.max_qty - store_stats.avg_qty)
            if imbalance > 0:
                redistribute_potential = min(int(imbalance * 0.5), deficit)
                purchase_required = max(0, deficit - redistribute_potential)
    
    # Build explanation
    explain_bullets = []
    explain_bullets.append(f"Demanda semanal promedio (8 sem): {weekly_demand:,.0f} unidades")
    explain_bullets.append(f"Demanda horizonte ({horizon_weeks} sem): {horizon_demand:,.0f} unidades")
    explain_bullets.append(f"Stock actual: {total_stock:,.0f} (CD: {stock_cd:,.0f} + Tiendas: {stock_stores:,.0f})")
    explain_bullets.append(f"Cobertura actual: {coverage_weeks:.1f} semanas")
    explain_bullets.append(f"Stock de seguridad ({int(safety_pct*100)}%): {safety_stock:,.0f} unidades")
    
    if deficit > 0:
        explain_bullets.append(f"Déficit proyectado: {deficit:,.0f} unidades")
    else:
        explain_bullets.append("Stock suficiente para el horizonte seleccionado")
    
    # Next best action
    next_action = "Mantener inventario actual"
    if decision == "BUY_NOW":
        if redistribute_potential > 0:
            next_action = f"Redistribuir primero (~{redistribute_potential:,.0f} u), luego comprar {purchase_required:,.0f} u"
        else:
            next_action = f"Comprar {purchase_required:,.0f} unidades"
    elif decision == "REVIEW":
        if redistribute_potential > 0:
            next_action = f"Evaluar redistribución (~{redistribute_potential:,.0f} u) antes de decidir compra"
        else:
            next_action = f"Revisar necesidad de compra: {deficit:,.0f} u de déficit potencial"
    
    explain_bullets.append(f"Acción recomendada: {next_action}")
    
    # SKU count in category
    sku_count = Product.query.filter_by(category=category).count()
    
    return jsonify({
        'category': category,
        'sku_count': sku_count,
        'weekly_demand': weekly_demand,
        'horizon_weeks': horizon_weeks,
        'horizon_demand': horizon_demand,
        'lead_time_weeks': lead_time_weeks,
        'safety_pct': safety_pct,
        'safety_stock': safety_stock,
        'stock_cd': stock_cd,
        'stock_stores': stock_stores,
        'total_stock': total_stock,
        'coverage_weeks': coverage_weeks,
        'deficit': deficit,
        'redistribute_potential': redistribute_potential,
        'purchase_required': purchase_required,
        'decision': decision,
        'risk_level': risk_level,
        'reason_code': reason_code,
        'next_action': next_action,
        'explain_bullets': explain_bullets
    })


@app.route('/api/forecast/category/batch', methods=['POST'])
@login_required
@require_permission('forecast_v2:run')
def api_forecast_category_batch():
    """
    Category Batch Forecast - Run forecast for multiple categories.
    """
    today = date.today()
    
    try:
        horizon_weeks = int(request.form.get('horizon_weeks', 8))
    except ValueError:
        horizon_weeks = 8
    try:
        lead_time_weeks = float(request.form.get('lead_time_weeks', 4))
    except ValueError:
        lead_time_weeks = 4.0
    try:
        safety_pct = float(request.form.get('safety_pct', 0.10))
    except ValueError:
        safety_pct = 0.10
    
    category_list = []
    
    if 'file' in request.files and request.files['file'].filename:
        file = request.files['file']
        try:
            if file.filename.endswith('.csv'):
                df = pd.read_csv(file)
            else:
                df = pd.read_excel(file)
            
            cat_col = None
            for col in df.columns:
                if col.lower() in ['category', 'categoria', 'categoría', 'cat']:
                    cat_col = col
                    break
            if cat_col is None and len(df.columns) > 0:
                cat_col = df.columns[0]
            
            if cat_col:
                category_list = [str(c).strip() for c in df[cat_col].dropna().tolist() if str(c).strip()]
        except Exception as e:
            return jsonify({"error": f"Error reading file: {str(e)}"}), 400
    
    text_input = request.form.get('categories_text', '').strip()
    if text_input:
        text_cats = [c.strip() for c in text_input.replace(',', '\n').split('\n') if c.strip()]
        category_list.extend(text_cats)
    
    category_list = list(dict.fromkeys(category_list))[:50]
    
    if not category_list:
        return jsonify({"error": "No categories provided"}), 400
    
    lookback_start = today - timedelta(weeks=8)
    
    # Prefetch aggregates for all categories in one query
    demand_data = db.session.query(
        SalesWeeklyAgg.category,
        func.sum(SalesWeeklyAgg.units).label('total_units'),
        func.count(func.distinct(SalesWeeklyAgg.week_start)).label('weeks_count')
    ).filter(
        SalesWeeklyAgg.category.in_(category_list),
        SalesWeeklyAgg.week_start >= lookback_start
    ).group_by(SalesWeeklyAgg.category).all()
    
    demand_map = {d.category: {'units': d.total_units or 0, 'weeks': d.weeks_count or 1} for d in demand_data}
    
    cd_stock_data = db.session.query(
        Product.category,
        func.sum(StockCD.quantity).label('stock')
    ).join(StockCD, Product.id == StockCD.product_id).filter(
        Product.category.in_(category_list)
    ).group_by(Product.category).all()
    
    cd_map = {d.category: d.stock or 0 for d in cd_stock_data}
    
    store_stock_data = db.session.query(
        Product.category,
        func.sum(StockSnapshot.quantity).label('stock')
    ).join(StockSnapshot, Product.id == StockSnapshot.product_id).filter(
        Product.category.in_(category_list)
    ).group_by(Product.category).all()
    
    store_map = {d.category: d.stock or 0 for d in store_stock_data}
    
    results = []
    buy_now_count = 0
    review_count = 0
    do_not_buy_count = 0
    
    for cat in category_list:
        demand_info = demand_map.get(cat, {'units': 0, 'weeks': 1})
        weekly_demand = round(demand_info['units'] / max(demand_info['weeks'], 1), 1)
        horizon_demand = round(weekly_demand * horizon_weeks)
        
        stock_cd = cd_map.get(cat, 0)
        stock_stores = store_map.get(cat, 0)
        total_stock = stock_cd + stock_stores
        
        safety_stock = round(weekly_demand * lead_time_weeks * safety_pct)
        raw_deficit = horizon_demand + safety_stock - total_stock
        deficit = max(0, raw_deficit)
        
        coverage_weeks = round(total_stock / weekly_demand, 1) if weekly_demand > 0 else 999
        
        decision = "DO_NOT_BUY"
        reason = "Stock adecuado"
        
        if deficit > 0:
            if coverage_weeks < lead_time_weeks:
                decision = "BUY_NOW"
                reason = "Cobertura bajo lead time"
                buy_now_count += 1
            elif coverage_weeks < (lead_time_weeks + horizon_weeks * 0.5):
                decision = "REVIEW"
                reason = "Cobertura parcial"
                review_count += 1
            else:
                do_not_buy_count += 1
        else:
            do_not_buy_count += 1
        
        results.append({
            'category': cat,
            'weekly_demand': weekly_demand,
            'total_stock': total_stock,
            'deficit': deficit,
            'coverage_weeks': coverage_weeks,
            'decision': decision,
            'reason': reason
        })
    
    return jsonify({
        'total': len(results),
        'buy_now_count': buy_now_count,
        'review_count': review_count,
        'do_not_buy_count': do_not_buy_count,
        'results': results
    })


@app.route('/export_cd_remanente', methods=['GET'])
@login_required
@require_permission('distribution:export')
def export_cd_remanente():
    """Exporta a Excel el remanente de CD solo de los SKUs usados en una corrida específica."""
    today = date.today()
    
    run_id = request.args.get("run_id", "").strip()
    if not run_id:
        latest_run = (
            PredictionRun.query
            .filter(PredictionRun.run_id.isnot(None))
            .filter(PredictionRun.run_id != "")
            .filter(PredictionRun.run_type == 'distribution')
            .order_by(PredictionRun.created_at.desc())
            .first()
        )
        run_id = latest_run.run_id if latest_run else None

    if not run_id:
        df_empty = pd.DataFrame([{
            "SKU": "",
            "Producto": "",
            "Stock CD disponible": "",
            "Corrida": "",
            "Fecha": today.strftime("%Y-%m-%d"),
        }])
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df_empty.to_excel(writer, index=False, sheet_name="Remanente")
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name=f"remanente_cd_{today.strftime('%Y%m%d')}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    preds = (
        db.session.query(Prediction.product_id)
        .filter(Prediction.run_id == run_id)
        .distinct()
        .all()
    )
    product_ids = [p.product_id for p in preds]

    cd_rows = (
        db.session.query(StockCD, Product)
        .join(Product, StockCD.product_id == Product.id)
        .filter(
            StockCD.as_of_date == today,
            StockCD.product_id.in_(product_ids)
        )
        .order_by(Product.sku.asc())
        .all()
    )

    data = []
    for cd, prod in cd_rows:
        data.append({
            "SKU": prod.sku,
            "Producto": prod.name,
            "Stock CD disponible": cd.quantity,
            "Corrida": run_id[:8],
            "Fecha": today.strftime("%Y-%m-%d"),
        })

    if not data:
        data.append({
            "SKU": "",
            "Producto": "",
            "Stock CD disponible": "",
            "Corrida": run_id[:8] if run_id else "",
            "Fecha": today.strftime("%Y-%m-%d"),
        })

    df = pd.DataFrame(data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Remanente")
    output.seek(0)

    run_id_short = run_id[:8] if run_id else "norun"
    return send_file(
        output,
        as_attachment=True,
        download_name=f"FastDSO_CD_Remainder_{run_id_short}_{today.strftime('%Y%m%d')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def process_macro_sales_upload(job_id, payload):
    """Background task to process macro sales upload - aggregates to weekly data."""
    import os
    
    filepath = payload.get('filepath')
    load_mode = payload.get('load_mode', 'replace_range')
    user_id = payload.get('user_id')
    
    session = get_isolated_session()
    try:
        update_job_status(job_id, progress=5, message='Leyendo archivo...')
        
        if filepath.endswith('.csv'):
            df = pd.read_csv(filepath, dtype=str)
        else:
            df = pd.read_excel(filepath, dtype=str)
        
        total_rows = len(df)
        update_job_status(job_id, progress=10, message=f'Procesando {total_rows} filas...')
        
        df.columns = [str(c).strip().lower() for c in df.columns]
        
        needed = {'sku', 'store', 'date', 'quantity'}
        if not needed.issubset(set(df.columns)):
            update_job_status(job_id, status='error', progress=100, 
                            message='Archivo debe tener columnas: sku, store, date, quantity')
            return
        
        df['sku'] = df['sku'].astype(str).str.strip()
        df['store'] = df['store'].astype(str).str.strip()
        df['quantity'] = pd.to_numeric(df['quantity'], errors='coerce').fillna(0).astype(int)
        df['date'] = pd.to_datetime(df['date'], errors='coerce')
        
        has_product_name = 'product_name' in df.columns
        has_category = 'category' in df.columns
        has_duv = 'duv' in df.columns
        has_duc = 'duc' in df.columns
        
        if has_product_name:
            df['product_name'] = df['product_name'].astype(str).str.strip()
        if has_category:
            df['category'] = df['category'].astype(str).str.strip()
            df['category'] = df['category'].replace(['', 'nan', 'None'], None)
        
        duv_map = {}
        duc_map = {}
        if has_duv or has_duc:
            if has_duv:
                df['duv_parsed'] = pd.to_numeric(df['duv'], errors='coerce')
                df.loc[df['duv_parsed'] < 0, 'duv_parsed'] = None
            if has_duc:
                df['duc_parsed'] = pd.to_numeric(df['duc'], errors='coerce')
                df.loc[df['duc_parsed'] < 0, 'duc_parsed'] = None
            
            for sku in df['sku'].unique():
                sku_rows = df[df['sku'] == sku]
                if has_duv:
                    duv_vals = sku_rows['duv_parsed'].dropna()
                    if len(duv_vals) > 0:
                        duv_map[sku] = int(duv_vals.iloc[0])
                if has_duc:
                    duc_vals = sku_rows['duc_parsed'].dropna()
                    if len(duc_vals) > 0:
                        duc_map[sku] = int(duc_vals.iloc[0])
        
        df = df.dropna(subset=['date'])
        df['week_start'] = df['date'].apply(lambda d: d - timedelta(days=d.weekday()))
        
        update_job_status(job_id, progress=20, message='Preparando productos y tiendas...')
        
        existing_products = {p.sku: p for p in session.query(Product).all()}
        existing_stores = {s.name: s for s in session.query(Store).all()}
        
        new_products = []
        new_stores = []
        
        unique_skus = df['sku'].unique()
        for sku in unique_skus:
            if sku not in existing_products:
                row = df[df['sku'] == sku].iloc[0]
                name = row.get('product_name', sku) if has_product_name else sku
                cat = row.get('category') if has_category else None
                new_products.append({'sku': sku, 'name': name, 'category': cat})
        
        unique_stores = df['store'].unique()
        for store_name in unique_stores:
            if store_name not in existing_stores:
                new_stores.append({'name': store_name})
        
        if new_products:
            session.execute(Product.__table__.insert(), new_products)
            session.commit()
            existing_products = {p.sku: p for p in session.query(Product).all()}
        
        if new_stores:
            session.execute(Store.__table__.insert(), new_stores)
            session.commit()
            existing_stores = {s.name: s for s in session.query(Store).all()}
        
        update_job_status(job_id, progress=35, message='Actualizando categorías de productos...')
        
        if has_category:
            category_updates = 0
            for sku in unique_skus:
                prod = existing_products.get(sku)
                if prod and not prod.category:
                    cat_val = df[df['sku'] == sku]['category'].dropna().unique()
                    if len(cat_val) > 0 and cat_val[0]:
                        session.query(Product).filter(Product.id == prod.id).update({'category': cat_val[0]})
                        category_updates += 1
            if category_updates > 0:
                session.commit()
        
        update_job_status(job_id, progress=50, message='Agregando ventas por semana...')
        
        df['product_id'] = df['sku'].map(lambda x: existing_products[x].id)
        df['store_id'] = df['store'].map(lambda x: existing_stores[x].id)
        
        agg_df = df.groupby(['product_id', 'store_id', 'week_start']).agg({
            'quantity': 'sum'
        }).reset_index()
        agg_df.rename(columns={'quantity': 'units'}, inplace=True)
        
        if has_category:
            cat_map = {}
            for sku, prod in existing_products.items():
                cat_map[prod.id] = prod.category
            agg_df['category'] = agg_df['product_id'].map(cat_map)
        else:
            agg_df['category'] = None
        
        update_job_status(job_id, progress=65, message=f'Guardando {len(agg_df)} registros semanales...')
        
        weeks_in_upload = set(agg_df['week_start'].unique())
        
        if load_mode == 'replace_range' and weeks_in_upload:
            session.query(SalesWeeklyAgg).filter(
                SalesWeeklyAgg.week_start.in_(weeks_in_upload)
            ).delete(synchronize_session=False)
            session.commit()
        
        chunk_size = 1000
        records = agg_df.to_dict('records')
        total_inserted = 0
        
        for i in range(0, len(records), chunk_size):
            chunk = records[i:i+chunk_size]
            
            if load_mode == 'append_range':
                for rec in chunk:
                    existing = session.query(SalesWeeklyAgg).filter(
                        SalesWeeklyAgg.product_id == rec['product_id'],
                        SalesWeeklyAgg.store_id == rec['store_id'],
                        SalesWeeklyAgg.week_start == rec['week_start']
                    ).first()
                    if existing:
                        existing.units += rec['units']
                    else:
                        session.add(SalesWeeklyAgg(**rec))
            else:
                session.execute(SalesWeeklyAgg.__table__.insert(), chunk)
            
            session.commit()
            total_inserted += len(chunk)
            progress = 65 + int(30 * (i + chunk_size) / len(records))
            update_job_status(job_id, progress=min(progress, 95), 
                            message=f'Insertando... {total_inserted}/{len(records)}')
        
        update_job_status(job_id, progress=96, message='Actualizando lifecycle SKUs...')
        
        lifecycle_store_df = df.groupby(['product_id', 'store_id']).agg({
            'date': 'max'
        }).reset_index()
        lifecycle_store_df.rename(columns={'date': 'last_sale_date_store'}, inplace=True)
        
        lifecycle_global_df = df.groupby(['product_id']).agg({
            'date': 'max'
        }).reset_index()
        lifecycle_global_df.rename(columns={'date': 'last_sale_date_global'}, inplace=True)
        
        for _, row in lifecycle_store_df.iterrows():
            pid = int(row['product_id'])
            sid = int(row['store_id'])
            raw_date = row['last_sale_date_store']
            if pd.isna(raw_date):
                continue
            last_date = raw_date.date() if hasattr(raw_date, 'date') else raw_date
            if not last_date:
                continue
            
            existing = session.query(SkuStoreLifecycle).filter(
                SkuStoreLifecycle.product_id == pid,
                SkuStoreLifecycle.store_id == sid
            ).first()
            
            if existing:
                if existing.last_sale_date_store is None or last_date > existing.last_sale_date_store:
                    existing.last_sale_date_store = last_date
                    existing.updated_at = datetime.utcnow()
            else:
                session.add(SkuStoreLifecycle(
                    product_id=pid,
                    store_id=sid,
                    last_sale_date_store=last_date,
                    updated_at=datetime.utcnow()
                ))
        
        session.commit()
        
        sku_to_pid = {sku: prod.id for sku, prod in existing_products.items()}
        pid_to_sku = {prod.id: sku for sku, prod in existing_products.items()}
        
        for _, row in lifecycle_global_df.iterrows():
            pid = int(row['product_id'])
            raw_date = row['last_sale_date_global']
            sku = pid_to_sku.get(pid)
            
            last_date = None
            if not pd.isna(raw_date):
                last_date = raw_date.date() if hasattr(raw_date, 'date') else raw_date
            
            duv_val = duv_map.get(sku) if sku else None
            duc_val = duc_map.get(sku) if sku else None
            
            if not last_date and duv_val is None and duc_val is None:
                continue
            
            existing = session.query(SkuLifecycle).filter(
                SkuLifecycle.product_id == pid
            ).first()
            
            if existing:
                if last_date and (existing.last_sale_date_global is None or last_date > existing.last_sale_date_global):
                    existing.last_sale_date_global = last_date
                if duv_val is not None:
                    existing.days_since_last_sale_global = duv_val
                if duc_val is not None:
                    existing.days_since_last_purchase_global = duc_val
                existing.updated_at = datetime.utcnow()
            else:
                session.add(SkuLifecycle(
                    product_id=pid,
                    last_sale_date_global=last_date,
                    days_since_last_sale_global=duv_val,
                    days_since_last_purchase_global=duc_val,
                    updated_at=datetime.utcnow()
                ))
        
        session.commit()
        
        processed_pids = set(lifecycle_global_df['product_id'].unique())
        for sku, duv_val in duv_map.items():
            pid = sku_to_pid.get(sku)
            if not pid or pid in processed_pids:
                continue
            duc_val = duc_map.get(sku)
            existing = session.query(SkuLifecycle).filter(SkuLifecycle.product_id == pid).first()
            if existing:
                existing.days_since_last_sale_global = duv_val
                if duc_val is not None:
                    existing.days_since_last_purchase_global = duc_val
                existing.updated_at = datetime.utcnow()
            else:
                session.add(SkuLifecycle(
                    product_id=pid,
                    days_since_last_sale_global=duv_val,
                    days_since_last_purchase_global=duc_val,
                    updated_at=datetime.utcnow()
                ))
            processed_pids.add(pid)
        
        for sku, duc_val in duc_map.items():
            pid = sku_to_pid.get(sku)
            if not pid or pid in processed_pids:
                continue
            existing = session.query(SkuLifecycle).filter(SkuLifecycle.product_id == pid).first()
            if existing:
                existing.days_since_last_purchase_global = duc_val
                existing.updated_at = datetime.utcnow()
            else:
                session.add(SkuLifecycle(
                    product_id=pid,
                    days_since_last_purchase_global=duc_val,
                    updated_at=datetime.utcnow()
                ))
        
        session.commit()
        
        try:
            os.remove(filepath)
        except:
            pass
        
        session.close()
        
        duv_count = len(duv_map)
        duc_count = len(duc_map)
        msg = f'Ventas macro cargadas: {len(agg_df)} registros semanales de {total_rows} filas originales (lifecycle actualizado'
        if duv_count or duc_count:
            msg += f', DUV: {duv_count}, DUC: {duc_count}'
        msg += ')'
        update_job_status(
            job_id,
            status='done',
            progress=100,
            message=msg,
            result={
                'records_created': len(agg_df),
                'total_rows': total_rows,
                'weeks': len(weeks_in_upload),
                'new_products': len(new_products),
                'new_stores': len(new_stores)
            }
        )
        
    except Exception as e:
        session.rollback()
        session.close()
        update_job_status(job_id, status='error', progress=100, message=f'Error: {str(e)[:400]}')
        raise


def process_sales_upload(job_id, payload):
    """Background task to process Sales upload with bulk operations and prediction generation. Uses isolated session."""
    import os
    from uuid import uuid4
    
    filepath = payload.get('filepath')
    analysis_mode = payload.get('analysis_mode', 'sma3_min3')
    meta = payload.get('meta', {})
    user_id = payload.get('user_id')
    original_filename = payload.get('original_filename', 'unknown')
    
    session = get_isolated_session()
    try:
        update_job_status(job_id, progress=10, message='Leyendo archivo...')
        
        if filepath.endswith('.csv'):
            df = pd.read_csv(filepath, dtype=str)
        else:
            df = pd.read_excel(filepath, dtype=str)
        
        total_rows = len(df)
        update_job_status(job_id, progress=15, message=f'Procesando {total_rows} filas...')
        
        df.columns = [str(c).strip().lower() for c in df.columns]
        needed = {'sku', 'product_name', 'store', 'quantity', 'date'}
        if not needed.issubset(set(df.columns)):
            update_job_status(job_id, status='error', progress=100, 
                            message='Archivo debe tener columnas: sku, product_name, store, quantity, date')
            return
        
        df['sku'] = df['sku'].astype(str).str.strip()
        df['product_name'] = df['product_name'].astype(str).str.strip()
        df['store'] = df['store'].astype(str).str.strip()
        df['quantity'] = pd.to_numeric(df['quantity'], errors='coerce').fillna(0).astype(int)
        df['date'] = pd.to_datetime(df['date'], errors='coerce')
        
        # Optional category column
        has_category = 'category' in df.columns
        if has_category:
            df['category'] = df['category'].astype(str).str.strip()
            df['category'] = df['category'].replace(['', 'nan', 'None', 'NaN'], None)
        
        sales_run_id = str(uuid.uuid4())
        sales_run = Run(
            run_id=sales_run_id,
            run_type='sales_upload',
            user_id=user_id,
            folio=meta.get("folio"),
            responsable=meta.get("responsable"),
            categoria=meta.get("categoria"),
            notes=meta.get("fecha_doc"),
            status='completed',
            rows_count=total_rows
        )
        session.add(sales_run)
        session.commit()
        
        update_job_status(job_id, progress=25, message='Cargando productos y tiendas existentes...')
        
        unique_skus = list(df['sku'].unique())
        unique_store_names = list(df['store'].unique())
        
        existing_products = {}
        batch_size = 500
        for i in range(0, len(unique_skus), batch_size):
            sku_batch = unique_skus[i:i + batch_size]
            for p in session.query(Product).filter(Product.sku.in_(sku_batch)).all():
                existing_products[p.sku] = p.id
        
        existing_stores = {}
        for i in range(0, len(unique_store_names), batch_size):
            store_batch = unique_store_names[i:i + batch_size]
            for s in session.query(Store).filter(Store.name.in_(store_batch)).all():
                existing_stores[s.name] = s.id
        
        # Build unique SKU dataframe with optional category
        cols_for_unique = ['sku', 'product_name']
        if has_category:
            cols_for_unique.append('category')
        unique_sku_df = df[cols_for_unique].drop_duplicates('sku')
        
        # Prepare category map from upload data for existing product updates
        sku_category_map = {}
        if has_category:
            for _, row in unique_sku_df.iterrows():
                cat_val = row.get('category')
                if cat_val and cat_val not in [None, 'None', 'nan', '']:
                    sku_category_map[row['sku']] = cat_val
        
        new_products = []
        for _, row in unique_sku_df.iterrows():
            if row['sku'] not in existing_products:
                prod_data = {'sku': row['sku'], 'name': row['product_name']}
                if has_category:
                    cat_val = row.get('category')
                    if cat_val and cat_val not in [None, 'None', 'nan', '']:
                        prod_data['category'] = cat_val
                new_products.append(prod_data)
        
        if new_products:
            update_job_status(job_id, progress=30, message=f'Creando {len(new_products)} productos nuevos...')
            session.execute(Product.__table__.insert(), new_products)
            session.commit()
            new_skus = [np['sku'] for np in new_products]
            for i in range(0, len(new_skus), batch_size):
                sku_batch = new_skus[i:i + batch_size]
                for p in session.query(Product).filter(Product.sku.in_(sku_batch)).all():
                    existing_products[p.sku] = p.id
        
        # Update category for existing products with NULL category
        if has_category and sku_category_map:
            category_updates = 0
            category_mismatches = 0
            existing_skus_to_check = [sku for sku in sku_category_map.keys() if sku in existing_products]
            for i in range(0, len(existing_skus_to_check), batch_size):
                sku_batch = existing_skus_to_check[i:i + batch_size]
                for prod in session.query(Product).filter(Product.sku.in_(sku_batch)).all():
                    incoming_cat = sku_category_map.get(prod.sku)
                    if incoming_cat:
                        if not prod.category:
                            prod.category = incoming_cat
                            category_updates += 1
                        elif prod.category != incoming_cat:
                            print(f"Category mismatch for SKU {prod.sku}: existing={prod.category}, incoming={incoming_cat}")
                            category_mismatches += 1
            if category_updates > 0:
                session.commit()
                print(f"Updated category for {category_updates} existing products, {category_mismatches} mismatches logged")
        
        new_stores = [{'name': s} for s in unique_store_names if s not in existing_stores]
        
        if new_stores:
            update_job_status(job_id, progress=35, message=f'Creando {len(new_stores)} tiendas nuevas...')
            session.execute(Store.__table__.insert(), new_stores)
            session.commit()
            new_names = [ns['name'] for ns in new_stores]
            for i in range(0, len(new_names), batch_size):
                name_batch = new_names[i:i + batch_size]
                for s in session.query(Store).filter(Store.name.in_(name_batch)).all():
                    existing_stores[s.name] = s.id
        
        update_job_status(job_id, progress=45, message='Preparando registros de ventas...')
        
        dist_records = []
        for _, row in df.iterrows():
            pid = existing_products.get(row['sku'])
            sid = existing_stores.get(row['store'])
            if pid and sid and pd.notna(row['date']):
                dist_records.append({
                    'product_id': pid,
                    'store_id': sid,
                    'quantity': int(row['quantity']),
                    'event_date': row['date'].date(),
                    'run_id': sales_run_id
                })
        
        update_job_status(job_id, progress=55, message=f'Insertando {len(dist_records)} registros...')
        
        insert_batch = 1000
        for i in range(0, len(dist_records), insert_batch):
            batch = dist_records[i:i + insert_batch]
            session.execute(DistributionRecord.__table__.insert(), batch)
            progress = 55 + int(20 * (i + len(batch)) / len(dist_records))
            update_job_status(job_id, progress=progress, message=f'Insertando {i + len(batch)}/{len(dist_records)}...')
        
        session.commit()
        session.close()
        
        update_job_status(job_id, progress=80, message='Generando predicciones...')
        
        run_id, n_preds, _, _, _ = generate_predictions(mode=analysis_mode, meta=meta, df=df, sales_run_id=sales_run_id)
        
        update_job_status(job_id, progress=95, message='Finalizando...')
        
        try:
            os.remove(filepath)
        except:
            pass
        
        msg = f'Ventas cargadas: {len(dist_records)} registros. Predicciones: {n_preds}'
        update_job_status(
            job_id,
            status='done',
            progress=100,
            message=msg,
            result={
                'records_created': len(dist_records),
                'predictions': n_preds,
                'total_rows': total_rows,
                'sales_run_id': sales_run_id,
                'prediction_run_id': run_id
            }
        )
        
    except Exception as e:
        session.rollback()
        session.close()
        update_job_status(job_id, status='error', progress=100, message=f'Error: {str(e)[:400]}')
        raise


@app.route('/upload', methods=['GET', 'POST'])
@login_required
@require_permission('sales:upload')
def upload():
    import os
    
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    
    has_stock = db.session.query(StockSnapshot.id).first() is not None
    require_stock_confirm = not has_stock

    if request.method == 'POST':
        if require_stock_confirm and not request.form.get('confirm_no_stock'):
            if is_ajax:
                return jsonify({'ok': False, 'message': 'No hay stock de tienda cargado. Confirma que quieres continuar.', 'category': 'warning'}), 400
            flash('No hay stock de tienda cargado. Confirma que quieres continuar.', 'warning')
            return redirect(url_for('upload'))

        file = request.files.get('file')
        if not file:
            if is_ajax:
                return jsonify({'ok': False, 'message': 'Sube un archivo CSV o Excel.', 'category': 'warning'}), 400
            flash('Sube un archivo CSV o Excel.', 'warning')
            return redirect(url_for('upload'))

        filename = file.filename.lower()
        if not (filename.endswith('.csv') or filename.endswith('.xlsx') or filename.endswith('.xls')):
            if is_ajax:
                return jsonify({'ok': False, 'message': 'Formato no soportado. Usa .csv o .xlsx', 'category': 'danger'}), 400
            flash('Formato no soportado. Usa .csv o .xlsx', 'danger')
            return redirect(url_for('upload'))
        
        try:
            if filename.endswith('.csv'):
                df_preview = pd.read_csv(file, dtype=str, nrows=5)
                file.seek(0)
            else:
                df_preview = pd.read_excel(file, dtype=str, nrows=5)
                file.seek(0)
        except Exception as e:
            if is_ajax:
                return jsonify({'ok': False, 'message': f'Error leyendo archivo: {str(e)[:200]}', 'category': 'danger'}), 400
            flash(f'Error leyendo archivo: {str(e)[:200]}', 'danger')
            return redirect(url_for('upload'))
        
        df_preview.columns = [str(c).strip().lower() for c in df_preview.columns]
        cols = set(df_preview.columns)
        
        sales_columns = {'store', 'date', 'quantity'}
        has_sales_columns = sales_columns.issubset(cols)
        
        if has_sales_columns:
            if is_ajax:
                return jsonify({
                    'ok': False, 
                    'message': 'Este módulo no ingesta ventas. Para cargar historial de ventas, usa "Ventas macro" en el menú Carga de archivos.', 
                    'category': 'warning'
                }), 400
            flash('Este módulo no ingesta ventas. Para cargar historial de ventas, usa "Ventas macro" en el menú Carga de archivos.', 'warning')
            return redirect(url_for('upload'))
        
        if 'sku' not in cols:
            if is_ajax:
                return jsonify({'ok': False, 'message': 'Archivo debe contener columna "sku".', 'category': 'danger'}), 400
            flash('Archivo debe contener columna "sku".', 'danger')
            return redirect(url_for('upload'))
        
        upload_dir = os.path.join(BASE_DIR, 'instance', 'uploads')
        os.makedirs(upload_dir, exist_ok=True)
        
        saved_filename = f"dist_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}_{file.filename}"
        filepath = os.path.join(upload_dir, saved_filename)
        file.save(filepath)
        
        analysis_mode = request.form.get('analysis_mode', 'sma3_min3')
        meta = {
            "folio": request.form.get('folio', '').strip() or None,
            "responsable": request.form.get('responsable', '').strip() or None,
            "categoria": request.form.get('categoria', '').strip() or None,
            "fecha_doc": request.form.get('fecha_doc', '').strip() or None,
        }
        user_id = current_user.id if current_user.is_authenticated else None
        
        # Request-file-only mode: always process as SKU list
        try:
            if filepath.endswith('.csv'):
                df_full = pd.read_csv(filepath, dtype=str)
            else:
                df_full = pd.read_excel(filepath, dtype=str)
            
            df_full.columns = [str(c).strip().lower() for c in df_full.columns]
            
            df_full['sku'] = df_full['sku'].astype(str).str.strip()
            df_full = df_full[df_full['sku'].notna() & (df_full['sku'] != '') & (df_full['sku'] != 'nan')]
            
            if len(df_full) == 0:
                if is_ajax:
                    return jsonify({'ok': False, 'message': 'Archivo no contiene SKUs válidos.', 'category': 'warning'}), 400
                flash('Archivo no contiene SKUs válidos.', 'warning')
                return redirect(url_for('upload'))
            
            has_name = 'product_name' in df_full.columns
            has_category = 'category' in df_full.columns
            
            upserted_count = 0
            for _, row in df_full.iterrows():
                sku_val = row['sku']
                product = Product.query.filter_by(sku=sku_val).first()
                
                if product:
                    if has_name and pd.notna(row.get('product_name')) and str(row['product_name']).strip():
                        if not product.name or product.name == sku_val:
                            product.name = str(row['product_name']).strip()
                    if has_category and pd.notna(row.get('category')) and str(row['category']).strip():
                        cat_val = str(row['category']).strip()
                        if cat_val not in ['', 'nan', 'None', 'NaN'] and not product.category:
                            product.category = cat_val
                else:
                    new_prod = Product(sku=sku_val)
                    if has_name and pd.notna(row.get('product_name')) and str(row['product_name']).strip():
                        new_prod.name = str(row['product_name']).strip()
                    else:
                        new_prod.name = sku_val
                    if has_category and pd.notna(row.get('category')) and str(row['category']).strip():
                        cat_val = str(row['category']).strip()
                        if cat_val not in ['', 'nan', 'None', 'NaN']:
                            new_prod.category = cat_val
                    db.session.add(new_prod)
                upserted_count += 1
            
            db.session.commit()
            
            sku_list = df_full['sku'].unique().tolist()
            
            if not sku_list:
                if is_ajax:
                    return jsonify({'ok': False, 'message': 'No se encontraron SKUs válidos en el archivo.', 'category': 'warning'}), 400
                flash('No se encontraron SKUs válidos en el archivo.', 'warning')
                return redirect(url_for('upload'))
            
            macro_count = db.session.query(func.count(SalesWeeklyAgg.id)).scalar() or 0
            if macro_count == 0:
                session['distribution_request_skus'] = sku_list
                session['distribution_request_mode'] = analysis_mode
                session['distribution_request_meta'] = meta
                
                msg = f'Archivo cargado: {len(sku_list)} SKUs. No hay ventas macro cargadas. Carga ventas macro primero.'
                
                if is_ajax:
                    return jsonify({
                        'ok': True,
                        'message': msg,
                        'category': 'warning',
                        'redirect_url': url_for('upload')
                    })
                flash(msg, 'warning')
                return redirect(url_for('upload'))
            
            run_id, n_preds, total_units = generate_predictions_from_macro(
                scope_skus=sku_list,
                mode=analysis_mode,
                meta=meta,
                user_id=user_id
            )
            
            session.pop('distribution_request_skus', None)
            session.pop('distribution_request_mode', None)
            session.pop('distribution_request_meta', None)
            
            try:
                os.remove(filepath)
            except:
                pass
            
            msg = f'Distribución generada: {n_preds} predicciones para {len(sku_list)} SKUs ({total_units:,} unidades)'
            
            if is_ajax:
                return jsonify({
                    'ok': True,
                    'message': msg,
                    'category': 'success',
                    'redirect_url': url_for('dashboard')
                })
            flash(msg, 'success')
            return redirect(url_for('dashboard'))
            
        except Exception as e:
            db.session.rollback()
            try:
                os.remove(filepath)
            except:
                pass
            if is_ajax:
                return jsonify({'ok': False, 'message': f'Error procesando archivo: {str(e)[:200]}', 'category': 'danger'}), 400
            flash(f'Error procesando archivo: {str(e)[:200]}', 'danger')
            return redirect(url_for('upload'))

    pending_skus = session.get('distribution_request_skus', [])
    return render_template('upload.html', require_stock_confirm=require_stock_confirm, pending_skus=pending_skus)


@app.route('/sales_macro_upload', methods=['GET', 'POST'])
@login_required
@require_permission('sales_macro:upload')
def sales_macro_upload():
    """Upload full catalog sales data for macro layer visibility and cold-start."""
    import os
    
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    
    stats = {}
    agg_count = db.session.query(func.count(SalesWeeklyAgg.id)).scalar() or 0
    if agg_count > 0:
        stats['total_skus'] = db.session.query(func.count(func.distinct(SalesWeeklyAgg.product_id))).scalar() or 0
        stats['total_stores'] = db.session.query(func.count(func.distinct(SalesWeeklyAgg.store_id))).scalar() or 0
        stats['total_weeks'] = db.session.query(func.count(func.distinct(SalesWeeklyAgg.week_start))).scalar() or 0
        stats['total_units'] = db.session.query(func.sum(SalesWeeklyAgg.units)).scalar() or 0
        
        min_week = db.session.query(func.min(SalesWeeklyAgg.week_start)).scalar()
        max_week = db.session.query(func.max(SalesWeeklyAgg.week_start)).scalar()
        if min_week and max_week:
            stats['date_range'] = f"{min_week.strftime('%d/%m/%Y')} - {max_week.strftime('%d/%m/%Y')}"
    
    if request.method == 'POST':
        file = request.files.get('file')
        if not file:
            if is_ajax:
                return jsonify({'ok': False, 'message': 'Sube un archivo CSV o Excel.', 'category': 'warning'}), 400
            flash('Sube un archivo CSV o Excel.', 'warning')
            return redirect(url_for('sales_macro_upload'))

        filename = file.filename.lower()
        if not (filename.endswith('.csv') or filename.endswith('.xlsx') or filename.endswith('.xls')):
            if is_ajax:
                return jsonify({'ok': False, 'message': 'Formato no soportado. Usa .csv o .xlsx', 'category': 'danger'}), 400
            flash('Formato no soportado. Usa .csv o .xlsx', 'danger')
            return redirect(url_for('sales_macro_upload'))
        
        upload_dir = os.path.join(BASE_DIR, 'instance', 'uploads')
        os.makedirs(upload_dir, exist_ok=True)
        
        saved_filename = f"macro_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}_{file.filename}"
        filepath = os.path.join(upload_dir, saved_filename)
        file.save(filepath)
        
        load_mode = request.form.get('load_mode', 'replace_range')
        user_id = current_user.id if current_user.is_authenticated else None
        
        payload = {
            'filepath': filepath,
            'load_mode': load_mode,
            'user_id': user_id,
            'original_filename': file.filename
        }
        
        job_id = create_and_run_job(
            job_type='upload_macro_sales',
            task_func=process_macro_sales_upload,
            payload=payload,
            user_id=user_id
        )
        
        if is_ajax:
            return jsonify({
                'ok': True,
                'job_id': job_id,
                'redirect_url': url_for('sales_macro_upload'),
                'message': 'Archivo recibido. Procesando ventas macro...',
                'category': 'success'
            })
        
        return redirect(url_for('job_status_view', job_id=job_id))

    return render_template('upload_sales_macro.html', stats=stats if stats else None)


@app.route('/stock', methods=['GET', 'POST'])
@login_required
@require_permission('stock_store:upload')
def upload_stock():
    if request.method == 'POST':
        f = request.files.get('file')
        if not f:
            flash('Sube un archivo CSV o Excel', 'warning')
            return redirect(url_for('upload_stock'))

        filename = f.filename.lower()
        try:
            if filename.endswith('.csv'):
                df = pd.read_csv(f, dtype=str)
            elif filename.endswith('.xlsx') or filename.endswith('.xls'):
                df = pd.read_excel(f, dtype=str)
            else:
                flash('Formato no soportado. Usa .csv o .xlsx', 'danger')
                return redirect(url_for('upload_stock'))
        except Exception as e:
            flash(f'Error leyendo archivo de stock: {e}', 'danger')
            return redirect(url_for('upload_stock'))

        # normalizar encabezados
        df.columns = [str(c).strip() for c in df.columns]
        cols_lower = {c.lower(): c for c in df.columns}

        sku_col = cols_lower.get('sku') or cols_lower.get('codigo')
        prod_col = cols_lower.get('producto') or cols_lower.get('product') or cols_lower.get('product_name') or cols_lower.get('nombre')
        cat_col = cols_lower.get('category') or cols_lower.get('categoria')

        if not sku_col:
            flash('El archivo debe tener columna "SKU" o "Codigo".', 'danger')
            return redirect(url_for('upload_stock'))

        # columnas de tiendas = todo lo que no es SKU, Producto, ni Category
        excluded_cols = [sku_col, prod_col, cat_col] if cat_col else [sku_col, prod_col]
        excluded_cols = [c for c in excluded_cols if c]  # filter None
        store_cols = [c for c in df.columns if c not in excluded_cols]
        if not store_cols:
            flash('No se encontraron columnas de tiendas.', 'danger')
            return redirect(url_for('upload_stock'))

        # limpieza básica
        df[sku_col] = df[sku_col].astype('string').str.strip()
        if prod_col:
            df[prod_col] = df[prod_col].astype('string').str.strip()
        
        # Optional category column normalization
        if cat_col:
            df[cat_col] = df[cat_col].astype('string').str.strip()
            df[cat_col] = df[cat_col].replace(['', 'nan', 'None', 'NaN', '<NA>'], None)

        # pasar columnas de tiendas a int
        for sc in store_cols:
            df[sc] = pd.to_numeric(df[sc], errors='coerce').fillna(0).astype(int)

        today = date.today()
        created = 0
        updated = 0

        # 🔥 CACHE para no golpear la BD por cada celda
        # 1) cache de productos existentes por SKU
        existing_products = {
            p.sku: p for p in Product.query.all()
        }

        # 2) cache de tiendas existentes por nombre
        existing_stores = {
            s.name: s for s in Store.query.all()
        }

        # 3) cache de snapshots del día por (product_id, store_id)
        existing_snapshots = {
            (ss.product_id, ss.store_id): ss
            for ss in StockSnapshot.query.filter_by(as_of_date=today).all()
        }

        # ahora iteramos el dataframe
        for _, row in df.iterrows():
            sku = row[sku_col]
            if not sku:
                continue

            pname = row[prod_col] if prod_col and pd.notna(row[prod_col]) else sku
            cat_val = None
            if cat_col:
                raw_cat = row.get(cat_col)
                if raw_cat and pd.notna(raw_cat) and str(raw_cat) not in ['', 'nan', 'None', 'NaN', '<NA>']:
                    cat_val = str(raw_cat).strip()

            # producto: usar cache o crear
            prod = existing_products.get(sku)
            if not prod:
                prod = Product(sku=sku, name=str(pname))
                if cat_val:
                    prod.category = cat_val
                db.session.add(prod)
                db.session.flush()  # para tener prod.id
                existing_products[sku] = prod  # meterlo al cache
            elif cat_val:
                # Update category for existing product if NULL
                if not prod.category:
                    prod.category = cat_val
                elif prod.category != cat_val:
                    print(f"Category mismatch for SKU {sku}: existing={prod.category}, incoming={cat_val}")

            # por cada tienda
            for sc in store_cols:
                store_name = sc  # el nombre de la columna es el nombre de la tienda
                qty = int(row[sc])

                # tienda: usar cache o crear
                store = existing_stores.get(store_name)
                if not store:
                    store = Store(name=store_name)
                    db.session.add(store)
                    db.session.flush()
                    existing_stores[store_name] = store

                key = (prod.id, store.id)
                snapshot = existing_snapshots.get(key)

                if snapshot:
                    # actualizar
                    snapshot.quantity = qty
                    updated += 1
                else:
                    # crear
                    ss = StockSnapshot(
                        as_of_date=today,
                        product_id=prod.id,
                        store_id=store.id,
                        quantity=qty
                    )
                    db.session.add(ss)
                    existing_snapshots[key] = ss
                    created += 1

        db.session.commit()
        log_audit(
            action="stock_store.upload",
            message=f"Stock tiendas cargado: {created} nuevos, {updated} actualizados",
            entity_type="StockSnapshot",
            metadata={
                "filename": filename,
                "created": created,
                "updated": updated,
                "distinct_skus": len(df),
                "store_count": len(store_cols),
                "as_of_date": str(today)
            }
        )
        flash(f'Stock cargado. Nuevos: {created}. Actualizados: {updated}.', 'success')
        return redirect(url_for('dashboard'))

    # GET
    return render_template('upload_stock.html')

from datetime import date
from sqlalchemy import func
import io

@app.route('/purchase_projection', methods=['GET', 'POST'])
@login_required
@require_permission('forecast_v2:view')
def purchase_projection():
    # Parámetros por defecto
    lead_time_w = 2
    safety_pct = 0

    if request.method == 'POST':
        try:
            lead_time_w = max(int(request.form.get('lead_time_w', 2)), 0)
        except:
            lead_time_w = 2
        try:
            safety_pct = max(float(request.form.get('safety_pct', 0)), 0.0)
        except:
            safety_pct = 0.0

    # Última semana objetivo con predicciones
    latest_week = db.session.query(func.max(Prediction.target_period_start)).scalar()
    preview = []
    if not latest_week:
        return render_template('purchase_projection.html',
                               latest_week=None, preview=preview,
                               lead_time_w=lead_time_w, safety_pct=safety_pct)

    # Demanda total por SKU (sumando todas las tiendas) de la última corrida
    preds = (
        db.session.query(Product.id, Product.sku, Product.name, func.sum(Prediction.quantity).label('demand'))
        .join(Product, Prediction.product_id == Product.id)
        .filter(Prediction.target_period_start == latest_week)
        .group_by(Product.id, Product.sku, Product.name)
        .order_by(Product.sku.asc())
        .all()
    )

    # Stock CD de hoy
    today = date.today()
    cd_rows = (
        db.session.query(StockCD.product_id, StockCD.quantity)
        .filter(StockCD.as_of_date == today)
        .all()
    )
    cd_map = {pid: qty for pid, qty in cd_rows}

    # Cálculo:
    # necesidad = max(demanda * (lead_time + 1) * (1 + safety) - stock_cd, 0)
    result = []
    for pid, sku, name, demand in preds:
        demand = int(demand or 0)
        stock_cd = int(cd_map.get(pid, 0))
        mult = (lead_time_w + 1)  # demanda de horizonte: esta semana + LT semanas
        safety_factor = 1.0 + (safety_pct / 100.0)
        need = max(int(round(demand * mult * safety_factor)) - stock_cd, 0)

        result.append({
            "product_id": pid,
            "SKU": str(sku),
            "Producto": name,
            "Demanda base (últ. dist.)": demand,
            "Stock CD hoy": stock_cd,
            "Lead time (sem)": lead_time_w,
            "Colchón (%)": safety_pct,
            "Necesidad Neta": need,
        })

    # Preview top 20 con necesidad > 0
    result_sorted = sorted(result, key=lambda x: x["Necesidad Neta"], reverse=True)
    preview = [r for r in result_sorted if r["Necesidad Neta"] > 0][:20]

    return render_template('purchase_projection.html',
                           latest_week=latest_week,
                           preview=preview,
                           lead_time_w=lead_time_w,
                           safety_pct=safety_pct)

@app.route('/export_purchase_projection', methods=['POST'])
@login_required
@require_permission('forecast_v2:run')
def export_purchase_projection():
    try:
        lead_time_w = max(int(request.form.get('lead_time_w', 2)), 0)
    except:
        lead_time_w = 2
    try:
        safety_pct = max(float(request.form.get('safety_pct', 0)), 0.0)
    except:
        safety_pct = 0.0

    latest_week = db.session.query(func.max(Prediction.target_period_start)).scalar()
    if not latest_week:
        flash('No hay distribución reciente para proyectar.', 'warning')
        return redirect(url_for('purchase_projection'))

    preds = (
        db.session.query(Product.id, Product.sku, Product.name, func.sum(Prediction.quantity).label('demand'))
        .join(Product, Prediction.product_id == Product.id)
        .filter(Prediction.target_period_start == latest_week)
        .group_by(Product.id, Product.sku, Product.name)
        .order_by(Product.sku.asc())
        .all()
    )

    today = date.today()
    cd_rows = (
        db.session.query(StockCD.product_id, StockCD.quantity)
        .filter(StockCD.as_of_date == today)
        .all()
    )
    cd_map = {pid: qty for pid, qty in cd_rows}

    rows = []
    for pid, sku, name, demand in preds:
        demand = int(demand or 0)
        stock_cd = int(cd_map.get(pid, 0))
        mult = (lead_time_w + 1)
        safety_factor = 1.0 + (safety_pct / 100.0)
        need = max(int(round(demand * mult * safety_factor)) - stock_cd, 0)

        rows.append({
            "SKU": str(sku),
            "Producto": name,
            "Demanda base (últ. dist.)": demand,
            "Stock CD hoy": stock_cd,
            "Lead time (sem)": lead_time_w,
            "Colchón (%)": safety_pct,
            "Necesidad Neta": need,
            "Semana objetivo": latest_week.strftime("%Y-%m-%d"),
            "Fecha cálculo": today.strftime("%Y-%m-%d"),
        })

    # Excel (SKU como texto)
    import pandas as pd
    output = io.BytesIO()
    df = pd.DataFrame(rows)
    df["SKU"] = df["SKU"].astype(str)
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Proyeccion")
        ws = writer.sheets["Proyeccion"]
        for cell in ws["A"]:
            cell.number_format = "@"
    output.seek(0)

    fname = f"proyeccion_compra_{today.strftime('%Y%m%d')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

# ======================================================
# RESET DE DATOS
# ======================================================

@app.route('/reset_sales', methods=['POST'])
@login_required
@require_permission('admin:reset')
def reset_sales():
    count = db.session.query(DistributionRecord).count()
    db.session.query(DistributionRecord).delete()
    db.session.commit()
    log_audit(
        action="reset.sales",
        message=f"Eliminados {count} registros de ventas",
        entity_type="DistributionRecord",
        metadata={"deleted_count": count}
    )
    flash('✅ Se eliminaron todos los registros de ventas cargadas.', 'warning')
    return redirect(url_for('dashboard'))


@app.route('/reset_store_stock', methods=['POST'])
@login_required
@require_permission('admin:reset')
def reset_store_stock():
    count = db.session.query(StockSnapshot).count()
    db.session.query(StockSnapshot).delete()
    db.session.commit()
    log_audit(
        action="reset.stock_store",
        message=f"Eliminados {count} snapshots de stock tiendas",
        entity_type="StockSnapshot",
        metadata={"deleted_count": count}
    )
    flash('✅ Se eliminó todo el stock de tiendas.', 'warning')
    return redirect(url_for('dashboard'))


@app.route('/reset_predictions', methods=['POST'])
@login_required
@require_permission('admin:reset')
def reset_predictions():
    count = db.session.query(Prediction).count()
    db.session.query(Prediction).delete()
    db.session.commit()
    log_audit(
        action="reset.predictions",
        message=f"Eliminadas {count} predicciones",
        entity_type="Prediction",
        metadata={"deleted_count": count}
    )
    flash('✅ Se eliminaron todas las distribuciones sugeridas.', 'warning')
    return redirect(url_for('dashboard'))


@app.route('/reset_stock_cd', methods=['POST'])
@login_required
@require_permission('admin:reset')
def reset_stock_cd():
    count = StockCD.query.count()
    StockCD.query.delete()
    db.session.commit()
    log_audit(
        action="reset.stock_cd",
        message=f"Eliminados {count} registros stock CD",
        entity_type="StockCD",
        metadata={"deleted_count": count}
    )
    flash('Stock CD completamente reiniciado.', 'success')
    return redirect(url_for('dashboard'))

def process_stock_cd_upload(job_id, payload):
    """
    Background task to process Stock CD upload with bulk operations. Uses isolated session.
    
    Modes:
    - append: Add quantities to existing SKUs, create new ones. Does NOT touch SKUs not in file.
    - replace_today: Delete all stock for snapshot_date, then insert new rows.
    - replace_all: Delete ALL stock history, then insert new rows.
    
    Append mode scenario:
    - Start: SKU A=10, SKU B=20 (snapshot_date X)
    - Upload: SKU A +5
    - Result: SKU A=15, SKU B=20 (unchanged), same snapshot_date
    """
    from datetime import date
    import os
    
    filepath = payload.get('filepath')
    modo = payload.get('modo', 'replace_today')
    fecha_doc_str = payload.get('fecha_doc')
    user_id = payload.get('user_id')
    original_filename = payload.get('original_filename', 'unknown')
    
    session = get_isolated_session()
    try:
        update_job_status(job_id, progress=5, message='Determinando fecha de snapshot...')
        
        if fecha_doc_str:
            snapshot_date = date.fromisoformat(fecha_doc_str)
        else:
            max_date_row = session.query(func.max(StockCD.as_of_date)).scalar()
            snapshot_date = max_date_row if max_date_row else date.today()
        
        update_job_status(job_id, progress=10, message='Leyendo archivo...')
        
        if filepath.endswith('.csv'):
            df = pd.read_csv(filepath, dtype=str)
        else:
            df = pd.read_excel(filepath, dtype=str)
        
        total_rows = len(df)
        update_job_status(job_id, progress=20, message=f'Procesando {total_rows} filas...')
        
        df.columns = [str(c).strip().lower() for c in df.columns]
        if 'sku' not in df.columns or 'quantity' not in df.columns:
            update_job_status(job_id, status='error', progress=100, message='Archivo debe tener columnas sku y quantity')
            return
        
        df['sku'] = df['sku'].astype(str).str.strip()
        df['quantity'] = pd.to_numeric(df['quantity'], errors='coerce').fillna(0).astype(int)
        
        name_col = None
        for col in ['product_name', 'producto', 'name', 'nombre']:
            if col in df.columns:
                name_col = col
                break
        
        # Optional category column
        has_category = 'category' in df.columns
        if has_category:
            df['category'] = df['category'].astype(str).str.strip()
            df['category'] = df['category'].replace(['', 'nan', 'None', 'NaN'], None)
        
        update_job_status(job_id, progress=30, message='Verificando productos...')
        
        unique_skus = list(df['sku'].unique())
        sku_to_product_id = {}
        
        batch_size = 500
        for i in range(0, len(unique_skus), batch_size):
            sku_batch = unique_skus[i:i + batch_size]
            for p in session.query(Product).filter(Product.sku.in_(sku_batch)).all():
                sku_to_product_id[p.sku] = p.id
        
        update_job_status(job_id, progress=40, message='Procesando SKUs...')
        
        # Build category map from upload data
        sku_category_map = {}
        if has_category:
            for sku in unique_skus:
                match = df[df['sku'] == sku]
                if not match.empty:
                    cat_val = match.iloc[0].get('category')
                    if cat_val and cat_val not in [None, 'None', 'nan', '']:
                        sku_category_map[sku] = cat_val
        
        new_products = []
        for sku in unique_skus:
            if sku not in sku_to_product_id:
                pname = None
                if name_col:
                    match = df[df['sku'] == sku]
                    if not match.empty and pd.notna(match.iloc[0].get(name_col)):
                        pname = str(match.iloc[0][name_col]).strip()
                prod_data = {'sku': sku, 'name': pname if pname else f"SKU {sku}"}
                if sku in sku_category_map:
                    prod_data['category'] = sku_category_map[sku]
                new_products.append(prod_data)
        
        if new_products:
            update_job_status(job_id, progress=50, message=f'Creando {len(new_products)} productos nuevos...')
            session.execute(Product.__table__.insert(), new_products)
            session.commit()
            new_skus = [np['sku'] for np in new_products]
            for i in range(0, len(new_skus), batch_size):
                sku_batch = new_skus[i:i + batch_size]
                for p in session.query(Product).filter(Product.sku.in_(sku_batch)).all():
                    sku_to_product_id[p.sku] = p.id
        
        # Update category for existing products with NULL category
        if has_category and sku_category_map:
            category_updates = 0
            category_mismatches = 0
            existing_skus_to_check = [sku for sku in sku_category_map.keys() if sku in sku_to_product_id]
            for i in range(0, len(existing_skus_to_check), batch_size):
                sku_batch = existing_skus_to_check[i:i + batch_size]
                for prod in session.query(Product).filter(Product.sku.in_(sku_batch)).all():
                    incoming_cat = sku_category_map.get(prod.sku)
                    if incoming_cat:
                        if not prod.category:
                            prod.category = incoming_cat
                            category_updates += 1
                        elif prod.category != incoming_cat:
                            print(f"Category mismatch for SKU {prod.sku}: existing={prod.category}, incoming={incoming_cat}")
                            category_mismatches += 1
            if category_updates > 0:
                session.commit()
                print(f"Stock CD: Updated category for {category_updates} products, {category_mismatches} mismatches logged")
        
        update_job_status(job_id, progress=60, message='Preparando stock...')
        
        created = 0
        updated = 0
        
        if modo == 'replace_all':
            session.query(StockCD).delete()
            session.commit()
        elif modo == 'replace_today':
            session.query(StockCD).filter_by(as_of_date=snapshot_date).delete()
            session.commit()
        
        existing_stock = {}
        if modo == 'append':
            for sc in session.query(StockCD).filter_by(as_of_date=snapshot_date).all():
                existing_stock[sc.product_id] = sc
        
        update_job_status(job_id, progress=70, message='Insertando stock...')
        
        stock_inserts = []
        
        for _, row in df.iterrows():
            sku = row['sku']
            qty = int(row['quantity'])
            pid = sku_to_product_id.get(sku)
            if not pid:
                continue
            
            if modo == 'append' and pid in existing_stock:
                existing_stock[pid].quantity += qty
                existing_stock[pid].as_of_date = snapshot_date
                updated += 1
            else:
                stock_inserts.append({
                    'as_of_date': snapshot_date,
                    'product_id': pid,
                    'quantity': qty
                })
                created += 1
        
        if stock_inserts:
            insert_batch = 1000
            for i in range(0, len(stock_inserts), insert_batch):
                batch = stock_inserts[i:i + insert_batch]
                session.execute(StockCD.__table__.insert(), batch)
                progress = 70 + int(20 * (i + len(batch)) / len(stock_inserts))
                update_job_status(job_id, progress=progress, message=f'Insertando {i + len(batch)}/{len(stock_inserts)}...')
        
        session.commit()
        
        update_job_status(job_id, progress=95, message='Finalizando...')
        
        try:
            os.remove(filepath)
        except:
            pass
        
        mode_label = {'append': 'sumado', 'replace_today': 'reemplazado hoy', 'replace_all': 'reemplazado todo'}.get(modo, modo)
        update_job_status(
            job_id,
            status='done',
            progress=100,
            message=f'Stock CD {mode_label} ({snapshot_date}). Nuevos: {created}, Actualizados: {updated}',
            result={'created': created, 'updated': updated, 'total_rows': total_rows, 'snapshot_date': str(snapshot_date)}
        )
        
    except Exception as e:
        session.rollback()
        update_job_status(job_id, status='error', progress=100, message=f'Error: {str(e)[:400]}')
        raise
    finally:
        session.close()


@app.route('/stock_cd', methods=['GET', 'POST'])
@login_required
@require_permission('stock_cd:upload')
def upload_stock_cd():
    from datetime import date
    import os
    
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    if request.method == 'POST':
        file = request.files.get('file')
        if not file:
            if is_ajax:
                return jsonify({'ok': False, 'message': 'Sube un archivo CSV o Excel.', 'category': 'warning'}), 400
            flash('Sube un archivo CSV o Excel.', 'warning')
            return redirect(url_for('upload_stock_cd'))

        filename = file.filename.lower()
        if not (filename.endswith('.csv') or filename.endswith('.xlsx') or filename.endswith('.xls')):
            if is_ajax:
                return jsonify({'ok': False, 'message': 'Formato no soportado. Usa .csv o .xlsx', 'category': 'danger'}), 400
            flash('Formato no soportado. Usa .csv o .xlsx', 'danger')
            return redirect(url_for('upload_stock_cd'))
        
        upload_dir = os.path.join(BASE_DIR, 'instance', 'uploads')
        os.makedirs(upload_dir, exist_ok=True)
        
        saved_filename = f"stockcd_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}_{file.filename}"
        filepath = os.path.join(upload_dir, saved_filename)
        file.save(filepath)
        
        modo = request.form.get('mode', 'replace_today')
        fecha_doc_str = request.form.get('fecha_doc', '').strip()
        user_id = current_user.id if current_user.is_authenticated else None
        
        payload = {
            'filepath': filepath,
            'modo': modo,
            'fecha_doc': fecha_doc_str if fecha_doc_str else None,
            'user_id': user_id,
            'original_filename': file.filename
        }
        
        job_id = create_and_run_job(
            job_type='upload_stock_cd',
            task_func=process_stock_cd_upload,
            payload=payload,
            user_id=user_id
        )
        
        if is_ajax:
            flash('Stock CD cargado exitosamente.', 'success')
            return jsonify({
                'ok': True,
                'job_id': job_id,
                'redirect_url': url_for('dashboard'),
                'message': 'Archivo recibido. Procesando...',
                'category': 'success'
            })
        
        return redirect(url_for('job_status_view', job_id=job_id))

    return render_template('upload_stock_cd.html')

@app.route('/stock_query', methods=['GET'])
@login_required
@require_permission('stock:query')
def stock_query():
    sku = (request.args.get("sku") or "").strip()
    scope = (request.args.get("scope") or "cd").strip()  # cd | tiendas
    store_name = (request.args.get("store") or "").strip()

    result = None
    stores = Store.query.order_by(Store.name.asc()).all()

    if sku:
        # SKU SIEMPRE como texto
        product = Product.query.filter_by(sku=sku).first()

        if product:
            if scope == "cd":
                # Tomamos el registro más reciente (no solo hoy)
                cd_row = (StockCD.query
                          .filter_by(product_id=product.id)
                          .order_by(StockCD.as_of_date.desc())
                          .first())
                result = {
                    "sku": product.sku,
                    "product": product.name,
                    "scope": "cd",
                    "as_of_date": cd_row.as_of_date if cd_row else None,
                    "quantity": int(cd_row.quantity) if cd_row else 0
                }

            else:
                # Stock de tiendas: último snapshot por tienda
                q = (db.session.query(StockSnapshot, Store)
                     .join(Store, StockSnapshot.store_id == Store.id)
                     .filter(StockSnapshot.product_id == product.id))

                if store_name:
                    q = q.filter(Store.name == store_name)

                rows = (q.order_by(StockSnapshot.as_of_date.desc()).all())

                # quedarnos con el más reciente por tienda
                latest_by_store = {}
                for snap, st in rows:
                    if st.id not in latest_by_store:
                        latest_by_store[st.id] = {
                            "store": st.name,
                            "as_of_date": snap.as_of_date,
                            "quantity": int(snap.quantity)
                        }

                result = {
                    "sku": product.sku,
                    "product": product.name,
                    "scope": "tiendas",
                    "stores": list(latest_by_store.values())
                }

        else:
            result = {"error": f"No existe producto con SKU: {sku}"}

    return render_template(
        "stock_query.html",
        stores=stores,
        selected_scope=scope,
        selected_store=store_name,
        sku=sku,
        result=result
    )

@app.route('/export_predictions', methods=['GET'])
@login_required
@require_permission('distribution:export')
def export_predictions():
    from io import BytesIO
    import pandas as pd

    run_id = request.args.get("run_id", "").strip()
    debug_mode = request.args.get("debug", "0") == "1"
    
    is_admin = current_user.role in ('Admin', 'Management')
    include_debug = debug_mode and is_admin
    
    if not run_id:
        latest_run = (
            PredictionRun.query
            .filter(PredictionRun.run_id.isnot(None))
            .filter(PredictionRun.run_id != "")
            .filter(PredictionRun.run_type == 'distribution')
            .order_by(PredictionRun.created_at.desc())
            .first()
        )
        run_id = latest_run.run_id if latest_run else None

    if not run_id:
        flash('No hay predicciones para exportar.', 'warning')
        return redirect(url_for('dashboard'))

    preds = (
        db.session.query(Prediction, Product, Store)
        .join(Product, Prediction.product_id == Product.id)
        .join(Store, Prediction.store_id == Store.id)
        .filter(Prediction.run_id == run_id)
        .filter(Prediction.quantity > 0)
        .order_by(Product.sku.asc(), Store.name.asc())
        .all()
    )

    rows = []
    for p, prod, st in preds:
        folio = ""
        responsable = ""
        categoria = ""
        fecha_doc = ""

        parts = (p.model_name or "").split("—", 1)
        base_model = parts[0].strip()
        meta_part = parts[1].strip() if len(parts) > 1 else ""

        if meta_part:
            meta_items = [m.strip() for m in meta_part.split("|")]
            for item in meta_items:
                low = item.lower()
                if low.startswith("folio:"):
                    folio = item.split(":", 1)[1].strip()
                elif low.startswith("resp:"):
                    responsable = item.split(":", 1)[1].strip()
                elif low.startswith("cat:"):
                    categoria = item.split(":", 1)[1].strip()
                elif low.startswith("fecha doc:"):
                    fecha_doc = item.split(":", 1)[1].strip()

        row = {
            "SKU": str(prod.sku),
            "Producto": prod.name,
            "Tienda": st.name,
            "Cantidad sugerida": p.quantity,
            "Modelo": base_model,
            "Folio": folio,
            "Responsable": responsable,
            "Categoría": categoria,
            "Fecha documento": fecha_doc,
            "Corrida": run_id[:8],
        }
        
        if include_debug:
            debug_data = p.get_debug()
            row["w0_units"] = debug_data.get('w0_units', '')
            row["w1_units"] = debug_data.get('w1_units', '')
            row["sma_mean"] = debug_data.get('sma_mean', '')
            row["weeks_target"] = debug_data.get('weeks_target', '')
            row["target_units"] = debug_data.get('target_units', '')
            row["stock_store_used"] = debug_data.get('stock_store_used', '')
            row["cd_stock_used"] = debug_data.get('cd_stock_used', '')
            row["suggested_before_caps"] = debug_data.get('suggested_before_caps', '')
            row["suggested_final"] = debug_data.get('suggested_final', '')
            row["reason_code"] = debug_data.get('reason_code', '')
            row["store_rank"] = debug_data.get('store_rank', '')
            row["total_sales_in_window"] = debug_data.get('total_sales_in_window', '')
        
        rows.append(row)

    if not rows:
        flash('No hay filas para exportar.', 'warning')
        return redirect(url_for('dashboard'))

    df = pd.DataFrame(rows)
    df["SKU"] = df["SKU"].astype(str)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Distribucion")
        ws = writer.sheets["Distribucion"]
        for cell in ws["A"]:
            cell.number_format = "@"

    output.seek(0)
    today = date.today()
    suffix = "_DEBUG" if include_debug else ""
    fname = f"FastDSO_Distribution_{run_id[:8]}_{today.strftime('%Y%m%d')}{suffix}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@app.route('/runs', methods=['GET'])
@login_required
@require_permission('runs:view')
def runs():
    """Display history of all runs with pagination, search, and filter chips."""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    search = (request.args.get('search') or '').strip()
    run_type_filter = (request.args.get('run_type') or '').strip()

    per_page = per_page if per_page in [10, 25, 50] else 10

    runs_q = Run.query

    if run_type_filter:
        runs_q = runs_q.filter(Run.run_type == run_type_filter)

    if search:
        runs_q = runs_q.filter(
            or_(
                Run.run_id.ilike(f"%{search}%"),
                Run.folio.ilike(f"%{search}%"),
                Run.responsable.ilike(f"%{search}%"),
                Run.categoria.ilike(f"%{search}%"),
                Run.run_type.ilike(f"%{search}%")
            )
        )

    total = runs_q.count()
    runs_list = (
        runs_q
        .order_by(Run.created_at.desc())
        .offset((page - 1) * per_page)
        .limit(per_page)
        .all()
    )
    pagination = Pagination(page, per_page, total, runs_list)

    pending_approvals = Run.query.filter_by(status='PENDING_APPROVAL').order_by(Run.submitted_at.desc()).all() if current_user.has_permission('runs:approve') else []
    pending_count = len(pending_approvals)

    return render_template(
        'runs.html',
        runs=runs_list,
        pagination=pagination,
        search=search,
        per_page=per_page,
        run_type_filter=run_type_filter,
        pending_approvals=pending_approvals,
        pending_count=pending_count,
        can_approve=current_user.has_permission('runs:approve'),
        can_submit=current_user.has_permission('runs:submit')
    )


@app.route('/run/<run_id>', methods=['GET'])
@login_required
@require_permission('runs:view')
def view_run(run_id):
    """View details of a specific run."""
    run = Run.query.filter_by(run_id=run_id).first_or_404()

    if run.run_type == 'distribution':
        predictions = (
            db.session.query(Prediction, Product, Store)
            .join(Product, Prediction.product_id == Product.id)
            .join(Store, Prediction.store_id == Store.id)
            .filter(Prediction.run_id == run_id)
            .filter(Prediction.quantity > 0)
            .order_by(Prediction.quantity.desc())
            .limit(100)
            .all()
        )
        return render_template('run_detail.html', run=run, predictions=predictions)

    elif run.run_type == 'forecast_v2':
        forecast_results = (
            ForecastResult.query
            .filter_by(run_id=run_id)
            .order_by(ForecastResult.suggested.desc())
            .limit(100)
            .all()
        )
        return render_template('run_detail.html', run=run, forecast_results=forecast_results)

    elif run.run_type == 'sales_upload':
        sales_summary = (
            db.session.query(
                Product.sku,
                Product.name,
                db.func.sum(DistributionRecord.quantity).label('total_qty')
            )
            .join(Product, DistributionRecord.product_id == Product.id)
            .filter(DistributionRecord.run_id == run_id)
            .group_by(Product.id, Product.sku, Product.name)
            .order_by(db.func.sum(DistributionRecord.quantity).desc())
            .limit(50)
            .all()
        )
        return render_template('run_detail.html', run=run, sales_summary=sales_summary)

    return render_template('run_detail.html', run=run)


@app.route('/runs/<run_id>/activate', methods=['POST'])
@login_required
@require_permission('runs:view')
def activate_run(run_id):
    """Set a run as active (only one active at a time)."""
    if current_user.role not in ['Admin', 'Management']:
        flash('Solo Admin o Management pueden activar corridas.', 'danger')
        return redirect(url_for('runs'))
    
    run = Run.query.filter_by(run_id=run_id).first_or_404()
    
    if run.run_type != 'distribution':
        flash('Solo corridas de distribución pueden ser activadas.', 'warning')
        return redirect(url_for('runs'))
    
    Run.query.filter(Run.run_type == 'distribution').update({'is_active': False})
    run.is_active = True
    db.session.commit()
    
    log_audit(
        action='run.activate',
        entity_type='run',
        entity_id=run.id,
        run_id=run_id,
        status='success',
        message=f'Corrida {run_id[:8]} activada'
    )
    
    flash(f'Corrida {run_id[:8]} activada exitosamente.', 'success')
    return redirect(url_for('runs'))


@app.route('/runs/<run_id>/status', methods=['POST'])
@login_required
@require_permission('runs:view')
def change_run_status(run_id):
    """Change run status (Admin/Manager only)."""
    if current_user.role not in ['Admin', 'Management']:
        flash('Solo Admin o Management pueden cambiar el estado.', 'danger')
        return redirect(url_for('runs'))
    
    new_status = request.form.get('status', '').strip()
    allowed_statuses = ['DRAFT', 'SIMULATED', 'APPROVED', 'EXECUTED', 'ARCHIVED', 'completed', 'failed']
    
    if new_status not in allowed_statuses:
        flash('Estado inválido.', 'danger')
        return redirect(url_for('runs'))
    
    run = Run.query.filter_by(run_id=run_id).first_or_404()
    old_status = run.status
    run.status = new_status
    db.session.commit()
    
    log_audit(
        action='run.status_change',
        entity_type='run',
        entity_id=run.id,
        run_id=run_id,
        status='success',
        message=f'Estado cambiado de {old_status} a {new_status}'
    )
    
    flash(f'Estado de corrida cambiado a {new_status}.', 'success')
    return redirect(url_for('runs'))


# ------------------ Run Workflow: Submit for Approval ------------------
@app.route('/runs/<run_id>/submit_for_approval', methods=['POST'])
@login_required
@require_permission('runs:submit')
def submit_for_approval(run_id):
    """Submit a distribution run for manager approval."""
    run = Run.query.filter_by(run_id=run_id).first_or_404()
    
    if run.run_type != 'distribution':
        flash('Solo se pueden enviar a autorización corridas de distribución', 'warning')
        return redirect(url_for('runs'))
    
    if run.status not in ['DRAFT', 'completed']:
        flash('Solo se pueden enviar a autorización corridas en estado BORRADOR', 'warning')
        return redirect(url_for('runs'))
    
    urgency = request.form.get('urgency', 'MEDIUM')
    submit_note = request.form.get('submit_note', '').strip()
    
    if urgency not in ['LOW', 'MEDIUM', 'URGENT']:
        urgency = 'MEDIUM'
    
    run.status = 'PENDING_APPROVAL'
    run.urgency = urgency
    run.submit_note = submit_note
    run.submitted_by_user_id = current_user.id
    run.submitted_at = datetime.utcnow()
    db.session.commit()
    
    log_audit(
        action='run.submit_for_approval',
        entity_type='run',
        entity_id=run.id,
        run_id=run_id,
        status='success',
        message=f'Enviado a autorización con urgencia {urgency}'
    )
    
    flash('Corrida enviada a autorización exitosamente', 'success')
    return redirect(url_for('runs'))


@app.route('/runs/<run_id>/approve_and_send', methods=['POST'])
@login_required
@require_permission('runs:approve')
def approve_and_send(run_id):
    """Approve a run and send it to FastPlanner."""
    run = Run.query.filter_by(run_id=run_id).first_or_404()
    
    if run.run_type != 'distribution':
        flash('Solo se pueden aprobar corridas de distribución', 'warning')
        return redirect(url_for('runs'))
    
    if run.status != 'PENDING_APPROVAL':
        flash('Solo se pueden aprobar corridas en estado PENDIENTE DE APROBACIÓN', 'warning')
        return redirect(url_for('runs'))
    
    predictions_count = Prediction.query.filter_by(run_id=run.run_id).filter(Prediction.quantity > 0).count()
    if predictions_count == 0:
        flash('No hay cantidades positivas para enviar a bodega', 'warning')
        return redirect(url_for('runs'))
    
    folio = run.folio or f"PLAN-{run.run_id[:8].upper()}"
    existing = DistributionPlan.query.filter_by(folio=folio).first()
    if existing:
        folio = f"{folio}-{datetime.utcnow().strftime('%H%M%S')}"
    
    plan = DistributionPlan(
        folio=folio,
        source_run_id=run.run_id,
        status='APPROVED',
        urgency=run.urgency or 'MEDIUM',
        notes_commercial=run.submit_note,
        created_by_user_id=run.submitted_by_user_id,
        approved_by_user_id=current_user.id,
        created_at=run.submitted_at or datetime.utcnow(),
        approved_at=datetime.utcnow()
    )
    db.session.add(plan)
    db.session.flush()
    
    predictions = Prediction.query.filter_by(run_id=run.run_id).filter(Prediction.quantity > 0).all()
    for pred in predictions:
        line = DistributionPlanLine(
            plan_id=plan.id,
            product_id=pred.product_id,
            store_id=pred.store_id,
            qty_planned=pred.quantity
        )
        db.session.add(line)
    
    activity = PlanActivityLog(
        plan_id=plan.id,
        action='APPROVED_AND_SENT',
        user_id=current_user.id,
        comment=f'Plan aprobado y enviado a FastPlanner por {current_user.username}'
    )
    db.session.add(activity)
    
    run.status = 'APPROVED'
    run.approved_by_user_id = current_user.id
    run.approved_at = datetime.utcnow()
    run.planner_plan_id = plan.id
    
    db.session.commit()
    
    log_audit(
        action='run.approved_and_sent',
        entity_type='run',
        entity_id=run.id,
        run_id=run_id,
        status='success',
        message=f'Aprobado y enviado a FastPlanner como {plan.folio}'
    )
    
    flash(f'Corrida aprobada y enviada a FastPlanner como {plan.folio}', 'success')
    return redirect(url_for('runs'))


@app.route('/runs/<run_id>/reject', methods=['POST'])
@login_required
@require_permission('runs:approve')
def reject_run(run_id):
    """Reject a run pending approval."""
    run = Run.query.filter_by(run_id=run_id).first_or_404()
    
    if run.run_type != 'distribution':
        flash('Solo se pueden rechazar corridas de distribución', 'warning')
        return redirect(url_for('runs'))
    
    if run.status != 'PENDING_APPROVAL':
        flash('Solo se pueden rechazar corridas en estado PENDIENTE DE APROBACIÓN', 'warning')
        return redirect(url_for('runs'))
    
    reject_comment = request.form.get('reject_comment', '').strip()
    if not reject_comment:
        flash('Se requiere un comentario para rechazar', 'warning')
        return redirect(url_for('runs'))
    
    run.status = 'REJECTED'
    run.rejected_by_user_id = current_user.id
    run.rejected_at = datetime.utcnow()
    run.reject_comment = reject_comment
    db.session.commit()
    
    log_audit(
        action='run.rejected',
        entity_type='run',
        entity_id=run.id,
        run_id=run_id,
        status='success',
        message=f'Rechazado: {reject_comment[:50]}'
    )
    
    flash('Corrida rechazada', 'info')
    return redirect(url_for('runs'))


@app.route('/runs/compare', methods=['GET'])
@login_required
@require_permission('runs:view')
def compare_runs():
    """Compare two distribution runs side by side."""
    left_id = request.args.get('left', '').strip()
    right_id = request.args.get('right', '').strip()
    
    distribution_runs = (
        Run.query
        .filter(Run.run_type == 'distribution')
        .order_by(Run.created_at.desc())
        .limit(50)
        .all()
    )
    
    left_run = None
    right_run = None
    comparison_data = []
    left_kpis = {}
    right_kpis = {}
    delta_kpis = {}
    
    if left_id and right_id:
        left_run = Run.query.filter_by(run_id=left_id).first()
        right_run = Run.query.filter_by(run_id=right_id).first()
        
        if left_run and right_run:
            left_preds = (
                db.session.query(
                    Product.sku,
                    Product.name,
                    Store.name.label('store_name'),
                    Prediction.quantity
                )
                .join(Product, Prediction.product_id == Product.id)
                .join(Store, Prediction.store_id == Store.id)
                .filter(Prediction.run_id == left_id)
                .all()
            )
            
            right_preds = (
                db.session.query(
                    Product.sku,
                    Product.name,
                    Store.name.label('store_name'),
                    Prediction.quantity
                )
                .join(Product, Prediction.product_id == Product.id)
                .join(Store, Prediction.store_id == Store.id)
                .filter(Prediction.run_id == right_id)
                .all()
            )
            
            left_map = {(str(p.sku), p.store_name): (p.name, p.quantity) for p in left_preds}
            right_map = {(str(p.sku), p.store_name): (p.name, p.quantity) for p in right_preds}
            
            all_keys = set(left_map.keys()) | set(right_map.keys())
            
            for key in all_keys:
                sku, store = key
                left_name, left_qty = left_map.get(key, ('', 0))
                right_name, right_qty = right_map.get(key, ('', 0))
                name = left_name or right_name
                delta = right_qty - left_qty
                
                comparison_data.append({
                    'sku': sku,
                    'name': name,
                    'store': store,
                    'left_qty': left_qty,
                    'right_qty': right_qty,
                    'delta': delta
                })
            
            comparison_data.sort(key=lambda x: abs(x['delta']), reverse=True)
            
            left_total = sum(p.quantity for p in left_preds)
            right_total = sum(p.quantity for p in right_preds)
            left_skus = len(set(p.sku for p in left_preds))
            right_skus = len(set(p.sku for p in right_preds))
            left_stores = len(set(p.store_name for p in left_preds))
            right_stores = len(set(p.store_name for p in right_preds))
            
            left_kpis = {'total_units': left_total, 'distinct_skus': left_skus, 'distinct_stores': left_stores}
            right_kpis = {'total_units': right_total, 'distinct_skus': right_skus, 'distinct_stores': right_stores}
            delta_kpis = {
                'total_units': right_total - left_total,
                'distinct_skus': right_skus - left_skus,
                'distinct_stores': right_stores - left_stores
            }
    
    return render_template(
        'runs_compare.html',
        distribution_runs=distribution_runs,
        left_id=left_id,
        right_id=right_id,
        left_run=left_run,
        right_run=right_run,
        comparison_data=comparison_data[:200],
        left_kpis=left_kpis,
        right_kpis=right_kpis,
        delta_kpis=delta_kpis
    )


# ======================================================
# ADMIN USER MANAGEMENT
# ======================================================

@app.route('/admin/users', methods=['GET'])
@login_required
@require_permission('admin:users')
def admin_users():
    """Admin page to manage users."""
    users = User.query.order_by(User.username.asc()).all()
    return render_template('admin_users.html', users=users, roles=ROLES)


@app.route('/admin/users/create', methods=['POST'])
@login_required
@require_permission('admin:users')
def admin_create_user():
    """Create a new user."""
    username = (request.form.get('username') or '').strip().lower()
    password = request.form.get('password') or ''
    role = request.form.get('role') or 'Viewer'
    
    if not username or not password:
        flash('Usuario y contraseña son requeridos.', 'danger')
        return redirect(url_for('admin_users'))
    
    if role not in ROLES:
        flash('Rol inválido.', 'danger')
        return redirect(url_for('admin_users'))
    
    existing = User.query.filter_by(username=username).first()
    if existing:
        flash(f'El usuario "{username}" ya existe.', 'danger')
        return redirect(url_for('admin_users'))
    
    new_user = User(
        username=username,
        password_hash=generate_password_hash(password),
        role=role,
        is_active=True
    )
    db.session.add(new_user)
    db.session.commit()
    
    flash(f'Usuario "{username}" creado con rol {role}.', 'success')
    return redirect(url_for('admin_users'))


@app.route('/admin/users/<int:user_id>/update', methods=['POST'])
@login_required
@require_permission('admin:users')
def admin_update_user(user_id):
    """Update user role and active status."""
    user = User.query.get_or_404(user_id)
    
    if user.username == 'admin' and current_user.id != user.id:
        flash('No puedes modificar al usuario admin.', 'danger')
        return redirect(url_for('admin_users'))
    
    role = request.form.get('role')
    is_active = request.form.get('is_active') == 'on'
    
    if role and role in ROLES:
        user.role = role
    
    if user.username != 'admin':
        user.is_active = is_active
    
    db.session.commit()
    flash(f'Usuario "{user.username}" actualizado.', 'success')
    return redirect(url_for('admin_users'))


@app.route('/admin/users/<int:user_id>/reset-password', methods=['POST'])
@login_required
@require_permission('admin:users')
def admin_reset_password(user_id):
    """Reset user password."""
    user = User.query.get_or_404(user_id)
    new_password = request.form.get('new_password') or ''
    
    if not new_password:
        flash('La nueva contraseña es requerida.', 'danger')
        return redirect(url_for('admin_users'))
    
    user.password_hash = generate_password_hash(new_password)
    db.session.commit()
    
    flash(f'Contraseña de "{user.username}" actualizada.', 'success')
    return redirect(url_for('admin_users'))


# ======================================================
# AUDIT TRAIL
# ======================================================

@app.route('/audit', methods=['GET'])
@login_required
@require_permission('audit:view')
def audit_view():
    """View audit trail with filters and pagination."""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 25, type=int)
    
    action_filter = request.args.get('action', '').strip()
    user_filter = request.args.get('user', '').strip()
    status_filter = request.args.get('status', '').strip()
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()
    
    query = AuditLog.query
    
    if action_filter:
        query = query.filter(AuditLog.action.ilike(f'%{action_filter}%'))
    if user_filter:
        query = query.filter(AuditLog.username_snapshot.ilike(f'%{user_filter}%'))
    if status_filter:
        query = query.filter(AuditLog.status == status_filter)
    if date_from:
        try:
            from_date = datetime.strptime(date_from, '%Y-%m-%d')
            query = query.filter(AuditLog.created_at >= from_date)
        except:
            pass
    if date_to:
        try:
            to_date = datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(AuditLog.created_at < to_date)
        except:
            pass
    
    total = query.count()
    logs = query.order_by(AuditLog.created_at.desc()).offset((page - 1) * per_page).limit(per_page).all()
    
    pagination = Pagination(page, per_page, total, logs)
    
    unique_actions = db.session.query(AuditLog.action).distinct().order_by(AuditLog.action).all()
    unique_actions = [a[0] for a in unique_actions]
    
    return render_template('audit.html', 
                           logs=logs, 
                           pagination=pagination,
                           unique_actions=unique_actions,
                           filters={
                               'action': action_filter,
                               'user': user_filter,
                               'status': status_filter,
                               'date_from': date_from,
                               'date_to': date_to
                           })


# ======================================================
# REBALANCING (Store-to-Store) ROUTES
# ======================================================
import uuid
import math

def compute_rebalancing_suggestions(
    weeks_window=4,
    target_woc_min=1.5,
    target_woc_target=2.5,
    target_woc_max=6.0,
    retain_woc=4.0,
    stock_floor=1,
    min_transfer_qty=2,
    store_filter=None
):
    """
    Compute store-to-store rebalancing suggestions.
    Returns list of dicts: {product_id, sku, name, from_store_id, from_store, to_store_id, to_store, qty, sales_rate_to, woc_from, woc_to, score, reason}
    Optimized: Uses indexed lookups instead of O(n*m) scans.
    """
    today = date.today()
    window_start_week = today - timedelta(days=weeks_window * 7)
    window_start_week = window_start_week - timedelta(days=window_start_week.weekday())
    
    app.logger.info(f"Redistribution using SalesWeeklyAgg from {window_start_week} to {today}")
    
    latest_stock_date = db.session.query(db.func.max(StockSnapshot.as_of_date)).scalar()
    if not latest_stock_date:
        return []
    
    stock_map = {}
    stores_per_product_stock = defaultdict(set)
    stock_q = (
        db.session.query(
            StockSnapshot.product_id,
            StockSnapshot.store_id,
            StockSnapshot.quantity
        )
        .filter(StockSnapshot.as_of_date == latest_stock_date)
        .all()
    )
    for pid, sid, qty in stock_q:
        stock_map[(pid, sid)] = qty
        stores_per_product_stock[pid].add(sid)
    
    sales_q = (
        db.session.query(
            SalesWeeklyAgg.product_id,
            SalesWeeklyAgg.store_id,
            db.func.sum(SalesWeeklyAgg.units).label('total_units'),
            db.func.count(db.func.distinct(SalesWeeklyAgg.week_start)).label('weeks_count')
        )
        .filter(SalesWeeklyAgg.week_start >= window_start_week)
        .group_by(SalesWeeklyAgg.product_id, SalesWeeklyAgg.store_id)
        .all()
    )
    
    sales_rate_map = {}
    stores_per_product_sales = defaultdict(set)
    for pid, sid, total_units, weeks_count in sales_q:
        rate = total_units / max(weeks_count, 1)
        sales_rate_map[(pid, sid)] = rate
        stores_per_product_sales[pid].add(sid)
    
    product_info = {p.id: (p.sku, p.name) for p in Product.query.all()}
    store_info = {s.id: s.name for s in Store.query.all()}
    
    all_products = set(stores_per_product_stock.keys()) | set(stores_per_product_sales.keys())
    
    suggestions = []
    
    for pid in all_products:
        stores_for_pid = stores_per_product_stock.get(pid, set()) | stores_per_product_sales.get(pid, set())
        
        store_data = []
        for sid in stores_for_pid:
            stock = stock_map.get((pid, sid), 0)
            rate = sales_rate_map.get((pid, sid), 0)
            woc = stock / max(rate, 0.01)
            store_data.append({
                'store_id': sid,
                'stock': stock,
                'rate': rate,
                'woc': woc
            })
        
        receivers = []
        for sd in store_data:
            if sd['rate'] > 0 and sd['woc'] < target_woc_min:
                need = max(math.ceil(target_woc_target * sd['rate']) - sd['stock'], 0)
                if need > 0:
                    receivers.append({
                        **sd,
                        'need': need,
                        'priority': (sd['rate'], -sd['woc'])
                    })
        receivers.sort(key=lambda x: x['priority'], reverse=True)
        
        donors = []
        for sd in store_data:
            if sd['woc'] > target_woc_max:
                keep_units = max(math.ceil(retain_woc * sd['rate']), stock_floor)
                give_units = max(sd['stock'] - keep_units, 0)
                if give_units > 0:
                    donors.append({
                        **sd,
                        'give': give_units
                    })
        donors.sort(key=lambda x: x['give'], reverse=True)
        
        for receiver in receivers:
            if store_filter and store_info.get(receiver['store_id']) != store_filter:
                continue
            
            remaining_need = receiver['need']
            for donor in donors:
                if remaining_need <= 0:
                    break
                if donor['give'] <= 0:
                    continue
                if donor['store_id'] == receiver['store_id']:
                    continue
                
                transfer = min(remaining_need, donor['give'])
                
                is_extreme = receiver['woc'] < 0.5
                if transfer < min_transfer_qty and not is_extreme:
                    continue
                
                score = receiver['rate'] * 10 + (target_woc_min - receiver['woc']) * 5
                
                reason = f"WOC {receiver['woc']:.1f} < {target_woc_min}, rate {receiver['rate']:.1f}/wk"
                
                suggestions.append({
                    'product_id': pid,
                    'sku': product_info.get(pid, ('?', '?'))[0],
                    'name': product_info.get(pid, ('?', '?'))[1],
                    'from_store_id': donor['store_id'],
                    'from_store': store_info.get(donor['store_id'], '?'),
                    'to_store_id': receiver['store_id'],
                    'to_store': store_info.get(receiver['store_id'], '?'),
                    'qty': transfer,
                    'sales_rate_to': receiver['rate'],
                    'woc_from': donor['woc'],
                    'woc_to': receiver['woc'],
                    'score': score,
                    'reason': reason
                })
                
                remaining_need -= transfer
                donor['give'] -= transfer
    
    suggestions.sort(key=lambda x: x['score'], reverse=True)
    return suggestions


@app.route('/rebalancing', methods=['GET', 'POST'])
@login_required
@require_permission('rebalancing:view')
def rebalancing():
    """Store-to-store rebalancing suggestions."""
    stores = Store.query.order_by(Store.name.asc()).all()
    
    # Support SKU prefill from query param (for forecast broken stock CTA)
    prefill_sku = request.args.get('sku', '').strip()
    
    weeks_window = 4
    target_woc_min = 1.5
    target_woc_target = 2.5
    target_woc_max = 6.0
    retain_woc = 4.0
    stock_floor = 1
    min_transfer_qty = 2
    store_filter = ''
    
    suggestions = []
    run_info = None
    kpis = {'total_units': 0, 'num_moves': 0, 'num_skus': 0, 'num_receivers': 0}
    
    if request.method == 'POST':
        if not current_user.has_permission('rebalancing:run'):
            flash('No tienes permiso para ejecutar rebalanceos.', 'warning')
            return redirect(url_for('rebalancing'))
        
        try:
            weeks_window = int(request.form.get('weeks_window', 4))
        except:
            weeks_window = 4
        try:
            target_woc_min = float(request.form.get('target_woc_min', 1.5))
        except:
            target_woc_min = 1.5
        try:
            target_woc_target = float(request.form.get('target_woc_target', 2.5))
        except:
            target_woc_target = 2.5
        try:
            target_woc_max = float(request.form.get('target_woc_max', 6.0))
        except:
            target_woc_max = 6.0
        try:
            retain_woc = float(request.form.get('retain_woc', 4.0))
        except:
            retain_woc = 4.0
        try:
            stock_floor = int(request.form.get('stock_floor', 1))
        except:
            stock_floor = 1
        try:
            min_transfer_qty = int(request.form.get('min_transfer_qty', 2))
        except:
            min_transfer_qty = 2
        
        store_filter = request.form.get('store_filter', '').strip()
        simulate = request.form.get('simulate') == '1'
        
        suggestions = compute_rebalancing_suggestions(
            weeks_window=weeks_window,
            target_woc_min=target_woc_min,
            target_woc_target=target_woc_target,
            target_woc_max=target_woc_max,
            retain_woc=retain_woc,
            stock_floor=stock_floor,
            min_transfer_qty=min_transfer_qty,
            store_filter=store_filter or None
        )
        
        run_id = str(uuid.uuid4())
        params = {
            'weeks_window': weeks_window,
            'target_woc_min': target_woc_min,
            'target_woc_target': target_woc_target,
            'target_woc_max': target_woc_max,
            'retain_woc': retain_woc,
            'stock_floor': stock_floor,
            'min_transfer_qty': min_transfer_qty,
            'store_filter': store_filter or None
        }
        
        kpis['total_units'] = sum(s['qty'] for s in suggestions)
        kpis['num_moves'] = len(suggestions)
        kpis['num_skus'] = len(set(s['product_id'] for s in suggestions))
        kpis['num_receivers'] = len(set(s['to_store_id'] for s in suggestions))
        
        rebalance_run = RebalanceRun(
            run_id=run_id,
            created_by_user_id=current_user.id if current_user.is_authenticated else None,
            params_json=json.dumps(params),
            is_simulation=simulate
        )
        db.session.add(rebalance_run)
        db.session.flush()
        
        if suggestions:
            suggestion_records = [
                {
                    'run_id': run_id,
                    'product_id': s['product_id'],
                    'from_store_id': s['from_store_id'],
                    'to_store_id': s['to_store_id'],
                    'qty': s['qty'],
                    'sales_rate_to': s['sales_rate_to'],
                    'woc_from': s['woc_from'],
                    'woc_to': s['woc_to'],
                    'score': s['score'],
                    'reason': s['reason']
                }
                for s in suggestions
            ]
            batch_size = 1000
            for i in range(0, len(suggestion_records), batch_size):
                batch = suggestion_records[i:i + batch_size]
                db.session.execute(RebalanceSuggestion.__table__.insert(), batch)
        
        db.session.commit()
        
        if simulate:
            flash(f'SIMULACIÓN: {len(suggestions)} sugerencias calculadas.', 'info')
        else:
            log_audit(
                action='rebalancing.run',
                status='success',
                message=f'Generated {len(suggestions)} rebalancing suggestions',
                entity_type='RebalanceRun',
                run_id=run_id,
                metadata={
                    'params': params,
                    'num_suggestions': len(suggestions)
                }
            )
            flash(f'Se generaron {len(suggestions)} sugerencias de redistribución.', 'success')
        
        return redirect(url_for('rebalancing', run_id=run_id))
    
    else:
        requested_run_id = request.args.get('run_id', '').strip()
        if requested_run_id:
            target_run = RebalanceRun.query.filter_by(run_id=requested_run_id).first()
        else:
            target_run = RebalanceRun.query.order_by(RebalanceRun.created_at.desc()).first()
        
        if target_run:
            run_info = {
                'run_id': target_run.run_id,
                'created_at': target_run.created_at,
                'params': target_run.get_params(),
                'is_simulation': target_run.is_simulation
            }
            params = run_info['params']
            weeks_window = params.get('weeks_window', 4)
            target_woc_min = params.get('target_woc_min', 1.5)
            target_woc_target = params.get('target_woc_target', 2.5)
            target_woc_max = params.get('target_woc_max', 6.0)
            retain_woc = params.get('retain_woc', 4.0)
            stock_floor = params.get('stock_floor', 1)
            min_transfer_qty = params.get('min_transfer_qty', 2)
            store_filter = params.get('store_filter', '') or ''
            
            db_suggestions = RebalanceSuggestion.query.filter_by(run_id=target_run.run_id).all()
            for s in db_suggestions:
                suggestions.append({
                    'product_id': s.product_id,
                    'sku': s.product.sku if s.product else '?',
                    'name': s.product.name if s.product else '?',
                    'from_store_id': s.from_store_id,
                    'from_store': s.from_store.name if s.from_store else '?',
                    'to_store_id': s.to_store_id,
                    'to_store': s.to_store.name if s.to_store else '?',
                    'qty': s.qty,
                    'sales_rate_to': s.sales_rate_to,
                    'woc_from': s.woc_from,
                    'woc_to': s.woc_to,
                    'score': s.score,
                    'reason': s.reason
                })
            
            kpis['total_units'] = sum(s['qty'] for s in suggestions)
            kpis['num_moves'] = len(suggestions)
            kpis['num_skus'] = len(set(s['product_id'] for s in suggestions))
            kpis['num_receivers'] = len(set(s['to_store_id'] for s in suggestions))
    
    # Pagination for suggestions (10 per page)
    sug_per_page = 10
    try:
        sug_page = int(request.args.get('sug_page', 1))
    except:
        sug_page = 1
    sug_page = max(1, sug_page)
    
    sug_total = len(suggestions)
    sug_pages = max(1, (sug_total + sug_per_page - 1) // sug_per_page)
    sug_page = min(sug_page, sug_pages)
    
    start_idx = (sug_page - 1) * sug_per_page
    end_idx = start_idx + sug_per_page
    suggestions_page = suggestions[start_idx:end_idx]
    
    sug_pagination = {
        'page': sug_page,
        'pages': sug_pages,
        'total': sug_total,
        'per_page': sug_per_page,
        'has_prev': sug_page > 1,
        'has_next': sug_page < sug_pages,
        'prev_num': sug_page - 1,
        'next_num': sug_page + 1,
        'start': start_idx + 1 if sug_total > 0 else 0,
        'end': min(end_idx, sug_total)
    }
    
    is_simulation = run_info.get('is_simulation', False) if run_info else False
    
    return render_template(
        'rebalancing.html',
        stores=stores,
        suggestions=suggestions_page,
        sug_pagination=sug_pagination,
        kpis=kpis,
        run_info=run_info,
        is_simulation=is_simulation,
        prefill_sku=prefill_sku,
        params={
            'weeks_window': weeks_window,
            'target_woc_min': target_woc_min,
            'target_woc_target': target_woc_target,
            'target_woc_max': target_woc_max,
            'retain_woc': retain_woc,
            'stock_floor': stock_floor,
            'min_transfer_qty': min_transfer_qty,
            'store_filter': store_filter
        }
    )


@app.route('/export_rebalancing', methods=['GET'])
@login_required
@require_permission('rebalancing:view')
def export_rebalancing():
    """Export rebalancing suggestions to Excel."""
    run_id = request.args.get('run_id', '').strip()
    
    if not run_id:
        latest_run = RebalanceRun.query.order_by(RebalanceRun.created_at.desc()).first()
        if latest_run:
            run_id = latest_run.run_id
        else:
            flash('No hay corridas de redistribución disponibles.', 'warning')
            return redirect(url_for('rebalancing'))
    
    run = RebalanceRun.query.filter_by(run_id=run_id).first()
    if not run:
        flash('Corrida no encontrada.', 'danger')
        return redirect(url_for('rebalancing'))
    
    suggestions = RebalanceSuggestion.query.filter_by(run_id=run_id).order_by(RebalanceSuggestion.score.desc()).all()
    
    rows = []
    for s in suggestions:
        rows.append({
            'SKU': s.product.sku if s.product else '',
            'Producto': s.product.name if s.product else '',
            'Tienda Origen': s.from_store.name if s.from_store else '',
            'Tienda Destino': s.to_store.name if s.to_store else '',
            'Cantidad': s.qty,
            'Venta/Sem Destino': round(s.sales_rate_to, 2) if s.sales_rate_to else 0,
            'WOC Origen': round(s.woc_from, 2) if s.woc_from else 0,
            'WOC Destino': round(s.woc_to, 2) if s.woc_to else 0,
            'Score': round(s.score, 2) if s.score else 0,
            'Razón': s.reason or ''
        })
    
    df = pd.DataFrame(rows)
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Redistribución')
    output.seek(0)
    
    filename = f"redistribucion_{run_id[:8]}_{date.today().isoformat()}.xlsx"
    
    log_audit(
        action='rebalancing.export',
        status='success',
        message=f'Exported {len(rows)} rebalancing suggestions',
        entity_type='RebalanceRun',
        run_id=run_id,
        metadata={'rows': len(rows)}
    )
    
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ------------------ Manual Transfer Builder Routes ------------------

@app.route('/rebalancing/manual/validate', methods=['POST'])
@login_required
@require_permission('rebalancing:run')
def manual_transfer_validate():
    """Validate manual transfer items from CSV/paste input."""
    from_store_id = request.form.get('from_store_id', type=int)
    to_store_id = request.form.get('to_store_id', type=int)
    donor_floor = request.form.get('donor_floor', 1, type=int)
    
    if not from_store_id or not to_store_id:
        return jsonify({'error': 'Tienda origen y destino son requeridas'}), 400
    
    if from_store_id == to_store_id:
        return jsonify({'error': 'Tienda origen no puede ser igual a destino'}), 400
    
    from_store = Store.query.get(from_store_id)
    to_store = Store.query.get(to_store_id)
    
    if not from_store or not to_store:
        return jsonify({'error': 'Tienda no encontrada'}), 404
    
    items_text = request.form.get('items_text', '').strip()
    items_file = request.files.get('items_file')
    
    sku_qty_list = []
    
    if items_file and items_file.filename:
        try:
            filename = items_file.filename.lower()
            if filename.endswith('.csv'):
                import csv
                import io
                content = items_file.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(content))
                for row in reader:
                    sku = str(row.get('sku', row.get('SKU', ''))).strip()
                    qty_str = str(row.get('qty', row.get('QTY', row.get('cantidad', row.get('Cantidad', '0'))))).strip()
                    try:
                        qty = int(float(qty_str))
                    except:
                        qty = 0
                    if sku and qty > 0:
                        sku_qty_list.append((sku, qty))
            elif filename.endswith('.xlsx') or filename.endswith('.xls'):
                df = pd.read_excel(items_file)
                df.columns = [c.lower().strip() for c in df.columns]
                for _, row in df.iterrows():
                    sku = str(row.get('sku', '')).strip()
                    qty = int(row.get('qty', row.get('cantidad', 0)) or 0)
                    if sku and qty > 0:
                        sku_qty_list.append((sku, qty))
        except Exception as e:
            return jsonify({'error': f'Error leyendo archivo: {str(e)}'}), 400
    
    if items_text and not sku_qty_list:
        for line in items_text.split('\n'):
            line = line.strip()
            if not line:
                continue
            parts = line.replace(';', ',').split(',')
            if len(parts) >= 2:
                sku = parts[0].strip()
                try:
                    qty = int(float(parts[1].strip()))
                except:
                    qty = 0
                if sku and qty > 0:
                    sku_qty_list.append((sku, qty))
    
    if not sku_qty_list:
        return jsonify({'error': 'No se encontraron SKUs válidos en el archivo o texto'}), 400
    
    latest_date = db.session.query(db.func.max(StockSnapshot.as_of_date)).scalar()
    
    donor_stock = {}
    if latest_date:
        stock_rows = StockSnapshot.query.filter(
            StockSnapshot.store_id == from_store_id,
            StockSnapshot.as_of_date == latest_date
        ).all()
        for ss in stock_rows:
            if ss.product:
                donor_stock[ss.product.sku] = ss.quantity
    
    results = []
    for sku, qty in sku_qty_list:
        product = Product.query.filter_by(sku=sku).first()
        
        if not product:
            results.append({
                'sku': sku,
                'product_name': '-',
                'product_id': None,
                'qty': qty,
                'status': 'ERROR',
                'notes': 'SKU no encontrado',
                'donor_stock': 0
            })
            continue
        
        available = donor_stock.get(sku, 0)
        max_transferable = max(0, available - donor_floor)
        
        if qty > max_transferable:
            if available == 0:
                status = 'ERROR'
                notes = 'Sin stock en tienda origen'
            else:
                status = 'WARNING'
                notes = f'Excede stock disponible (max: {max_transferable})'
        else:
            status = 'OK'
            notes = ''
        
        results.append({
            'sku': sku,
            'product_name': product.name,
            'product_id': product.id,
            'qty': qty,
            'status': status,
            'notes': notes,
            'donor_stock': available
        })
    
    return jsonify({
        'from_store': from_store.name,
        'to_store': to_store.name,
        'from_store_id': from_store_id,
        'to_store_id': to_store_id,
        'items': results,
        'total_items': len(results),
        'ok_count': sum(1 for r in results if r['status'] == 'OK'),
        'warning_count': sum(1 for r in results if r['status'] == 'WARNING'),
        'error_count': sum(1 for r in results if r['status'] == 'ERROR')
    })


@app.route('/rebalancing/manual/add', methods=['POST'])
@login_required
@require_permission('rebalancing:run')
def manual_transfer_add():
    """Add validated items to the current manual transfer plan."""
    data = request.get_json()
    
    if not data:
        return jsonify({'error': 'Datos no proporcionados'}), 400
    
    from_store_id = data.get('from_store_id')
    to_store_id = data.get('to_store_id')
    items = data.get('items', [])
    
    if not from_store_id or not to_store_id:
        return jsonify({'error': 'Tienda origen y destino son requeridas'}), 400
    
    if from_store_id == to_store_id:
        return jsonify({'error': 'Tienda origen no puede ser igual a destino'}), 400
    
    valid_items = [i for i in items if i.get('product_id') and i.get('status') in ('OK', 'WARNING')]
    
    if not valid_items:
        return jsonify({'error': 'No hay items válidos para agregar'}), 400
    
    current_run = ManualTransferRun.query.filter_by(
        created_by_user_id=current_user.id,
        status='draft'
    ).first()
    
    if not current_run:
        run_id = str(uuid.uuid4())
        current_run = ManualTransferRun(
            run_id=run_id,
            created_by_user_id=current_user.id,
            status='draft'
        )
        db.session.add(current_run)
        db.session.flush()
    
    added_count = 0
    for item in valid_items:
        existing = ManualTransferItem.query.filter_by(
            run_id=current_run.run_id,
            product_id=item['product_id'],
            from_store_id=from_store_id,
            to_store_id=to_store_id
        ).first()
        
        if existing:
            existing.qty = item['qty']
            existing.status = item['status']
            existing.notes = item.get('notes', '')
        else:
            new_item = ManualTransferItem(
                run_id=current_run.run_id,
                product_id=item['product_id'],
                from_store_id=from_store_id,
                to_store_id=to_store_id,
                qty=item['qty'],
                status=item['status'],
                notes=item.get('notes', '')
            )
            db.session.add(new_item)
            added_count += 1
    
    db.session.commit()
    
    total_items = ManualTransferItem.query.filter_by(run_id=current_run.run_id).count()
    
    log_audit(
        action='manual_transfer.add',
        status='success',
        message=f'Added {added_count} items to manual transfer plan',
        entity_type='ManualTransferRun',
        run_id=current_run.run_id
    )
    
    return jsonify({
        'success': True,
        'run_id': current_run.run_id,
        'added_count': added_count,
        'total_items': total_items
    })


@app.route('/rebalancing/manual/items', methods=['GET'])
@login_required
@require_permission('rebalancing:view')
def manual_transfer_items():
    """Get current manual transfer plan items."""
    current_run = ManualTransferRun.query.filter_by(
        created_by_user_id=current_user.id,
        status='draft'
    ).first()
    
    if not current_run:
        return jsonify({'items': [], 'run_id': None})
    
    items = ManualTransferItem.query.filter_by(run_id=current_run.run_id).all()
    
    return jsonify({
        'run_id': current_run.run_id,
        'items': [{
            'id': i.id,
            'sku': i.product.sku if i.product else '',
            'product_name': i.product.name if i.product else '',
            'from_store': i.from_store.name if i.from_store else '',
            'to_store': i.to_store.name if i.to_store else '',
            'qty': i.qty,
            'status': i.status,
            'notes': i.notes
        } for i in items]
    })


@app.route('/rebalancing/manual/clear', methods=['POST'])
@login_required
@require_permission('rebalancing:run')
def manual_transfer_clear():
    """Clear the current manual transfer plan."""
    current_run = ManualTransferRun.query.filter_by(
        created_by_user_id=current_user.id,
        status='draft'
    ).first()
    
    if current_run:
        ManualTransferItem.query.filter_by(run_id=current_run.run_id).delete()
        db.session.delete(current_run)
        db.session.commit()
        
        log_audit(
            action='manual_transfer.clear',
            status='success',
            message='Cleared manual transfer plan',
            entity_type='ManualTransferRun',
            run_id=current_run.run_id
        )
    
    return jsonify({'success': True})


@app.route('/rebalancing/manual/export', methods=['GET'])
@login_required
@require_permission('rebalancing:view')
def manual_transfer_export():
    """Export manual transfer plan to Excel."""
    run_id = request.args.get('run_id', '').strip()
    
    if not run_id:
        current_run = ManualTransferRun.query.filter_by(
            created_by_user_id=current_user.id,
            status='draft'
        ).first()
        if current_run:
            run_id = current_run.run_id
    
    if not run_id:
        flash('No hay plan de transferencia manual activo.', 'warning')
        return redirect(url_for('rebalancing'))
    
    items = ManualTransferItem.query.filter_by(run_id=run_id).all()
    
    if not items:
        flash('El plan de transferencia está vacío.', 'warning')
        return redirect(url_for('rebalancing'))
    
    rows = []
    for i in items:
        rows.append({
            'SKU': str(i.product.sku) if i.product else '',
            'Producto': i.product.name if i.product else '',
            'Tienda Origen': i.from_store.name if i.from_store else '',
            'Tienda Destino': i.to_store.name if i.to_store else '',
            'Cantidad': i.qty,
            'Estado': i.status,
            'Notas': i.notes or ''
        })
    
    df = pd.DataFrame(rows)
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Transferencia Manual')
        
        workbook = writer.book
        worksheet = writer.sheets['Transferencia Manual']
        
        for col_cells in worksheet.columns:
            col_letter = col_cells[0].column_letter
            if col_letter == 'A':
                worksheet.column_dimensions[col_letter].number_format = '@'
                for cell in col_cells[1:]:
                    cell.number_format = '@'
    
    output.seek(0)
    
    filename = f"transferencia_manual_{run_id[:8]}_{date.today().isoformat()}.xlsx"
    
    log_audit(
        action='manual_transfer.export',
        status='success',
        message=f'Exported {len(rows)} manual transfer items',
        entity_type='ManualTransferRun',
        run_id=run_id
    )
    
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/rebalancing/manual/remove/<int:item_id>', methods=['POST'])
@login_required
@require_permission('rebalancing:run')
def manual_transfer_remove_item(item_id):
    """Remove a single item from the manual transfer plan."""
    item = ManualTransferItem.query.get(item_id)
    
    if not item:
        return jsonify({'error': 'Item no encontrado'}), 404
    
    run = ManualTransferRun.query.filter_by(run_id=item.run_id).first()
    if not run or run.created_by_user_id != current_user.id:
        return jsonify({'error': 'No tienes permiso para modificar este plan'}), 403
    
    db.session.delete(item)
    db.session.commit()
    
    return jsonify({'success': True})


@app.route('/rebalancing/manual/calculate', methods=['POST'])
@login_required
@require_permission('rebalancing:run')
def manual_transfer_calculate():
    """
    Assisted Manual Plan: Calculate transfer quantities for a list of SKUs.
    System determines optimal qty based on WOC logic - no user qty input needed.
    """
    to_store_id = request.form.get('to_store_id', type=int)
    from_store_id = request.form.get('from_store_id', type=int)
    
    weeks_window = request.form.get('weeks_window', 4, type=int)
    woc_min = request.form.get('woc_min', 1.5, type=float)
    woc_target = request.form.get('woc_target', 2.5, type=float)
    woc_max = request.form.get('woc_max', 6.0, type=float)
    retain_woc = request.form.get('retain_woc', 4.0, type=float)
    stock_floor = request.form.get('stock_floor', 1, type=int)
    min_transfer_qty = request.form.get('min_transfer_qty', 2, type=int)
    
    if not to_store_id:
        return jsonify({'error': 'Tienda destino es requerida'}), 400
    
    if from_store_id and from_store_id == to_store_id:
        return jsonify({'error': 'Tienda origen no puede ser igual a destino'}), 400
    
    to_store = Store.query.get(to_store_id)
    from_store = Store.query.get(from_store_id) if from_store_id else None
    
    if not to_store:
        return jsonify({'error': 'Tienda destino no encontrada'}), 404
    
    sku_text = request.form.get('sku_text', '').strip()
    sku_file = request.files.get('sku_file')
    
    sku_list = []
    
    if sku_file and sku_file.filename:
        try:
            filename = sku_file.filename.lower()
            if filename.endswith('.csv'):
                import csv
                import io
                content = sku_file.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(content))
                for row in reader:
                    sku = str(row.get('sku', row.get('SKU', ''))).strip()
                    if sku:
                        sku_list.append(sku)
            elif filename.endswith('.xlsx') or filename.endswith('.xls'):
                df = pd.read_excel(sku_file)
                df.columns = [c.lower().strip() for c in df.columns]
                for _, row in df.iterrows():
                    sku = str(row.get('sku', '')).strip()
                    if sku:
                        sku_list.append(sku)
        except Exception as e:
            return jsonify({'error': f'Error leyendo archivo: {str(e)}'}), 400
    
    if sku_text and not sku_list:
        for line in sku_text.split('\n'):
            line = line.strip().split(',')[0].split(';')[0].strip()
            if line:
                sku_list.append(line)
    
    if not sku_list:
        return jsonify({'error': 'No se encontraron SKUs válidos'}), 400
    
    sku_list = list(dict.fromkeys(sku_list))
    
    today = date.today()
    cutoff = today - timedelta(days=weeks_window * 7)
    
    latest_stock_date = db.session.query(db.func.max(StockSnapshot.as_of_date)).scalar()
    if not latest_stock_date:
        return jsonify({'error': 'No hay datos de stock disponibles'}), 400
    
    products_by_sku = {p.sku: p for p in Product.query.filter(Product.sku.in_(sku_list)).all()}
    product_ids = [p.id for p in products_by_sku.values()]
    
    stock_map = {}
    stock_q = (
        db.session.query(
            StockSnapshot.product_id,
            StockSnapshot.store_id,
            StockSnapshot.quantity
        )
        .filter(
            StockSnapshot.as_of_date == latest_stock_date,
            StockSnapshot.product_id.in_(product_ids)
        )
        .all()
    )
    for pid, sid, qty in stock_q:
        stock_map[(pid, sid)] = qty
    
    week_expr = db.func.strftime('%Y-%W', DistributionRecord.event_date)
    sales_q = (
        db.session.query(
            DistributionRecord.product_id,
            DistributionRecord.store_id,
            db.func.sum(DistributionRecord.quantity).label('total_qty'),
            db.func.count(db.func.distinct(week_expr)).label('num_weeks')
        )
        .filter(
            DistributionRecord.event_date >= cutoff,
            DistributionRecord.product_id.in_(product_ids)
        )
        .group_by(DistributionRecord.product_id, DistributionRecord.store_id)
        .all()
    )
    
    sales_rate_map = {}
    for pid, sid, total_qty, num_weeks in sales_q:
        rate = total_qty / max(num_weeks, 1)
        sales_rate_map[(pid, sid)] = rate
    
    store_info = {s.id: s.name for s in Store.query.all()}
    all_store_ids = list(store_info.keys())
    
    results = []
    
    for sku in sku_list:
        product = products_by_sku.get(sku)
        
        if not product:
            results.append({
                'sku': sku,
                'product_name': '-',
                'to_store': to_store.name,
                'donor_store': None,
                'transfer_qty': 0,
                'woc_dest': 0,
                'donor_give': 0,
                'status': 'NO-SKU',
                'notes': 'SKU no encontrado'
            })
            continue
        
        pid = product.id
        
        stock_dest = stock_map.get((pid, to_store_id), 0)
        rate_dest = sales_rate_map.get((pid, to_store_id), 0)
        woc_dest = stock_dest / max(rate_dest, 0.01) if rate_dest > 0 else 999
        
        if rate_dest == 0:
            results.append({
                'sku': sku,
                'product_name': product.name,
                'to_store': to_store.name,
                'donor_store': None,
                'transfer_qty': 0,
                'woc_dest': 0,
                'donor_give': 0,
                'status': 'NO-SALES',
                'notes': 'Sin historial de ventas en destino'
            })
            continue
        
        if woc_dest >= woc_min:
            need_units = 0
        else:
            need_units = max(math.ceil(woc_target * rate_dest) - stock_dest, 0)
        
        if need_units == 0:
            results.append({
                'sku': sku,
                'product_name': product.name,
                'to_store': to_store.name,
                'donor_store': None,
                'transfer_qty': 0,
                'woc_dest': round(woc_dest, 2),
                'donor_give': 0,
                'status': 'NO-NEED',
                'notes': f'WOC {woc_dest:.1f} >= {woc_min}'
            })
            continue
        
        candidate_donors = []
        search_store_ids = [from_store_id] if from_store_id else all_store_ids
        
        for sid in search_store_ids:
            if sid == to_store_id:
                continue
            
            stock_donor = stock_map.get((pid, sid), 0)
            if stock_donor <= 0:
                continue
            
            rate_donor = sales_rate_map.get((pid, sid), 0)
            keep_units = max(math.ceil(retain_woc * rate_donor), stock_floor) if rate_donor > 0 else stock_floor
            give_units = max(stock_donor - keep_units, 0)
            
            if give_units > 0:
                woc_donor = stock_donor / max(rate_donor, 0.01)
                candidate_donors.append({
                    'store_id': sid,
                    'store_name': store_info.get(sid, '?'),
                    'give': give_units,
                    'woc': woc_donor
                })
        
        candidate_donors.sort(key=lambda x: x['give'], reverse=True)
        
        if not candidate_donors:
            results.append({
                'sku': sku,
                'product_name': product.name,
                'to_store': to_store.name,
                'donor_store': from_store.name if from_store else None,
                'transfer_qty': 0,
                'woc_dest': round(woc_dest, 2),
                'donor_give': 0,
                'status': 'NO-DONOR',
                'notes': 'Sin tienda donante disponible'
            })
            continue
        
        best_donor = candidate_donors[0]
        transfer = min(need_units, best_donor['give'])
        
        is_stockout = stock_dest == 0 and rate_dest > 0
        if transfer < min_transfer_qty and not is_stockout:
            transfer = 0
            status = 'NO-NEED'
            notes = f'Transferencia ({transfer}) < mínimo ({min_transfer_qty})'
        else:
            status = 'OK'
            notes = ''
        
        results.append({
            'sku': sku,
            'product_name': product.name,
            'product_id': pid,
            'to_store': to_store.name,
            'to_store_id': to_store_id,
            'donor_store': best_donor['store_name'],
            'donor_store_id': best_donor['store_id'],
            'transfer_qty': transfer,
            'woc_dest': round(woc_dest, 2),
            'donor_give': best_donor['give'],
            'status': status,
            'notes': notes
        })
    
    ok_count = sum(1 for r in results if r['status'] == 'OK' and r.get('transfer_qty', 0) > 0)
    no_need_count = sum(1 for r in results if r['status'] == 'NO-NEED')
    no_donor_count = sum(1 for r in results if r['status'] in ('NO-DONOR', 'NO-SKU', 'NO-SALES'))
    
    return jsonify({
        'to_store': to_store.name,
        'from_store': from_store.name if from_store else None,
        'items': results,
        'total_items': len(results),
        'ok_count': ok_count,
        'no_need_count': no_need_count,
        'no_donor_count': no_donor_count
    })


@app.route('/rebalancing/manual/save', methods=['POST'])
@login_required
@require_permission('rebalancing:run')
def manual_transfer_save():
    """Save the calculated assisted manual plan to database."""
    data = request.get_json()
    
    if not data:
        return jsonify({'error': 'Datos no proporcionados'}), 400
    
    items = data.get('items', [])
    valid_items = [i for i in items if i.get('status') == 'OK' and i.get('transfer_qty', 0) > 0]
    
    if not valid_items:
        return jsonify({'error': 'No hay items válidos para guardar'}), 400
    
    run_id = str(uuid.uuid4())
    current_run = ManualTransferRun(
        run_id=run_id,
        created_by_user_id=current_user.id,
        status='complete'
    )
    db.session.add(current_run)
    
    count = 0
    for item in valid_items:
        mt_item = ManualTransferItem(
            run_id=run_id,
            product_id=item.get('product_id'),
            from_store_id=item.get('donor_store_id'),
            to_store_id=item.get('to_store_id'),
            qty=item.get('transfer_qty', 0),
            status='OK',
            notes=item.get('notes', '')
        )
        db.session.add(mt_item)
        count += 1
    
    db.session.commit()
    
    log_audit(
        action='manual_transfer.save_assisted',
        status='success',
        message=f'Saved assisted manual plan with {count} items',
        entity_type='ManualTransferRun',
        run_id=run_id
    )
    
    return jsonify({'success': True, 'count': count, 'run_id': run_id})


@app.route('/rebalancing/manual/export-calculated', methods=['POST'])
@login_required
@require_permission('rebalancing:run')
def manual_transfer_export_calculated():
    """Export the calculated assisted manual plan to Excel."""
    items_json = request.form.get('items', '[]')
    
    try:
        items = json.loads(items_json)
    except:
        items = []
    
    if not items:
        flash('No hay items para exportar', 'warning')
        return redirect(url_for('rebalancing'))
    
    rows = []
    for item in items:
        rows.append({
            'SKU': str(item.get('sku', '')),
            'Producto': item.get('product_name', ''),
            'Tienda Origen': item.get('donor_store', ''),
            'Tienda Destino': item.get('to_store', ''),
            'Cantidad': item.get('transfer_qty', 0),
            'WOC Destino': item.get('woc_dest', 0),
            'Donante Da': item.get('donor_give', 0),
            'Estado': item.get('status', '')
        })
    
    df = pd.DataFrame(rows)
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Plan Asistido')
        
        workbook = writer.book
        worksheet = writer.sheets['Plan Asistido']
        
        for col_cells in worksheet.columns:
            col_letter = col_cells[0].column_letter
            if col_letter == 'A':
                worksheet.column_dimensions[col_letter].number_format = '@'
                for cell in col_cells[1:]:
                    cell.number_format = '@'
    
    output.seek(0)
    
    to_store = items[0].get('to_store', 'unknown') if items else 'unknown'
    filename = f"plan_asistido_{to_store}_{date.today().isoformat()}.xlsx"
    
    log_audit(
        action='manual_transfer.export_calculated',
        status='success',
        message=f'Exported calculated plan with {len(rows)} items'
    )
    
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/export_simulation/<sim_type>', methods=['GET'])
@login_required
def export_simulation(sim_type):
    """Export simulation results to Excel with watermark."""
    sim_data = get_simulation_results(sim_type)
    
    if not sim_data:
        flash('No hay simulación activa para exportar.', 'warning')
        return redirect(url_for('dashboard'))
    
    results = sim_data.get('results', [])
    kpis = sim_data.get('kpis', {})
    meta = sim_data.get('meta', {})
    
    if sim_type == 'rebalancing':
        rows = []
        for s in results:
            rows.append({
                'SIMULACIÓN': 'SÍ - NO GUARDADO',
                'SKU': s.get('sku', ''),
                'Producto': s.get('name', ''),
                'Tienda Origen': s.get('from_store', ''),
                'Tienda Destino': s.get('to_store', ''),
                'Cantidad': s.get('qty', 0),
                'Venta/Sem Destino': round(s.get('sales_rate_to', 0), 2),
                'WOC Origen': round(s.get('woc_from', 0), 2),
                'WOC Destino': round(s.get('woc_to', 0), 2),
                'Score': round(s.get('score', 0), 2),
                'Razón': s.get('reason', '')
            })
        filename = f"SIMULACION_redistribucion_{date.today().isoformat()}.xlsx"
    elif sim_type == 'distribution':
        rows = []
        for r in results:
            rows.append({
                'SIMULACIÓN': 'SÍ - NO GUARDADO',
                'SKU': r.get('sku', ''),
                'Producto': r.get('product_name', ''),
                'Tienda': r.get('store', ''),
                'Cantidad Sugerida': r.get('quantity', 0),
                'Modelo': r.get('model_name', ''),
                'Semana Objetivo': r.get('target_week', '')
            })
        filename = f"SIMULACION_distribucion_{date.today().isoformat()}.xlsx"
    else:
        flash('Tipo de simulación no reconocido.', 'warning')
        return redirect(url_for('dashboard'))
    
    df = pd.DataFrame(rows)
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Simulación')
        
        workbook = writer.book
        worksheet = writer.sheets['Simulación']
        
        from openpyxl.styles import PatternFill, Font
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        for row in worksheet.iter_rows(min_row=2, max_row=len(rows)+1, min_col=1, max_col=1):
            for cell in row:
                cell.fill = yellow_fill
                cell.font = Font(bold=True, color='FF0000')
    
    output.seek(0)
    
    log_audit(
        action=f'simulation.export.{sim_type}',
        status='success',
        message=f'Exported simulation with {len(rows)} rows',
        metadata={'sim_type': sim_type, 'rows': len(rows)}
    )
    
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/clear_simulation/<sim_type>', methods=['GET'])
@login_required
def clear_simulation_route(sim_type):
    """Clear simulation data and redirect."""
    if sim_type == 'all':
        clear_simulation(None)
        flash('Todas las simulaciones limpiadas.', 'info')
    else:
        clear_simulation(sim_type)
        flash('Simulación limpiada.', 'info')
    redirect_to = request.args.get('redirect', 'dashboard')
    return redirect(url_for(redirect_to))


# ------------------ Slow Stock & Smart Reallocation Module ------------------

# Default parameters
SLOW_STOCK_PARAMS = {
    'HISTORY_WINDOW_WEEKS': 12,
    'RECENT_WINDOW_WEEKS': 4,
    'DEAD_DAYS_STORE': 60,
    'DEAD_DAYS_GLOBAL': 90,
    'SLOW_MIN_DAYS_STORE': 21,
    'DEAD_PURCHASE_DAYS': 120,
    'SLOW_RATE_THRESHOLD': 0.3,
    'MIN_WOC': 1.5,
    'MAX_WOC': 6.0,
    'RECEIVER_MIN_RATE': 0.5,
    'RECEIVER_WOC_MIN': 1.5,
    'TARGET_RECEIVER_WOC': 2.5,
    'DONOR_FLOOR': 1,
    'MIN_TRANSFER_QTY': 2
}


def compute_slow_stock_analysis(params=None):
    """
    Compute slow/dead stock analysis for stores and CD.
    Returns dict with store_analysis, cd_analysis, and transfer_suggestions.
    """
    if params is None:
        params = SLOW_STOCK_PARAMS.copy()
    
    today = date.today()
    
    # Get latest stock snapshots
    latest_store_date = db.session.query(db.func.max(StockSnapshot.as_of_date)).scalar()
    latest_cd_date = db.session.query(db.func.max(StockCD.as_of_date)).scalar()
    
    if not latest_store_date or not latest_cd_date:
        return {'store_analysis': [], 'cd_analysis': [], 'transfers': [], 'error': 'No stock data available'}
    
    # Build product and store lookups
    products = {p.id: p for p in Product.query.all()}
    stores = {s.id: s for s in Store.query.all()}
    
    # Get store stock (latest snapshot)
    store_stock = {}
    for ss in StockSnapshot.query.filter(StockSnapshot.as_of_date == latest_store_date).all():
        store_stock[(ss.product_id, ss.store_id)] = ss.quantity
    
    # Get CD stock (latest snapshot)
    cd_stock = {}
    for cd in StockCD.query.filter(StockCD.as_of_date == latest_cd_date).all():
        cd_stock[cd.product_id] = cd.quantity
    
    # Get lifecycle data
    lifecycle_global = {lc.product_id: lc for lc in SkuLifecycle.query.all()}
    lifecycle_store = {}
    for lcs in SkuStoreLifecycle.query.all():
        lifecycle_store[(lcs.product_id, lcs.store_id)] = lcs
    
    # Compute sales rates from SalesWeeklyAgg (Macro Sales)
    recent_window = params['RECENT_WINDOW_WEEKS']
    window_start_week = today - timedelta(weeks=recent_window)
    window_start_week = window_start_week - timedelta(days=window_start_week.weekday())
    
    app.logger.info(f"Slow Stock using SalesWeeklyAgg from {window_start_week} to {today}")
    
    # Sales rate per SKU-store from SalesWeeklyAgg
    sales_query = db.session.query(
        SalesWeeklyAgg.product_id,
        SalesWeeklyAgg.store_id,
        db.func.sum(SalesWeeklyAgg.units).label('total_units'),
        db.func.count(db.func.distinct(SalesWeeklyAgg.week_start)).label('weeks_count')
    ).filter(
        SalesWeeklyAgg.week_start >= window_start_week
    ).group_by(
        SalesWeeklyAgg.product_id,
        SalesWeeklyAgg.store_id
    ).all()
    
    sales_rate = {}
    for row in sales_query:
        weekly_rate = row.total_units / max(row.weeks_count, 1)
        sales_rate[(row.product_id, row.store_id)] = weekly_rate
    
    # Compute last sale from SalesWeeklyAgg max week with units>0 if not in lifecycle
    last_sale_agg = db.session.query(
        SalesWeeklyAgg.product_id,
        SalesWeeklyAgg.store_id,
        db.func.max(SalesWeeklyAgg.week_start).label('last_week')
    ).filter(
        SalesWeeklyAgg.units > 0
    ).group_by(
        SalesWeeklyAgg.product_id,
        SalesWeeklyAgg.store_id
    ).all()
    
    last_sale_store_fallback = {}
    last_sale_global_fallback = {}
    for row in last_sale_agg:
        last_week = row.last_week
        if last_week:
            last_sale_store_fallback[(row.product_id, row.store_id)] = last_week
            if row.product_id not in last_sale_global_fallback or last_week > last_sale_global_fallback[row.product_id]:
                last_sale_global_fallback[row.product_id] = last_week
    
    # Classify store-level status
    store_analysis = []
    donors = {}  # product_id -> list of (store_id, status, stock)
    
    for (pid, sid), qty in store_stock.items():
        if qty <= 0:
            continue
        
        product = products.get(pid)
        store = stores.get(sid)
        if not product or not store:
            continue
        
        # Get last sale date for this SKU-store
        lcs = lifecycle_store.get((pid, sid))
        if lcs and lcs.last_sale_date_store:
            last_sale = lcs.last_sale_date_store
        else:
            last_sale = last_sale_store_fallback.get((pid, sid))
        
        days_since_last_sale = (today - last_sale).days if last_sale else 9999
        rate = sales_rate.get((pid, sid), 0)
        
        # Calculate weeks of coverage (WOC)
        woc = qty / max(rate, 0.01) if rate > 0 else 9999
        
        # Classification: DEAD_STORE, SLOW_STORE, LOW_STOCK, HEALTHY_STORE
        # Priority: DEAD > SLOW > LOW_STOCK > HEALTHY
        if days_since_last_sale >= params['DEAD_DAYS_STORE']:
            status = 'DEAD_STORE'
        elif rate <= params['SLOW_RATE_THRESHOLD']:
            # Low sales rate means SLOW regardless of days
            status = 'SLOW_STORE'
        elif woc > params['MAX_WOC']:
            # Overstock (high WOC) is also SLOW
            status = 'SLOW_STORE'
        elif rate > params['SLOW_RATE_THRESHOLD'] and woc < params['MIN_WOC']:
            # Active sales (above slow threshold) but low coverage
            status = 'LOW_STOCK'
        else:
            status = 'HEALTHY_STORE'
        
        # Include all SKU-store pairs with stock
        store_analysis.append({
            'sku': product.sku,
            'product_name': product.name,
            'store': store.name,
            'stock': qty,
            'days_since_last_sale': days_since_last_sale,
            'sales_rate': round(rate, 2),
            'coverage_weeks': round(woc, 1) if woc < 9999 else None,
            'status': status,
            'product_id': pid,
            'store_id': sid
        })
        
        # Track donors (DEAD or SLOW only)
        if status in ['DEAD_STORE', 'SLOW_STORE']:
            if pid not in donors:
                donors[pid] = []
            donors[pid].append((sid, status, qty))
    
    # Classify CD-level status
    cd_analysis = []
    for pid, qty in cd_stock.items():
        if qty <= 0:
            continue
        
        product = products.get(pid)
        if not product:
            continue
        
        # Get global lifecycle data
        lc = lifecycle_global.get(pid)
        last_sale_global = lc.last_sale_date_global if lc and lc.last_sale_date_global else last_sale_global_fallback.get(pid)
        last_purchase = lc.last_purchase_date if lc else None
        
        days_since_last_sale_global = (today - last_sale_global).days if last_sale_global else 9999
        days_since_last_purchase = (today - last_purchase).days if last_purchase else 9999
        
        # Calculate CD WOC for this SKU (CD stock only, not total stock)
        global_rate = sum(sales_rate.get((pid, sid), 0) for sid in stores.keys())
        woc_cd = qty / max(global_rate, 0.01) if global_rate > 0 else 9999
        
        # Classification: DEAD_CD, SLOW_CD, HEALTHY_CD
        if days_since_last_sale_global >= params['DEAD_DAYS_GLOBAL'] or days_since_last_purchase >= params.get('DEAD_PURCHASE_DAYS', 120):
            status = 'DEAD_CD'
        elif global_rate <= params['SLOW_RATE_THRESHOLD'] or woc_cd > params['MAX_WOC']:
            status = 'SLOW_CD'
        else:
            status = 'HEALTHY_CD'
        
        # Check if there are active receivers for this SKU
        has_receivers = any(
            sales_rate.get((pid, sid), 0) >= params['RECEIVER_MIN_RATE']
            for sid in stores.keys()
        )
        
        recommendation = ''
        if status == 'DEAD_CD':
            if has_receivers:
                recommendation = 'Redistribuir a tiendas'
            else:
                recommendation = 'Considerar liquidación'
        elif status == 'SLOW_CD':
            recommendation = 'Revisar movimiento'
        
        # Only add non-healthy CD entries (DEAD_CD, SLOW_CD)
        if status in ['DEAD_CD', 'SLOW_CD']:
            cd_analysis.append({
                'sku': product.sku,
                'product_name': product.name,
                'stock_cd': qty,
                'days_since_last_sale_global': days_since_last_sale_global if days_since_last_sale_global < 9999 else None,
                'days_since_last_purchase': days_since_last_purchase if days_since_last_purchase < 9999 else None,
                'global_rate': round(global_rate, 2),
                'woc_cd': round(woc_cd, 1) if woc_cd < 9999 else None,
                'status': status,
                'recommendation': recommendation,
                'product_id': pid
            })
    
    # Generate transfer suggestions
    transfers = []
    
    for pid, donor_list in donors.items():
        product = products.get(pid)
        if not product:
            continue
        
        # Find receivers for this SKU
        receivers = []
        for sid in stores.keys():
            rate = sales_rate.get((pid, sid), 0)
            stock = store_stock.get((pid, sid), 0)
            woc = stock / max(rate, 0.01)
            
            # Qualify as receiver
            if rate >= params['RECEIVER_MIN_RATE'] or (rate > 0 and woc < params['RECEIVER_WOC_MIN']):
                need = max(int(params['TARGET_RECEIVER_WOC'] * rate) - stock, 0)
                if need > 0:
                    receivers.append((sid, rate, stock, woc, need))
        
        # Sort receivers by sales rate descending
        receivers.sort(key=lambda x: -x[1])
        
        # Sort donors by available qty descending
        donor_list.sort(key=lambda x: -x[2])
        
        for donor_sid, donor_status, donor_stock in donor_list:
            available = max(donor_stock - params['DONOR_FLOOR'], 0)
            if available < params['MIN_TRANSFER_QTY']:
                continue
            
            for i, (recv_sid, recv_rate, recv_stock, recv_woc, recv_need) in enumerate(receivers):
                if available <= 0:
                    break
                if recv_need <= 0:
                    continue
                if recv_sid == donor_sid:
                    continue
                
                # Determine transfer qty
                transfer_qty = min(available, recv_need)
                if transfer_qty < params['MIN_TRANSFER_QTY']:
                    # Allow qty 1 if receiver has 0 stock and has sales history
                    if recv_stock == 0 and recv_rate > 0:
                        transfer_qty = min(available, recv_need, 1) if transfer_qty >= 1 else 0
                    else:
                        continue
                
                if transfer_qty > 0:
                    transfers.append({
                        'sku': product.sku,
                        'product_name': product.name,
                        'from_store': stores[donor_sid].name,
                        'to_store': stores[recv_sid].name,
                        'qty': transfer_qty,
                        'donor_status': donor_status,
                        'receiver_sales_rate': round(recv_rate, 2),
                        'receiver_woc': round(recv_woc, 2),
                        'reason': f'{donor_status} → demanda activa',
                        'product_id': pid,
                        'from_store_id': donor_sid,
                        'to_store_id': recv_sid
                    })
                    
                    available -= transfer_qty
                    receivers[i] = (recv_sid, recv_rate, recv_stock + transfer_qty, 
                                   (recv_stock + transfer_qty) / max(recv_rate, 0.01),
                                   max(recv_need - transfer_qty, 0))
    
    # Build global-level analysis (aggregated per SKU)
    global_analysis = []
    sku_global_data = {}
    
    # Get all unique product IDs from both store_stock and cd_stock
    all_pids = set()
    for (pid, sid) in store_stock.keys():
        all_pids.add(pid)
    for pid in cd_stock.keys():
        all_pids.add(pid)
    
    # Aggregate store stock and sales per SKU
    for (pid, sid), qty in store_stock.items():
        if pid not in sku_global_data:
            sku_global_data[pid] = {'store_stock': 0, 'total_sales_rate': 0}
        sku_global_data[pid]['store_stock'] += qty
        sku_global_data[pid]['total_sales_rate'] += sales_rate.get((pid, sid), 0)
    
    # Ensure CD-only SKUs are included in global data
    for pid in all_pids:
        if pid not in sku_global_data:
            sku_global_data[pid] = {'store_stock': 0, 'total_sales_rate': 0}
    
    for pid, data in sku_global_data.items():
        product = products.get(pid)
        if not product:
            continue
        
        cd_qty = cd_stock.get(pid, 0)
        total_stock = data['store_stock'] + cd_qty
        total_rate = data['total_sales_rate']
        
        if total_stock <= 0:
            continue
        
        last_sale_global = last_sale_global_fallback.get(pid)
        days_since_last_sale = (today - last_sale_global).days if last_sale_global else 9999
        woc = total_stock / max(total_rate, 0.01) if total_rate > 0 else 9999
        
        # Global classification
        status = 'HEALTHY'
        if days_since_last_sale >= params['DEAD_DAYS_GLOBAL']:
            status = 'DEAD'
        elif woc > params['MAX_WOC']:
            status = 'SLOW'
        elif total_rate <= 0:
            status = 'DEAD'
        
        global_analysis.append({
            'sku': product.sku,
            'product_name': product.name,
            'store': 'GLOBAL',
            'stock': total_stock,
            'stock_cd': cd_qty,
            'stock_stores': data['store_stock'],
            'days_since_last_sale': days_since_last_sale if days_since_last_sale < 9999 else None,
            'sales_rate': round(total_rate, 2),
            'coverage_weeks': round(woc, 1) if woc < 9999 else None,
            'status': status,
            'product_id': pid
        })
    
    return {
        'store_analysis': store_analysis,
        'cd_analysis': cd_analysis,
        'transfers': transfers,
        'global_analysis': global_analysis
    }


def paginate_list(items, page, per_page=10):
    """
    Paginate a Python list, returning a dict with pagination metadata.
    Compatible with SQLAlchemy pagination interface for template consistency.
    """
    total = len(items)
    pages = max((total + per_page - 1) // per_page, 1)
    page = max(1, min(page, pages))
    start = (page - 1) * per_page
    end = start + per_page
    sliced = items[start:end]
    
    class ListPagination:
        def __init__(self, items, page, pages, total, per_page):
            self.items = items
            self.page = page
            self.pages = pages
            self.total = total
            self.per_page = per_page
            self.has_prev = page > 1
            self.has_next = page < pages
            self.prev_num = page - 1 if page > 1 else None
            self.next_num = page + 1 if page < pages else None
        
        def iter_pages(self, left_edge=2, left_current=2, right_current=3, right_edge=2):
            """Generate page numbers for pagination display."""
            last = 0
            for num in range(1, self.pages + 1):
                if num <= left_edge or \
                   (self.page - left_current <= num <= self.page + right_current) or \
                   num > self.pages - right_edge:
                    if last + 1 != num:
                        yield None
                    yield num
                    last = num
    
    return ListPagination(sliced, page, pages, total, per_page)


@app.route('/slow_stock', methods=['GET', 'POST'])
@login_required
@require_permission('slow_stock:view')
def slow_stock():
    """Slow Stock & Smart Reallocation module."""
    products = Product.query.order_by(Product.sku).all()
    stores = Store.query.order_by(Store.name).all()
    
    # Get latest run
    latest_run = SlowStockRun.query.order_by(SlowStockRun.created_at.desc()).first()
    
    results = None
    run_info = None
    
    if request.method == 'POST' and current_user.has_permission('slow_stock:run'):
        # Get parameters from form
        params = SLOW_STOCK_PARAMS.copy()
        try:
            params['HISTORY_WINDOW_WEEKS'] = int(request.form.get('history_window_weeks', 12))
            params['RECENT_WINDOW_WEEKS'] = int(request.form.get('recent_window_weeks', 4))
            params['DEAD_DAYS_STORE'] = int(request.form.get('dead_days_store', 60))
            params['DEAD_DAYS_GLOBAL'] = int(request.form.get('dead_days_global', 90))
            params['DEAD_PURCHASE_DAYS'] = int(request.form.get('dead_purchase_days', 120))
            params['SLOW_MIN_DAYS_STORE'] = int(request.form.get('slow_min_days_store', 21))
            params['SLOW_RATE_THRESHOLD'] = float(request.form.get('slow_rate_threshold', 0.3))
            params['MIN_WOC'] = float(request.form.get('min_woc', 1.5))
            params['MAX_WOC'] = float(request.form.get('max_woc', 6.0))
            params['RECEIVER_MIN_RATE'] = float(request.form.get('receiver_min_rate', 0.5))
            params['TARGET_RECEIVER_WOC'] = float(request.form.get('target_receiver_woc', 2.5))
        except ValueError:
            flash('Parámetros inválidos.', 'warning')
            return redirect(url_for('slow_stock'))
        
        # Run analysis
        results = compute_slow_stock_analysis(params)
        
        if 'error' in results:
            flash(results['error'], 'warning')
            return redirect(url_for('slow_stock'))
        
        # Save run
        run_id = str(uuid.uuid4())
        new_run = SlowStockRun(
            run_id=run_id,
            created_by_user_id=current_user.id,
            params_json=json.dumps(params),
            dead_store_count=len([r for r in results['store_analysis'] if r['status'] == 'DEAD_STORE']),
            slow_store_count=len([r for r in results['store_analysis'] if r['status'] == 'SLOW_STORE']),
            transfer_count=len(results['transfers']),
            dead_cd_count=len(results['cd_analysis'])
        )
        db.session.add(new_run)
        
        # Save suggestions
        for t in results['transfers']:
            sug = SlowStockSuggestion(
                run_id=run_id,
                product_id=t['product_id'],
                from_store_id=t['from_store_id'],
                to_store_id=t['to_store_id'],
                qty=t['qty'],
                donor_status=t['donor_status'],
                receiver_sales_rate=t['receiver_sales_rate'],
                receiver_woc=t['receiver_woc'],
                reason=t['reason']
            )
            db.session.add(sug)
        
        db.session.commit()
        
        log_audit(
            action='slow_stock.run',
            status='success',
            message=f'Analysis completed: {new_run.dead_store_count} dead, {new_run.slow_store_count} slow, {new_run.transfer_count} transfers',
            run_id=run_id,
            metadata=params
        )
        
        run_info = {
            'run_id': run_id,
            'created_at': new_run.created_at,
            'params': params
        }
        
        flash(f'Análisis completado. {new_run.dead_store_count} muerto(s), {new_run.slow_store_count} lento(s), {new_run.transfer_count} transferencia(s).', 'success')
    
    elif latest_run:
        # Load latest run results
        params = latest_run.get_params() or SLOW_STOCK_PARAMS.copy()
        results = compute_slow_stock_analysis(params)
        run_info = {
            'run_id': latest_run.run_id,
            'created_at': latest_run.created_at,
            'params': params
        }
    
    # Filter parameters
    sku_search = request.args.get('sku_search', '').strip()
    store_filter = request.args.get('store_filter', '').strip()
    status_filter = request.args.get('status_filter', '').strip()
    flagged_only = request.args.get('flagged', '').strip() == '1'
    category_filter = request.args.get('category', '').strip()
    
    # Pagination parameters (fixed 10 per page)
    slow_page = request.args.get('slow_page', 1, type=int)
    sug_page = request.args.get('sug_page', 1, type=int)
    cd_page = request.args.get('cd_page', 1, type=int)
    global_page = request.args.get('global_page', 1, type=int)
    
    # Paginate results with filters
    slow_pag = None
    sug_pag = None
    cd_pag = None
    global_pag = None
    
    # Build flagged product lookup if flagged filter is active
    flagged_product_ids = set()
    if flagged_only:
        flagged_prods = Product.query.filter(Product.eligible_for_distribution == False).all()
        flagged_product_ids = {p.id for p in flagged_prods}
    
    # Build category product lookup if category filter is active
    category_product_ids = set()
    if category_filter:
        cat_prods = Product.query.filter(Product.category == category_filter).all()
        category_product_ids = {p.id for p in cat_prods}
    
    if results:
        # Apply filters to store analysis
        store_data = results.get('store_analysis', [])
        if sku_search:
            store_data = [r for r in store_data if sku_search.lower() in r['sku'].lower() or sku_search.lower() in r['product_name'].lower()]
        if store_filter:
            store_data = [r for r in store_data if r['store'] == store_filter]
        if status_filter:
            store_data = [r for r in store_data if r['status'] == status_filter]
        if flagged_only:
            store_data = [r for r in store_data if r.get('product_id') in flagged_product_ids]
        if category_filter:
            store_data = [r for r in store_data if r.get('product_id') in category_product_ids]
        
        slow_pag = paginate_list(store_data, slow_page, 10)
        sug_pag = paginate_list(results.get('transfers', []), sug_page, 10)
        cd_pag = paginate_list(results.get('cd_analysis', []), cd_page, 10)
        
        # Apply filters to global analysis
        global_data = results.get('global_analysis', [])
        if sku_search:
            global_data = [r for r in global_data if sku_search.lower() in r['sku'].lower() or sku_search.lower() in r['product_name'].lower()]
        if status_filter:
            # Normalize filter values: DEAD_STORE/DEAD_CD -> DEAD, SLOW_STORE -> SLOW
            global_status_map = {'DEAD_STORE': 'DEAD', 'DEAD_CD': 'DEAD', 'SLOW_STORE': 'SLOW', 'HEALTHY': 'HEALTHY', 'DEAD': 'DEAD', 'SLOW': 'SLOW'}
            normalized_status = global_status_map.get(status_filter, status_filter)
            global_data = [r for r in global_data if r['status'] == normalized_status]
        if flagged_only:
            global_data = [r for r in global_data if r.get('product_id') in flagged_product_ids]
        if category_filter:
            global_data = [r for r in global_data if r.get('product_id') in category_product_ids]
        
        global_pag = paginate_list(global_data, global_page, 10)
    
    # Get flagged products for display
    flagged_products = Product.query.filter(Product.eligible_for_distribution == False).order_by(Product.sku).all()
    
    return render_template('slow_stock.html',
                           products=products,
                           stores=stores,
                           results=results,
                           run_info=run_info,
                           default_params=SLOW_STOCK_PARAMS,
                           slow_pag=slow_pag,
                           sug_pag=sug_pag,
                           cd_pag=cd_pag,
                           global_pag=global_pag,
                           slow_page=slow_page,
                           sug_page=sug_page,
                           cd_page=cd_page,
                           global_page=global_page,
                           sku_search=sku_search,
                           store_filter=store_filter,
                           status_filter=status_filter,
                           flagged_only=flagged_only,
                           category_filter=category_filter,
                           flagged_products=flagged_products)


@app.route('/slow_stock/upload_lifecycle', methods=['POST'])
@login_required
@require_permission('slow_stock:run')
def upload_lifecycle():
    """Upload lifecycle data (last purchase/sale dates)."""
    if 'file' not in request.files:
        flash('No se seleccionó archivo.', 'warning')
        return redirect(url_for('slow_stock'))
    
    file = request.files['file']
    if file.filename == '':
        flash('No se seleccionó archivo.', 'warning')
        return redirect(url_for('slow_stock'))
    
    upload_type = request.form.get('upload_type', 'global')  # 'global' or 'store'
    
    try:
        if file.filename.endswith('.csv'):
            df = pd.read_csv(file, dtype={'sku': str})
        else:
            df = pd.read_excel(file, dtype={'sku': str})
        
        df.columns = [c.strip().lower() for c in df.columns]
        
        if 'sku' not in df.columns:
            flash('El archivo debe contener la columna "sku".', 'warning')
            return redirect(url_for('slow_stock'))
        
        # Build product lookup
        products = {p.sku: p.id for p in Product.query.all()}
        stores_lookup = {s.name: s.id for s in Store.query.all()}
        
        updated = 0
        
        if upload_type == 'global':
            # Global lifecycle: sku, last_purchase_date, last_sale_date_global
            for _, row in df.iterrows():
                sku = str(row['sku']).strip()
                if sku not in products:
                    continue
                
                pid = products[sku]
                
                lc = SkuLifecycle.query.filter_by(product_id=pid).first()
                if not lc:
                    lc = SkuLifecycle(product_id=pid)
                    db.session.add(lc)
                
                if 'last_purchase_date' in df.columns and pd.notna(row['last_purchase_date']):
                    lc.last_purchase_date = pd.to_datetime(row['last_purchase_date']).date()
                if 'last_sale_date_global' in df.columns and pd.notna(row['last_sale_date_global']):
                    lc.last_sale_date_global = pd.to_datetime(row['last_sale_date_global']).date()
                
                updated += 1
        
        else:
            # Store-level: sku, store, last_sale_date_store
            if 'store' not in df.columns:
                flash('El archivo debe contener la columna "store".', 'warning')
                return redirect(url_for('slow_stock'))
            
            for _, row in df.iterrows():
                sku = str(row['sku']).strip()
                store_name = str(row['store']).strip()
                
                if sku not in products or store_name not in stores_lookup:
                    continue
                
                pid = products[sku]
                sid = stores_lookup[store_name]
                
                lcs = SkuStoreLifecycle.query.filter_by(product_id=pid, store_id=sid).first()
                if not lcs:
                    lcs = SkuStoreLifecycle(product_id=pid, store_id=sid)
                    db.session.add(lcs)
                
                if 'last_sale_date_store' in df.columns and pd.notna(row['last_sale_date_store']):
                    lcs.last_sale_date_store = pd.to_datetime(row['last_sale_date_store']).date()
                
                updated += 1
        
        db.session.commit()
        
        log_audit(
            action='slow_stock.lifecycle_upload',
            status='success',
            message=f'Uploaded {upload_type} lifecycle data: {updated} records',
            metadata={'upload_type': upload_type, 'records': updated}
        )
        
        flash(f'Datos de ciclo de vida cargados: {updated} registros.', 'success')
    
    except Exception as e:
        db.session.rollback()
        flash(f'Error al procesar archivo: {str(e)}', 'danger')
        log_audit(
            action='slow_stock.lifecycle_upload',
            status='fail',
            message=str(e)
        )
    
    return redirect(url_for('slow_stock'))


@app.route('/export_slow_stock/<export_type>')
@login_required
@require_permission('slow_stock:view')
def export_slow_stock(export_type):
    """Export slow stock analysis results."""
    params = SLOW_STOCK_PARAMS.copy()
    results = compute_slow_stock_analysis(params)
    
    if 'error' in results:
        flash(results['error'], 'warning')
        return redirect(url_for('slow_stock'))
    
    if export_type == 'store':
        data = results['store_analysis']
        filename = f'stock_lento_tiendas_{date.today().isoformat()}.xlsx'
        sheet_name = 'Stock Lento Tiendas'
    elif export_type == 'transfers':
        data = results['transfers']
        filename = f'sugerencias_transferencia_{date.today().isoformat()}.xlsx'
        sheet_name = 'Transferencias'
    elif export_type == 'cd':
        data = results['cd_analysis']
        filename = f'stock_muerto_cd_{date.today().isoformat()}.xlsx'
        sheet_name = 'Stock CD'
    else:
        flash('Tipo de exportación no válido.', 'warning')
        return redirect(url_for('slow_stock'))
    
    # Remove internal IDs from export
    export_data = []
    for row in data:
        clean_row = {k: v for k, v in row.items() if not k.endswith('_id')}
        export_data.append(clean_row)
    
    df = pd.DataFrame(export_data)
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    output.seek(0)
    
    log_audit(
        action='slow_stock.export',
        status='success',
        message=f'Exported {export_type}: {len(data)} rows',
        metadata={'export_type': export_type, 'rows': len(data)}
    )
    
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ------------------ Slow Stock Flagging Module ------------------

def compute_risk_score(product_id):
    """
    Compute risk score (0-100) for a SKU based on lifecycle and stock data.
    Returns: (risk_score, risk_reason)
    """
    import math
    
    today = date.today()
    product = db.session.get(Product, product_id)
    if not product:
        return 0, "Producto no encontrado"
    
    # Get lifecycle data
    lifecycle = SkuLifecycle.query.filter_by(product_id=product_id).first()
    
    # Get days since last sale (from lifecycle or SalesWeeklyAgg)
    days_since_last_sale = None
    if lifecycle and lifecycle.last_sale_date_global:
        days_since_last_sale = (today - lifecycle.last_sale_date_global).days
    else:
        # Fallback to SalesWeeklyAgg
        last_sale = db.session.query(func.max(SalesWeeklyAgg.week_start)).filter(
            SalesWeeklyAgg.product_id == product_id,
            SalesWeeklyAgg.units > 0
        ).scalar()
        if last_sale:
            days_since_last_sale = (today - last_sale).days
    
    # Get days since last purchase
    days_since_last_purchase = 0
    if lifecycle and lifecycle.last_purchase_date:
        days_since_last_purchase = (today - lifecycle.last_purchase_date).days
    
    # Get total stock (CD + stores)
    stock_cd = db.session.query(func.coalesce(func.sum(StockCD.quantity), 0)).filter(
        StockCD.product_id == product_id
    ).scalar() or 0
    
    stock_stores = db.session.query(func.coalesce(func.sum(StockSnapshot.quantity), 0)).filter(
        StockSnapshot.product_id == product_id
    ).scalar() or 0
    
    total_stock = int(stock_cd) + int(stock_stores)
    
    # Calculate score components
    score_sale = 0
    if days_since_last_sale is not None:
        score_sale = min(100, (days_since_last_sale / 120) * 100)
    else:
        score_sale = 100  # No sales ever = max score for sale component
    
    score_stock = min(40, math.log10(total_stock + 1) * 10)
    score_purchase = min(30, (days_since_last_purchase / 180) * 30) if days_since_last_purchase > 0 else 0
    
    risk_score = max(0, min(100, int(score_sale + score_stock + score_purchase)))
    
    # Build reason text
    reasons = []
    if days_since_last_sale is not None:
        reasons.append(f"Sin ventas hace {days_since_last_sale} días")
    else:
        reasons.append("Sin registro de ventas")
    reasons.append(f"Stock total: {total_stock:,}")
    if days_since_last_purchase > 0:
        reasons.append(f"Última compra hace {days_since_last_purchase} días")
    
    risk_reason = ", ".join(reasons)
    
    return risk_score, risk_reason


def get_risk_band(score):
    """Return risk band label based on score."""
    if score >= 70:
        return 'HIGH'
    elif score >= 40:
        return 'MEDIUM'
    return 'LOW'


@app.route('/slow_stock/flag', methods=['POST'])
@login_required
@require_permission('slow_stock:run')
def flag_slow_stock():
    """Flag a product as managed by Slow Stock (exclude from distribution)."""
    sku = request.form.get('sku', '').strip()
    product_id = request.form.get('product_id')
    
    # Find product
    product = None
    if product_id:
        product = db.session.get(Product, int(product_id))
    elif sku:
        product = Product.query.filter_by(sku=sku).first()
    
    if not product:
        flash('Producto no encontrado.', 'warning')
        return redirect(request.referrer or url_for('slow_stock'))
    
    # Store previous state for audit
    prev_eligible = product.eligible_for_distribution
    
    # Compute risk score
    risk_score, risk_reason = compute_risk_score(product.id)
    
    # Update product
    product.eligible_for_distribution = False
    product.risk_score = risk_score
    product.risk_reason = risk_reason
    
    try:
        db.session.commit()
        
        log_audit(
            action='slow_stock.flag',
            status='success',
            message=f'SKU {product.sku} flagged for slow stock management',
            entity_type='Product',
            entity_id=product.id,
            metadata={
                'sku': product.sku,
                'risk_score': risk_score,
                'risk_reason': risk_reason,
                'prev_eligible_for_distribution': prev_eligible,
                'new_eligible_for_distribution': False
            }
        )
        
        flash(f'SKU {product.sku} marcado para gestión de stock lento (Riesgo: {risk_score}).', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al marcar producto: {str(e)}', 'danger')
        log_audit(
            action='slow_stock.flag',
            status='fail',
            message=str(e),
            entity_type='Product',
            entity_id=product.id
        )
    
    return redirect(request.referrer or url_for('slow_stock'))


@app.route('/slow_stock/flag_bulk', methods=['POST'])
@login_required
@require_permission('slow_stock:run')
def flag_slow_stock_bulk():
    """Bulk flag products as managed by Slow Stock."""
    skus = request.form.getlist('skus')
    product_ids_raw = request.form.get('product_ids', '')
    
    # Parse comma-separated product IDs
    product_ids = []
    if product_ids_raw:
        product_ids = [pid.strip() for pid in product_ids_raw.split(',') if pid.strip()]
    
    # Collect products
    products = []
    if product_ids:
        for pid in product_ids:
            try:
                p = db.session.get(Product, int(pid))
                if p:
                    products.append(p)
            except (ValueError, TypeError):
                continue
    elif skus:
        for sku in skus:
            p = Product.query.filter_by(sku=sku.strip()).first()
            if p:
                products.append(p)
    
    if not products:
        flash('No se encontraron productos para marcar.', 'warning')
        return redirect(request.referrer or url_for('slow_stock'))
    
    flagged = 0
    for product in products:
        prev_eligible = product.eligible_for_distribution
        risk_score, risk_reason = compute_risk_score(product.id)
        
        product.eligible_for_distribution = False
        product.risk_score = risk_score
        product.risk_reason = risk_reason
        flagged += 1
    
    try:
        db.session.commit()
        
        log_audit(
            action='slow_stock.flag_bulk',
            status='success',
            message=f'Bulk flagged {flagged} products for slow stock management',
            metadata={
                'count': flagged,
                'skus': [p.sku for p in products]
            }
        )
        
        flash(f'{flagged} SKUs marcados para gestión de stock lento.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al marcar productos: {str(e)}', 'danger')
        log_audit(
            action='slow_stock.flag_bulk',
            status='fail',
            message=str(e)
        )
    
    return_url = request.form.get('return_url')
    return redirect(return_url or request.referrer or url_for('slow_stock'))


@app.route('/dashboard_category/no_movement/flag_all', methods=['POST'])
@login_required
@require_permission('slow_stock:run')
def flag_slow_stock_all_no_movement():
    """
    Bulk flag ALL no-movement SKUs for the given category/window/store filters.
    Processes in batches for performance.
    """
    category = request.form.get('category', '')
    window_days = int(request.form.get('window_days', 30))
    store_name = request.form.get('store', '')
    return_url = request.form.get('return_url')
    
    if not category:
        flash('Debe seleccionar una categoría.', 'warning')
        return redirect(return_url or url_for('dashboard_category'))
    
    today = date.today()
    window_start = today - timedelta(days=window_days)
    
    store_id_filter = None
    if store_name:
        store_obj = Store.query.filter_by(name=store_name).first()
        if store_obj:
            store_id_filter = store_obj.id
    
    # Build the same query as category_no_movement but get all product IDs
    sales_subq = db.session.query(
        SalesWeeklyAgg.product_id,
        func.sum(SalesWeeklyAgg.units).label('total_sales')
    ).filter(
        SalesWeeklyAgg.category == category,
        SalesWeeklyAgg.week_start >= window_start
    )
    if store_id_filter:
        sales_subq = sales_subq.filter(SalesWeeklyAgg.store_id == store_id_filter)
    sales_subq = sales_subq.group_by(SalesWeeklyAgg.product_id).subquery()
    
    cd_stock_subq = db.session.query(
        StockCD.product_id,
        func.sum(StockCD.quantity).label('stock_cd')
    ).join(Product, StockCD.product_id == Product.id).filter(
        Product.category == category
    ).group_by(StockCD.product_id).subquery()
    
    store_stock_subq = db.session.query(
        StockSnapshot.product_id,
        func.sum(StockSnapshot.quantity).label('stock_stores')
    ).join(Product, StockSnapshot.product_id == Product.id).filter(
        Product.category == category
    )
    if store_id_filter:
        store_stock_subq = store_stock_subq.filter(StockSnapshot.store_id == store_id_filter)
    store_stock_subq = store_stock_subq.group_by(StockSnapshot.product_id).subquery()
    
    # Get all product IDs matching no-movement criteria
    product_ids = db.session.query(Product.id).outerjoin(
        sales_subq, Product.id == sales_subq.c.product_id
    ).outerjoin(
        cd_stock_subq, Product.id == cd_stock_subq.c.product_id
    ).outerjoin(
        store_stock_subq, Product.id == store_stock_subq.c.product_id
    ).filter(
        Product.category == category,
        Product.eligible_for_distribution == True,
        func.coalesce(sales_subq.c.total_sales, 0) == 0,
        (func.coalesce(cd_stock_subq.c.stock_cd, 0) + func.coalesce(store_stock_subq.c.stock_stores, 0)) > 0
    ).all()
    
    product_ids = [pid[0] for pid in product_ids]
    
    if not product_ids:
        flash('No se encontraron SKUs sin movimiento para marcar.', 'info')
        return redirect(return_url or url_for('dashboard_category', category=category, window=window_days, store=store_name))
    
    # Process in batches
    flagged = 0
    batch_size = 100
    
    for i in range(0, len(product_ids), batch_size):
        batch = product_ids[i:i + batch_size]
        products = Product.query.filter(Product.id.in_(batch)).all()
        
        for product in products:
            risk_score, risk_reason = compute_risk_score(product.id)
            product.eligible_for_distribution = False
            product.risk_score = risk_score
            product.risk_reason = risk_reason
            flagged += 1
    
    try:
        db.session.commit()
        
        log_audit(
            action='slow_stock.flag_all_no_movement',
            status='success',
            message=f'Bulk flagged {flagged} no-movement products for category {category}',
            metadata={
                'count': flagged,
                'category': category,
                'window_days': window_days,
                'store': store_name
            }
        )
        
        flash(f'{flagged} SKUs sin movimiento marcados para gestión de stock lento.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al marcar productos: {str(e)}', 'danger')
        log_audit(
            action='slow_stock.flag_all_no_movement',
            status='fail',
            message=str(e)
        )
    
    return redirect(return_url or url_for('category_no_movement', category=category, window_days=window_days, store=store_name))


@app.route('/dashboard_category/no_movement/export_selected', methods=['POST'])
@login_required
@require_permission('dashboard:view')
def category_no_movement_export_selected():
    """Export selected no-movement SKUs to Excel."""
    product_ids_raw = request.form.get('product_ids', '')
    category = request.form.get('category', 'export')
    
    product_ids = []
    if product_ids_raw:
        product_ids = [pid.strip() for pid in product_ids_raw.split(',') if pid.strip()]
    
    if not product_ids:
        flash('No se seleccionaron productos para exportar.', 'warning')
        return redirect(request.referrer or url_for('dashboard_category'))
    
    # Fetch products with stock info
    products = Product.query.filter(Product.id.in_([int(pid) for pid in product_ids])).all()
    
    rows = []
    for p in products:
        stock_cd = db.session.query(func.coalesce(func.sum(StockCD.quantity), 0)).filter(
            StockCD.product_id == p.id
        ).scalar() or 0
        
        stock_stores = db.session.query(func.coalesce(func.sum(StockSnapshot.quantity), 0)).filter(
            StockSnapshot.product_id == p.id
        ).scalar() or 0
        
        rows.append({
            'SKU': p.sku,
            'Producto': p.name,
            'Categoría': p.category,
            'Stock CD': stock_cd,
            'Stock Tiendas': stock_stores,
            'Stock Total': stock_cd + stock_stores,
            'Marcado Stock Lento': 'Sí' if not p.eligible_for_distribution else 'No',
            'Riesgo': p.risk_score if p.risk_score else ''
        })
    
    df = pd.DataFrame(rows)
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Seleccionados')
    output.seek(0)
    
    filename = f"sin_movimiento_seleccionados_{category}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@app.route('/slow_stock/unflag', methods=['POST'])
@login_required
@require_permission('slow_stock:run')
def unflag_slow_stock():
    """Unflag a product (re-enable for distribution)."""
    sku = request.form.get('sku', '').strip()
    product_id = request.form.get('product_id')
    
    product = None
    if product_id:
        product = db.session.get(Product, int(product_id))
    elif sku:
        product = Product.query.filter_by(sku=sku).first()
    
    if not product:
        flash('Producto no encontrado.', 'warning')
        return redirect(request.referrer or url_for('slow_stock'))
    
    prev_eligible = product.eligible_for_distribution
    prev_score = product.risk_score
    
    product.eligible_for_distribution = True
    product.risk_score = None
    product.risk_reason = None
    
    try:
        db.session.commit()
        
        log_audit(
            action='slow_stock.unflag',
            status='success',
            message=f'SKU {product.sku} unflagged from slow stock management',
            entity_type='Product',
            entity_id=product.id,
            metadata={
                'sku': product.sku,
                'prev_risk_score': prev_score,
                'prev_eligible_for_distribution': prev_eligible
            }
        )
        
        flash(f'SKU {product.sku} rehabilitado para distribución.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al rehabilitar producto: {str(e)}', 'danger')
    
    return redirect(request.referrer or url_for('slow_stock'))


# ------------------ Store Health Index Module ------------------

STORE_HEALTH_WEIGHTS = {
    'FILL_RATE': 0.30,
    'STOCKOUT_RATE': 0.30,
    'OVERSTOCK_RATE': 0.20,
    'SALES_VELOCITY': 0.20,
    'MAX_WOC': 8.0
}


def compute_store_health_index(weights=None, sku_scope='core'):
    """
    Compute a health score (0-100) for each store based on:
    - Fill rate (% SKUs with stock > 0)
    - Stockout rate (% SKUs with stock = 0)
    - Overstock rate (% SKUs over MAX_WOC)
    - Sales velocity (average weekly sales)
    - Redistribution dependency (SKUs received via transfers)
    
    sku_scope: 'core' (default), 'runs', or 'full'
    - core: SKUs with global sales in last 90 days OR in recent dist runs OR has stock in stores
    - runs: SKUs that appeared in last 5 distribution runs
    - full: All SKUs in catalog
    """
    if weights is None:
        weights = STORE_HEALTH_WEIGHTS.copy()
    
    # Normalize weights to ensure they sum to 1.0
    weight_sum = (weights.get('FILL_RATE', 0.30) + weights.get('STOCKOUT_RATE', 0.30) +
                  weights.get('OVERSTOCK_RATE', 0.20) + weights.get('SALES_VELOCITY', 0.20))
    if weight_sum > 0:
        weights['FILL_RATE'] = weights.get('FILL_RATE', 0.30) / weight_sum
        weights['STOCKOUT_RATE'] = weights.get('STOCKOUT_RATE', 0.30) / weight_sum
        weights['OVERSTOCK_RATE'] = weights.get('OVERSTOCK_RATE', 0.20) / weight_sum
        weights['SALES_VELOCITY'] = weights.get('SALES_VELOCITY', 0.20) / weight_sum
    
    stores = {s.id: s for s in Store.query.all()}
    all_products = {p.id: p for p in Product.query.all()}
    
    # Get flagged products (excluded from distribution) - exclude from health metrics
    flagged_product_ids = {p.id for p in Product.query.filter(Product.eligible_for_distribution == False).all()}
    
    if not stores or not all_products:
        return {'error': 'Sin tiendas o productos registrados', 'stores': []}
    
    # Build SKU scope based on selected mode
    today = date.today()
    scope_window_days = 90
    scope_runs_count = 5
    
    # Get SKUs with global sales in last N days - prefer SalesWeeklyAgg (macro layer) if available
    sales_cutoff = today - timedelta(days=scope_window_days)
    skus_with_sales = set()
    
    macro_count = db.session.query(func.count(SalesWeeklyAgg.id)).scalar() or 0
    if macro_count > 0:
        sales_query = db.session.query(SalesWeeklyAgg.product_id.distinct()).filter(
            SalesWeeklyAgg.week_start >= sales_cutoff
        ).all()
    else:
        sales_query = db.session.query(DistributionRecord.product_id.distinct()).filter(
            DistributionRecord.event_date >= sales_cutoff
        ).all()
    skus_with_sales = {r[0] for r in sales_query}
    
    # Get SKUs that appeared in recent distribution runs
    skus_in_runs = set()
    recent_runs = (
        Run.query
        .filter(Run.run_type == 'distribution')
        .order_by(Run.created_at.desc())
        .limit(scope_runs_count)
        .all()
    )
    if recent_runs:
        run_ids = [r.run_id for r in recent_runs]
        run_skus = db.session.query(Prediction.product_id.distinct()).filter(
            Prediction.run_id.in_(run_ids)
        ).all()
        skus_in_runs = {r[0] for r in run_skus}
    
    # Get SKUs with stock > 0 in any store (exclude CD)
    skus_with_store_stock = set()
    store_ids = list(stores.keys())
    stock_skus = db.session.query(StockSnapshot.product_id.distinct()).filter(
        StockSnapshot.store_id.in_(store_ids),
        StockSnapshot.quantity > 0
    ).all()
    skus_with_store_stock = {r[0] for r in stock_skus}
    
    # Build scope set based on mode
    all_product_ids = set(all_products.keys())
    
    if sku_scope == 'full':
        scope_skus = all_product_ids
    elif sku_scope == 'runs':
        scope_skus = skus_in_runs if skus_in_runs else all_product_ids
    else:  # 'core' - default
        scope_skus = skus_with_sales | skus_in_runs | skus_with_store_stock
        if not scope_skus:
            scope_skus = all_product_ids
    
    # Exclude flagged products from scope (flagged = managed by Slow Stock)
    scope_skus = scope_skus - flagged_product_ids
    
    # Compute exclusion stats
    excluded_skus = all_product_ids - scope_skus
    
    # Categorize excluded SKUs
    cd_only_no_sales = 0
    not_in_stores = 0
    not_in_runs = 0
    
    for pid in excluded_skus:
        has_store_stock = pid in skus_with_store_stock
        has_sales = pid in skus_with_sales
        has_runs = pid in skus_in_runs
        
        if not has_store_stock and not has_sales:
            cd_only_no_sales += 1
        if not has_store_stock:
            not_in_stores += 1
        if not has_runs:
            not_in_runs += 1
    
    scope_info = {
        'mode': sku_scope,
        'mode_label': {'core': 'Surtido Activo', 'runs': 'SKUs en Corridas', 'full': 'Catálogo Completo'}.get(sku_scope, sku_scope),
        'total_catalog': len(all_product_ids),
        'skus_in_scope': len(scope_skus),
        'skus_excluded': len(excluded_skus),
        'excluded_cd_only': cd_only_no_sales,
        'excluded_no_store_stock': not_in_stores,
        'excluded_no_runs': not_in_runs,
        'skus_with_sales': len(skus_with_sales),
        'skus_in_runs': len(skus_in_runs),
        'skus_with_stock': len(skus_with_store_stock)
    }
    
    # Filter products to scope
    products = {pid: all_products[pid] for pid in scope_skus if pid in all_products}
    total_skus = len(products)
    max_woc = weights.get('MAX_WOC', 8.0)
    
    # Get latest stock snapshot for each store-product pair using subquery for latest date
    latest_dates = db.session.query(
        StockSnapshot.store_id,
        StockSnapshot.product_id,
        db.func.max(StockSnapshot.as_of_date).label('max_date')
    ).group_by(
        StockSnapshot.store_id,
        StockSnapshot.product_id
    ).subquery()
    
    stock_query = db.session.query(StockSnapshot).join(
        latest_dates,
        db.and_(
            StockSnapshot.store_id == latest_dates.c.store_id,
            StockSnapshot.product_id == latest_dates.c.product_id,
            StockSnapshot.as_of_date == latest_dates.c.max_date
        )
    ).all()
    
    store_stock = {}
    for ss in stock_query:
        if ss.store_id not in store_stock:
            store_stock[ss.store_id] = {}
        store_stock[ss.store_id][ss.product_id] = ss.quantity
    
    today = date.today()
    recent_window = 4
    cutoff_week = today - timedelta(weeks=recent_window)
    cutoff_week = cutoff_week - timedelta(days=cutoff_week.weekday())
    
    sales_query = db.session.query(
        SalesWeeklyAgg.store_id,
        SalesWeeklyAgg.product_id,
        db.func.sum(SalesWeeklyAgg.units).label('total_units')
    ).filter(
        SalesWeeklyAgg.week_start >= cutoff_week
    ).group_by(
        SalesWeeklyAgg.store_id,
        SalesWeeklyAgg.product_id
    ).all()
    
    sales_rate = {}
    store_total_sales = {}
    for row in sales_query:
        weekly_rate = row.total_units / recent_window if recent_window > 0 else 0
        if row.store_id not in sales_rate:
            sales_rate[row.store_id] = {}
            store_total_sales[row.store_id] = 0
        sales_rate[row.store_id][row.product_id] = weekly_rate
        store_total_sales[row.store_id] += weekly_rate
    
    redistribution_count = {}
    try:
        transfers = SlowStockSuggestion.query.all()
        for t in transfers:
            if t.to_store_id not in redistribution_count:
                redistribution_count[t.to_store_id] = set()
            redistribution_count[t.to_store_id].add(t.product_id)
    except:
        pass
    
    max_sales = max(store_total_sales.values()) if store_total_sales else 0
    has_sales_data = max_sales > 0
    
    store_results = []
    
    for sid, store in stores.items():
        stock_data = store_stock.get(sid, {})
        store_sales = sales_rate.get(sid, {})
        
        skus_with_stock = sum(1 for qty in stock_data.values() if qty > 0)
        skus_in_stockout = sum(1 for pid in products.keys() if stock_data.get(pid, 0) <= 0)
        
        skus_overstocked = 0
        for pid, qty in stock_data.items():
            if qty > 0:
                rate = store_sales.get(pid, 0)
                if rate > 0:
                    woc = qty / rate
                    if woc > max_woc:
                        skus_overstocked += 1
                elif qty > 0:
                    skus_overstocked += 1
        
        fill_rate = (skus_with_stock / total_skus * 100) if total_skus > 0 else 0
        stockout_rate = (skus_in_stockout / total_skus * 100) if total_skus > 0 else 0
        overstock_rate = (skus_overstocked / total_skus * 100) if total_skus > 0 else 0
        
        avg_weekly_sales = store_total_sales.get(sid, 0)
        # Normalize velocity against chain maximum; use 0 if no comparison baseline available
        if max_sales > 0 and avg_weekly_sales > 0:
            sales_velocity_score = min((avg_weekly_sales / max_sales * 100), 100)
        else:
            sales_velocity_score = 0
        
        redist_skus = len(redistribution_count.get(sid, set()))
        
        fill_score = fill_rate
        stockout_score = 100 - stockout_rate
        overstock_score = 100 - overstock_rate
        velocity_score = sales_velocity_score
        
        # If no sales data available chain-wide, redistribute velocity weight to other metrics
        if has_sales_data:
            health_score = (
                fill_score * weights.get('FILL_RATE', 0.30) +
                stockout_score * weights.get('STOCKOUT_RATE', 0.30) +
                overstock_score * weights.get('OVERSTOCK_RATE', 0.20) +
                velocity_score * weights.get('SALES_VELOCITY', 0.20)
            )
        else:
            # Redistribute weights to 3 available metrics when no sales data
            adj_total = weights.get('FILL_RATE', 0.30) + weights.get('STOCKOUT_RATE', 0.30) + weights.get('OVERSTOCK_RATE', 0.20)
            health_score = (
                fill_score * (weights.get('FILL_RATE', 0.30) / adj_total) +
                stockout_score * (weights.get('STOCKOUT_RATE', 0.30) / adj_total) +
                overstock_score * (weights.get('OVERSTOCK_RATE', 0.20) / adj_total)
            ) if adj_total > 0 else 0
        
        health_score = max(0, min(100, round(health_score, 1)))
        
        if health_score >= 80:
            status = 'HEALTHY'
        elif health_score >= 50:
            status = 'WARNING'
        else:
            status = 'CRITICAL'
        
        alerts = []
        if stockout_rate > 30:
            alerts.append(f'{round(stockout_rate)}% quiebres')
        if overstock_rate > 25:
            alerts.append(f'{round(overstock_rate)}% sobrestock')
        if fill_rate < 50:
            alerts.append('Surtido bajo')
        if redist_skus > 5:
            alerts.append(f'{redist_skus} SKUs redistribuidos')
        if avg_weekly_sales == 0:
            alerts.append('Sin ventas recientes')
        
        store_results.append({
            'store_id': sid,
            'store_name': store.name,
            'health_score': health_score,
            'status': status,
            'fill_rate': round(fill_rate, 1),
            'stockout_rate': round(stockout_rate, 1),
            'overstock_rate': round(overstock_rate, 1),
            'avg_weekly_sales': round(avg_weekly_sales, 1),
            'redist_dependency': redist_skus,
            'alerts': ', '.join(alerts) if alerts else 'Sin alertas'
        })
    
    store_results.sort(key=lambda x: x['health_score'], reverse=True)
    
    healthy_count = sum(1 for s in store_results if s['status'] == 'HEALTHY')
    warning_count = sum(1 for s in store_results if s['status'] == 'WARNING')
    critical_count = sum(1 for s in store_results if s['status'] == 'CRITICAL')
    avg_score = round(sum(s['health_score'] for s in store_results) / len(store_results), 1) if store_results else 0
    
    return {
        'stores': store_results,
        'summary': {
            'total_stores': len(store_results),
            'healthy_count': healthy_count,
            'warning_count': warning_count,
            'critical_count': critical_count,
            'avg_score': avg_score,
            'has_sales_data': has_sales_data
        },
        'weights': weights,
        'scope_info': scope_info
    }


@app.route('/store_health')
@login_required
@require_permission('store_health:view')
def store_health():
    """Store Health Index diagnostic module."""
    weights = STORE_HEALTH_WEIGHTS.copy()
    
    w_fill = request.args.get('w_fill', type=float)
    w_stockout = request.args.get('w_stockout', type=float)
    w_overstock = request.args.get('w_overstock', type=float)
    w_velocity = request.args.get('w_velocity', type=float)
    
    if w_fill is not None:
        weights['FILL_RATE'] = w_fill / 100
    if w_stockout is not None:
        weights['STOCKOUT_RATE'] = w_stockout / 100
    if w_overstock is not None:
        weights['OVERSTOCK_RATE'] = w_overstock / 100
    if w_velocity is not None:
        weights['SALES_VELOCITY'] = w_velocity / 100
    
    sku_scope = request.args.get('scope', 'core')
    if sku_scope not in ('core', 'runs', 'full'):
        sku_scope = 'core'
    
    results = compute_store_health_index(weights, sku_scope=sku_scope)
    
    sort_by = request.args.get('sort', 'score')
    sort_order = request.args.get('order', 'desc')
    
    stores_data = results.get('stores', [])
    
    if sort_by == 'name':
        stores_data.sort(key=lambda x: x['store_name'], reverse=(sort_order == 'desc'))
    elif sort_by == 'fill':
        stores_data.sort(key=lambda x: x['fill_rate'], reverse=(sort_order == 'desc'))
    elif sort_by == 'stockout':
        stores_data.sort(key=lambda x: x['stockout_rate'], reverse=(sort_order == 'desc'))
    else:
        stores_data.sort(key=lambda x: x['health_score'], reverse=(sort_order == 'desc'))
    
    page = request.args.get('page', 1, type=int)
    pag = paginate_list(stores_data, page, 10)
    
    return render_template('store_health.html',
                           results=results,
                           pag=pag,
                           page=page,
                           sort_by=sort_by,
                           sort_order=sort_order,
                           weights=weights,
                           default_weights=STORE_HEALTH_WEIGHTS,
                           sku_scope=sku_scope)


@app.route('/export_store_health')
@login_required
@require_permission('store_health:view')
def export_store_health():
    """Export store health index results."""
    sku_scope = request.args.get('scope', 'core')
    if sku_scope not in ('core', 'runs', 'full'):
        sku_scope = 'core'
    
    results = compute_store_health_index(sku_scope=sku_scope)
    
    if 'error' in results:
        flash(results['error'], 'warning')
        return redirect(url_for('store_health'))
    
    export_data = []
    for row in results['stores']:
        export_data.append({
            'Tienda': row['store_name'],
            'Puntaje Salud': row['health_score'],
            'Estado': row['status'],
            'Fill Rate (%)': row['fill_rate'],
            'Quiebres (%)': row['stockout_rate'],
            'Sobrestock (%)': row['overstock_rate'],
            'Venta Sem': row['avg_weekly_sales'],
            'SKUs Redistribuidos': row['redist_dependency'],
            'Alertas': row['alerts']
        })
    
    df = pd.DataFrame(export_data)
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Índice Salud Tiendas')
    output.seek(0)
    
    log_audit(
        action='store_health.export',
        status='success',
        message=f'Exported store health index: {len(export_data)} stores',
        metadata={'stores': len(export_data)}
    )
    
    return send_file(
        output,
        as_attachment=True,
        download_name=f'indice_salud_tiendas_{date.today().isoformat()}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ------------------ Explainability Module ------------------

EXPLAINABILITY_PARAMS = {
    'MIN_WOC': 1.5,
    'MAX_WOC': 6.0
}


def explain_distribution(product_id, store_id, prediction_qty, run_id=None):
    """
    Generate human-readable explanation for a distribution suggestion.
    Uses the same calculation logic as generate_predictions (SMA with optional stock adjustment).
    Returns a list of bullet points explaining the suggestion.
    """
    try:
        bullets = []
        
        product = Product.query.get(product_id)
        store = Store.query.get(store_id)
        
        if not product or not store:
            return ["Error: Producto o tienda no encontrados"]
        
        run_mode = "sma3_min3"
        run_meta = None
        use_stock = True
        if run_id:
            run_obj = Run.query.filter_by(run_id=run_id).first()
            if run_obj:
                run_mode = run_obj.mode or "sma3_min3"
                run_meta = f"{run_obj.folio or ''} {run_obj.responsable or ''} {run_obj.categoria or ''}".strip()
        
        if "ignore_stock" in run_mode:
            use_stock = False
        
        if run_mode.startswith("sma3"):
            win = 3
            if use_stock:
                method_label = "Promedio móvil 3 semanas (SMA3) ajustado por stock"
            else:
                method_label = "Promedio móvil 3 semanas (SMA3) sin ajuste de stock"
        elif run_mode.startswith("sma2"):
            win = 2
            if use_stock:
                method_label = "Promedio móvil 2 semanas (SMA2) ajustado por stock"
            else:
                method_label = "Promedio móvil 2 semanas (SMA2) sin ajuste de stock"
        else:
            win = 1
            if use_stock:
                method_label = "Última semana ajustado por stock"
            else:
                method_label = "Última semana sin ajuste de stock"
        
        # Primary window sales query
        sales_q = (
            db.session.query(
                db.func.strftime('%Y-%W', DistributionRecord.event_date).label('week'),
                db.func.sum(DistributionRecord.quantity).label('qty')
            )
            .filter(
                DistributionRecord.product_id == product_id,
                DistributionRecord.store_id == store_id
            )
            .group_by(db.func.strftime('%Y-%W', DistributionRecord.event_date))
            .order_by(db.func.strftime('%Y-%W', DistributionRecord.event_date).desc())
            .limit(win)
            .all()
        )
        
        weekly_totals = [int(r.qty) for r in sales_q]
        total_weeks = len(weekly_totals)
        total_sales = sum(weekly_totals) if weekly_totals else 0
        primary_sma_value = round(total_sales / max(total_weeks, 1), 1)
        
        store_stock = 0
        stock_date = "N/A"
        if use_stock:
            latest_stock = (
                StockSnapshot.query
                .filter(
                    StockSnapshot.product_id == product_id,
                    StockSnapshot.store_id == store_id
                )
                .order_by(StockSnapshot.as_of_date.desc())
                .first()
            )
            store_stock = latest_stock.quantity if latest_stock else 0
            stock_date = latest_stock.as_of_date if latest_stock else "N/A"
        
        # Stockout-aware backoff logic
        FALLBACK_WINDOW_WEEKS = 12
        backoff_applied = False
        fallback_sma_value = 0.0
        sma_value = primary_sma_value
        
        if total_sales == 0 and store_stock == 0:
            # Query fallback window (12 weeks)
            fallback_q = (
                db.session.query(
                    db.func.strftime('%Y-%W', DistributionRecord.event_date).label('week'),
                    db.func.sum(DistributionRecord.quantity).label('qty')
                )
                .filter(
                    DistributionRecord.product_id == product_id,
                    DistributionRecord.store_id == store_id
                )
                .group_by(db.func.strftime('%Y-%W', DistributionRecord.event_date))
                .order_by(db.func.strftime('%Y-%W', DistributionRecord.event_date).desc())
                .limit(FALLBACK_WINDOW_WEEKS)
                .all()
            )
            fallback_totals = [int(r.qty) for r in fallback_q]
            fallback_sum = sum(fallback_totals) if fallback_totals else 0
            if fallback_sum > 0:
                fallback_sma_value = round(fallback_sum / min(len(fallback_totals), FALLBACK_WINDOW_WEEKS), 1)
                sma_value = fallback_sma_value
                backoff_applied = True
        
        latest_cd_date = db.session.query(db.func.max(StockCD.as_of_date)).scalar()
        cd_stock = 0
        if latest_cd_date:
            cd_row = StockCD.query.filter(
                StockCD.product_id == product_id,
                StockCD.as_of_date == latest_cd_date
            ).first()
            cd_stock = cd_row.quantity if cd_row else 0
        
        if use_stock:
            raw_suggested = max(int(round(sma_value)) - store_stock, 0)
        else:
            raw_suggested = int(round(sma_value))
        
        bullets.append(f"Método: {method_label}")
        if run_meta:
            bullets.append(f"Run: {run_meta}")
        
        if backoff_applied:
            bullets.append(f"BACKOFF STOCKOUT: Tienda sin stock y sin ventas en ventana primaria ({win} semanas)")
            bullets.append(f"Se usa promedio de 12 semanas en lugar de SMA{win}")
            bullets.append(f"Ventas últimas 12 semanas: {int(fallback_sum)} unidades")
            bullets.append(f"Promedio semanal (fallback 12 sem): {sma_value} unidades")
        elif total_weeks > 0:
            weeks_str = ", ".join([str(w) for w in weekly_totals])
            bullets.append(f"Ventas últimas {total_weeks} semanas: [{weeks_str}]")
            bullets.append(f"Promedio semanal (SMA{win}): {sma_value} unidades")
        else:
            bullets.append("Sin historial de ventas reciente")
        
        if use_stock:
            bullets.append(f"Stock tienda actual: {store_stock} unidades (fecha: {stock_date})")
            if backoff_applied:
                bullets.append(f"Cálculo: Fallback ({sma_value}) - Stock tienda ({store_stock}) = {raw_suggested}")
            else:
                bullets.append(f"Cálculo: SMA{win} ({sma_value}) - Stock tienda ({store_stock}) = {raw_suggested}")
        else:
            bullets.append("Ajuste de stock: desactivado para este run")
            bullets.append(f"Cálculo: SMA{win} ({sma_value}) = {raw_suggested}")
        
        bullets.append(f"Stock CD disponible: {cd_stock} unidades")
        
        if cd_stock > 0:
            if prediction_qty < raw_suggested:
                bullets.append(f"Limitado por disponibilidad CD: {prediction_qty} de {raw_suggested} unidades asignadas")
            elif prediction_qty == 0 and raw_suggested > 0:
                bullets.append("Sin stock disponible en CD para cubrir sugerencia")
        elif cd_stock == 0 and raw_suggested > 0:
            bullets.append("Sin stock en CD - no se puede enviar mercancía")
        
        bullets.append(f"Cantidad final sugerida: {prediction_qty} unidades")
        
        return bullets
    except Exception as e:
        return [f"Error generando explicación: {str(e)}"]


def explain_forecast(sku, horizon_weeks=8, lead_time_weeks=4, safety_pct=0.1, store_filter=None):
    """
    Generate human-readable explanation for a forecast purchase suggestion.
    Uses the same calculation logic as Forecast V2 page (avg_last4 weeks).
    Returns a list of bullet points explaining the suggestion.
    """
    try:
        today = date.today()
        bullets = []
        
        product = Product.query.filter_by(sku=sku).first()
        if not product:
            return ["Error: SKU no encontrado"]
        
        week_expr = db.func.strftime('%Y-%W', DistributionRecord.event_date)
        
        weekly_q = (
            db.session.query(
                week_expr.label('week'),
                db.func.sum(DistributionRecord.quantity).label('qty')
            )
            .filter(DistributionRecord.product_id == product.id)
        )
        
        store_name = store_filter
        if store_filter:
            store = Store.query.filter_by(name=store_filter).first()
            if store:
                weekly_q = weekly_q.filter(DistributionRecord.store_id == store.id)
        
        weekly_q = (
            weekly_q
            .group_by(week_expr)
            .order_by(week_expr.desc())
            .limit(4)
            .all()
        )
        
        weekly_values = [int(r.qty) for r in weekly_q]
        num_weeks = len(weekly_values)
        total_last4 = sum(weekly_values) if weekly_values else 0
        avg_last4 = round(total_last4 / max(num_weeks, 1), 1)
        
        latest_cd_date = db.session.query(db.func.max(StockCD.as_of_date)).scalar() or today
        cd_stock = db.session.query(
            db.func.coalesce(db.func.sum(StockCD.quantity), 0)
        ).filter(
            StockCD.product_id == product.id,
            StockCD.as_of_date == latest_cd_date
        ).scalar() or 0
        
        forecast_total = avg_last4 * horizon_weeks
        demand_lead_time = avg_last4 * lead_time_weeks
        safety_units = int(round(avg_last4 * lead_time_weeks * safety_pct))
        
        total_required = forecast_total + demand_lead_time + safety_units
        suggested_purchase = max(0, int(round(total_required - cd_stock)))
        
        woc_cd = round(cd_stock / max(avg_last4, 0.01), 1) if avg_last4 > 0 else 0
        
        bullets.append(f"SKU: {sku}")
        if store_name:
            bullets.append(f"Tienda: {store_name}")
        else:
            bullets.append("Todas las tiendas")
        
        if num_weeks > 0:
            weeks_str = ", ".join([str(w) for w in weekly_values])
            bullets.append(f"Ventas últimas {num_weeks} semanas: [{weeks_str}]")
        else:
            bullets.append("Sin historial de ventas reciente")
        
        bullets.append(f"Promedio semanal (últimas 4 sem): {avg_last4} unidades")
        bullets.append(f"Horizonte de cobertura: {horizon_weeks} semanas")
        bullets.append(f"Forecast horizonte: {avg_last4} × {horizon_weeks} = {int(forecast_total)} unidades")
        
        bullets.append(f"Lead time: {lead_time_weeks} semanas")
        bullets.append(f"Demanda lead time: {avg_last4} × {lead_time_weeks} = {int(demand_lead_time)} unidades")
        
        bullets.append(f"Stock de seguridad: {int(safety_pct * 100)}%")
        bullets.append(f"Unidades de seguridad: {safety_units} unidades")
        
        bullets.append(f"Stock CD actual: {cd_stock} unidades")
        bullets.append(f"Cobertura CD: {woc_cd} semanas")
        
        bullets.append(f"Requerimiento total: {int(total_required)} unidades")
        bullets.append(f"Cálculo compra: {int(total_required)} - {cd_stock} = {suggested_purchase}")
        bullets.append(f"Compra sugerida: {suggested_purchase} unidades")
        
        return bullets
    except Exception as e:
        return [f"Error generando explicación: {str(e)}"]


@app.route('/api/explain/distribution')
@login_required
def api_explain_distribution():
    """API endpoint for distribution explainability."""
    product_id = request.args.get('product_id', type=int)
    store_id = request.args.get('store_id', type=int)
    qty = request.args.get('qty', 0, type=int)
    run_id = request.args.get('run_id', '')
    
    if not product_id or not store_id:
        return jsonify({'error': 'Se requiere product_id y store_id', 'bullets': []}), 400
    
    product = Product.query.get(product_id)
    store = Store.query.get(store_id)
    
    if not product:
        return jsonify({'error': 'Producto no encontrado', 'bullets': []}), 404
    if not store:
        return jsonify({'error': 'Tienda no encontrada', 'bullets': []}), 404
    
    bullets = explain_distribution(product_id, store_id, qty, run_id)
    
    return jsonify({
        'sku': product.sku,
        'product_name': product.name,
        'store_name': store.name,
        'quantity': qty,
        'bullets': bullets
    })


@app.route('/api/explain/forecast')
@login_required
def api_explain_forecast():
    """API endpoint for forecast explainability."""
    sku = request.args.get('sku', '').strip()
    horizon_weeks = request.args.get('horizon_weeks', 8, type=int)
    lead_time_weeks = request.args.get('lead_time_weeks', 4, type=float)
    safety_pct = request.args.get('safety_pct', 0.1, type=float)
    store_filter = request.args.get('store', '').strip() or None
    
    if not sku:
        return jsonify({'error': 'Se requiere SKU', 'bullets': []}), 400
    
    product = Product.query.filter_by(sku=sku).first()
    if not product:
        return jsonify({'error': 'SKU no encontrado', 'bullets': []}), 404
    
    bullets = explain_forecast(sku, horizon_weeks, lead_time_weeks, safety_pct, store_filter)
    
    return jsonify({
        'sku': sku,
        'product_name': product.name,
        'store': store_filter or 'Todas',
        'bullets': bullets
    })


# ------------------ Alerts Module ------------------

ALERT_PARAMS = {
    'MIN_WOC': 1.5,
    'MAX_WOC': 6.0,
    'DEAD_DAYS_STORE': 60,
    'DEAD_DAYS_GLOBAL': 90,
    'RECENT_WINDOW_WEEKS': 4
}

_alerts_cache = {'data': None, 'time': 0}
ALERTS_CACHE_TTL = 45


def get_alerts_summary(store_filter=None):
    """
    Lightweight alerts summary with caching.
    Returns {high_count, medium_count, low_count, total_count}.
    
    Strategy: Cache the full global alerts list once (45s TTL), then
    derive store-filtered summaries from the cached list instantly.
    This means only 1 expensive compute_alerts() call per cache period,
    regardless of how many store-specific dashboard views are loaded.
    """
    import time
    
    now = time.time()
    
    if _alerts_cache['data'] is None or (now - _alerts_cache['time']) >= ALERTS_CACHE_TTL:
        try:
            _alerts_cache['data'] = compute_alerts()
            _alerts_cache['time'] = now
        except Exception:
            _alerts_cache['data'] = []
            _alerts_cache['time'] = now
    
    all_alerts = _alerts_cache['data']
    
    if store_filter:
        all_alerts = [a for a in all_alerts 
                     if a.get('location') == store_filter and a.get('location_type') == 'store']
    
    high_count = sum(1 for a in all_alerts if a['severity'] == 'HIGH')
    medium_count = sum(1 for a in all_alerts if a['severity'] == 'MEDIUM')
    low_count = sum(1 for a in all_alerts if a['severity'] == 'LOW')
    total_count = high_count + medium_count + low_count
    
    return {
        'high_count': high_count,
        'medium_count': medium_count,
        'low_count': low_count,
        'total_count': total_count
    }


def classify_sku_lifecycle(sales_30d_map, sales_90d_map, sales_ever_map, product_ids):
    """
    Classify SKUs by lifecycle status based on SalesWeeklyAgg data.
    
    Returns dict {product_id: lifecycle_status} where status is:
    - ACTIVE: sales in last 30 days
    - SLOW: no sales last 30d, but sales in 31-90 days
    - DEAD: no sales last 90d, but has historical sales
    - NEW: no sales ever (cold start candidates)
    """
    lifecycle = {}
    for pid in product_ids:
        sales_30 = sales_30d_map.get(pid, 0)
        sales_90 = sales_90d_map.get(pid, 0)
        sales_ever = sales_ever_map.get(pid, 0)
        
        if sales_30 > 0:
            lifecycle[pid] = 'ACTIVE'
        elif sales_90 > 0:
            lifecycle[pid] = 'SLOW'
        elif sales_ever > 0:
            lifecycle[pid] = 'DEAD'
        else:
            lifecycle[pid] = 'NEW'
    
    return lifecycle


def get_lifecycle_sales_maps(store_id_filter=None):
    """
    Build lifecycle classification maps from SalesWeeklyAgg.
    Returns (sales_30d_map, sales_90d_map, sales_ever_map) dicts of {product_id: units}.
    """
    today = date.today()
    cutoff_30d = today - timedelta(days=30)
    cutoff_90d = today - timedelta(days=90)
    
    base_query = db.session.query(
        SalesWeeklyAgg.product_id,
        db.func.sum(SalesWeeklyAgg.units).label('total_units')
    )
    if store_id_filter:
        base_query = base_query.filter(SalesWeeklyAgg.store_id == store_id_filter)
    
    sales_30d = base_query.filter(SalesWeeklyAgg.week_start >= cutoff_30d).group_by(SalesWeeklyAgg.product_id).all()
    sales_30d_map = {r.product_id: int(r.total_units) for r in sales_30d}
    
    base_query_90 = db.session.query(
        SalesWeeklyAgg.product_id,
        db.func.sum(SalesWeeklyAgg.units).label('total_units')
    )
    if store_id_filter:
        base_query_90 = base_query_90.filter(SalesWeeklyAgg.store_id == store_id_filter)
    
    sales_90d = base_query_90.filter(
        SalesWeeklyAgg.week_start >= cutoff_90d,
        SalesWeeklyAgg.week_start < cutoff_30d
    ).group_by(SalesWeeklyAgg.product_id).all()
    sales_90d_map = {r.product_id: int(r.total_units) for r in sales_90d}
    
    base_query_ever = db.session.query(
        SalesWeeklyAgg.product_id,
        db.func.sum(SalesWeeklyAgg.units).label('total_units')
    )
    if store_id_filter:
        base_query_ever = base_query_ever.filter(SalesWeeklyAgg.store_id == store_id_filter)
    
    sales_ever = base_query_ever.group_by(SalesWeeklyAgg.product_id).all()
    sales_ever_map = {r.product_id: int(r.total_units) for r in sales_ever}
    
    return sales_30d_map, sales_90d_map, sales_ever_map


def get_category_avg_demand(category, store_id_filter=None, weeks=4):
    """
    Get average weekly demand for a category (for cold start NEW SKUs).
    Returns average units per week per SKU in category.
    """
    today = date.today()
    cutoff = today - timedelta(days=weeks * 7)
    
    query = db.session.query(
        db.func.sum(SalesWeeklyAgg.units).label('total'),
        db.func.count(db.func.distinct(SalesWeeklyAgg.product_id)).label('sku_count')
    ).filter(
        SalesWeeklyAgg.category == category,
        SalesWeeklyAgg.week_start >= cutoff
    )
    if store_id_filter:
        query = query.filter(SalesWeeklyAgg.store_id == store_id_filter)
    
    result = query.first()
    if result and result.total and result.sku_count > 0:
        return float(result.total) / float(result.sku_count) / float(weeks)
    return 0.0


def compute_alerts(params=None):
    """
    Compute dynamic alerts using Macro Sales (SalesWeeklyAgg) and lifecycle tables.
    Returns list of alert dicts sorted by severity (HIGH > MEDIUM > LOW).
    
    Alert types:
    - PROJECTED_STOCKOUT: Store has stock but low WOC
    - OVERSTOCK: Stock exceeds WOC threshold
    - SILENT_SKU: No sales for extended period (uses lifecycle)
    - BROKEN_STOCK: Store has zero stock but recent demand (uses lifecycle)
    """
    if params is None:
        params = ALERT_PARAMS.copy()
    
    today = date.today()
    alerts = []
    severity_order = {'HIGH': 0, 'MEDIUM': 1, 'LOW': 2}
    
    latest_store_date = db.session.query(db.func.max(StockSnapshot.as_of_date)).scalar()
    latest_cd_date = db.session.query(db.func.max(StockCD.as_of_date)).scalar()
    
    if not latest_store_date:
        return alerts
    
    products = {p.id: p for p in Product.query.all()}
    stores = {s.id: s for s in Store.query.all()}
    
    store_stock = {}
    for ss in StockSnapshot.query.filter(StockSnapshot.as_of_date == latest_store_date).all():
        store_stock[(ss.product_id, ss.store_id)] = ss.quantity
    
    cd_stock = {}
    if latest_cd_date:
        for cd in StockCD.query.filter(StockCD.as_of_date == latest_cd_date).all():
            cd_stock[cd.product_id] = cd.quantity
    
    recent_window = params.get('RECENT_WINDOW_WEEKS', 4)
    cutoff_week = today - timedelta(weeks=recent_window)
    cutoff_week = cutoff_week - timedelta(days=cutoff_week.weekday())
    
    sales_query = db.session.query(
        SalesWeeklyAgg.product_id,
        SalesWeeklyAgg.store_id,
        db.func.sum(SalesWeeklyAgg.units).label('total_units')
    ).filter(
        SalesWeeklyAgg.week_start >= cutoff_week
    ).group_by(
        SalesWeeklyAgg.product_id,
        SalesWeeklyAgg.store_id
    ).all()
    
    sales_rate = {}
    for row in sales_query:
        weekly_rate = row.total_units / recent_window if recent_window > 0 else 0
        sales_rate[(row.product_id, row.store_id)] = weekly_rate
    
    last_sale_store = {}
    for lc in SkuStoreLifecycle.query.all():
        if lc.last_sale_date_store:
            last_sale_store[(lc.product_id, lc.store_id)] = lc.last_sale_date_store
    
    last_sale_global = {}
    for lc in SkuLifecycle.query.all():
        if lc.last_sale_date_global:
            last_sale_global[lc.product_id] = lc.last_sale_date_global
    
    min_woc = params.get('MIN_WOC', 1.5)
    max_woc = params.get('MAX_WOC', 6.0)
    dead_days_store = params.get('DEAD_DAYS_STORE', 60)
    dead_days_global = params.get('DEAD_DAYS_GLOBAL', 90)
    broken_stock_days = params.get('BROKEN_STOCK_DAYS', 45)
    
    for (pid, sid), qty in store_stock.items():
        if qty <= 0:
            continue
        
        product = products.get(pid)
        store = stores.get(sid)
        if not product or not store:
            continue
        
        rate = sales_rate.get((pid, sid), 0)
        woc = qty / rate if rate > 0 else 9999
        last_sale = last_sale_store.get((pid, sid))
        days_since = (today - last_sale).days if last_sale else 9999
        
        if rate > 0:
            if woc < 0.5:
                alerts.append({
                    'type': 'PROJECTED_STOCKOUT',
                    'severity': 'HIGH',
                    'sku': product.sku,
                    'product_name': product.name,
                    'category': product.category or '',
                    'location': store.name,
                    'location_type': 'store',
                    'woc': round(woc, 1),
                    'stock': qty,
                    'rate': round(rate, 2),
                    'reason': f'WOC {round(woc, 1)} sem < 0.5',
                    'action': 'Reponer urgente'
                })
            elif woc < min_woc:
                alerts.append({
                    'type': 'PROJECTED_STOCKOUT',
                    'severity': 'MEDIUM',
                    'sku': product.sku,
                    'product_name': product.name,
                    'category': product.category or '',
                    'location': store.name,
                    'location_type': 'store',
                    'woc': round(woc, 1),
                    'stock': qty,
                    'rate': round(rate, 2),
                    'reason': f'WOC {round(woc, 1)} sem < {min_woc}',
                    'action': 'Reponer'
                })
        
        if rate > 0:
            if woc > max_woc * 2:
                alerts.append({
                    'type': 'OVERSTOCK',
                    'severity': 'HIGH',
                    'sku': product.sku,
                    'product_name': product.name,
                    'category': product.category or '',
                    'location': store.name,
                    'location_type': 'store',
                    'woc': round(woc, 1) if woc < 9999 else None,
                    'stock': qty,
                    'rate': round(rate, 2),
                    'reason': f'WOC {round(woc, 1)} sem > {max_woc * 2}',
                    'action': 'Redistribuir'
                })
            elif woc > max_woc:
                alerts.append({
                    'type': 'OVERSTOCK',
                    'severity': 'MEDIUM',
                    'sku': product.sku,
                    'product_name': product.name,
                    'category': product.category or '',
                    'location': store.name,
                    'location_type': 'store',
                    'woc': round(woc, 1) if woc < 9999 else None,
                    'stock': qty,
                    'rate': round(rate, 2),
                    'reason': f'WOC {round(woc, 1)} sem > {max_woc}',
                    'action': 'Revisar inventario'
                })
        
        if days_since >= dead_days_store:
            if days_since >= dead_days_store * 2:
                severity = 'HIGH'
            elif days_since >= dead_days_store * 1.5:
                severity = 'MEDIUM'
            else:
                severity = 'LOW'
            alerts.append({
                'type': 'SILENT_SKU',
                'severity': severity,
                'sku': product.sku,
                'product_name': product.name,
                'category': product.category or '',
                'location': store.name,
                'location_type': 'store',
                'woc': round(woc, 1) if woc < 9999 else None,
                'stock': qty,
                'rate': round(rate, 2),
                'days_since_sale': days_since if days_since < 9999 else None,
                'reason': f'{days_since} días sin venta' if days_since < 9999 else 'Sin ventas registradas',
                'action': 'Liquidar' if severity == 'HIGH' else 'Revisar'
            })
    
    for (pid, sid), last_date in last_sale_store.items():
        stock_qty = store_stock.get((pid, sid), 0)
        if stock_qty > 0:
            continue
        
        product = products.get(pid)
        store = stores.get(sid)
        if not product or not store:
            continue
        
        days_since = (today - last_date).days if last_date else 9999
        if days_since <= broken_stock_days:
            rate = sales_rate.get((pid, sid), 0)
            if days_since <= 14:
                severity = 'HIGH'
            elif days_since <= 30:
                severity = 'MEDIUM'
            else:
                severity = 'LOW'
            alerts.append({
                'type': 'BROKEN_STOCK',
                'severity': severity,
                'sku': product.sku,
                'product_name': product.name,
                'category': product.category or '',
                'location': store.name,
                'location_type': 'store',
                'woc': None,
                'stock': 0,
                'rate': round(rate, 2) if rate else None,
                'days_since_sale': days_since,
                'reason': f'Quiebre: {days_since} días desde última venta',
                'action': 'Reponer'
            })
    
    for pid, qty in cd_stock.items():
        if qty <= 0:
            continue
        
        product = products.get(pid)
        if not product:
            continue
        
        global_rate = sum(sales_rate.get((pid, sid), 0) for sid in stores.keys())
        woc_cd = qty / global_rate if global_rate > 0 else 9999
        last_sale = last_sale_global.get(pid)
        days_since = (today - last_sale).days if last_sale else 9999
        
        if global_rate > 0 and woc_cd < 9999:
            if woc_cd > max_woc * 2:
                alerts.append({
                    'type': 'OVERSTOCK',
                    'severity': 'HIGH',
                    'sku': product.sku,
                    'product_name': product.name,
                    'category': product.category or '',
                    'location': 'Centro Distribución',
                    'location_type': 'cd',
                    'woc': round(woc_cd, 1),
                    'stock': qty,
                    'rate': round(global_rate, 2),
                    'reason': f'WOC CD {round(woc_cd, 1)} sem > {max_woc * 2}',
                    'action': 'Revisar compra'
                })
            elif woc_cd > max_woc:
                alerts.append({
                    'type': 'OVERSTOCK',
                    'severity': 'MEDIUM',
                    'sku': product.sku,
                    'product_name': product.name,
                    'category': product.category or '',
                    'location': 'Centro Distribución',
                    'location_type': 'cd',
                    'woc': round(woc_cd, 1),
                    'stock': qty,
                    'rate': round(global_rate, 2),
                    'reason': f'WOC CD {round(woc_cd, 1)} sem > {max_woc}',
                    'action': 'Revisar compra'
                })
        
        if days_since >= dead_days_global:
            if days_since >= dead_days_global * 2:
                severity = 'HIGH'
            elif days_since >= dead_days_global * 1.5:
                severity = 'MEDIUM'
            else:
                severity = 'LOW'
            alerts.append({
                'type': 'SILENT_SKU',
                'severity': severity,
                'sku': product.sku,
                'product_name': product.name,
                'category': product.category or '',
                'location': 'Centro Distribución',
                'location_type': 'cd',
                'woc': round(woc_cd, 1) if woc_cd < 9999 else None,
                'stock': qty,
                'rate': round(global_rate, 2),
                'days_since_sale': days_since if days_since < 9999 else None,
                'reason': f'{days_since} días sin venta global' if days_since < 9999 else 'Sin ventas registradas',
                'action': 'Liquidar' if severity == 'HIGH' else 'Revisar'
            })
    
    alerts.sort(key=lambda x: (severity_order.get(x['severity'], 99), x['sku']))
    
    return alerts


@app.route('/alerts')
@login_required
@require_permission('alerts:view')
def alerts_page():
    """Alerts module - dynamic alerts based on stock and sales velocity.
    Supports multi-value filters via comma-separated query params:
      severity=HIGH,MEDIUM
      type=PROJECTED_STOCKOUT,OVERSTOCK
    """
    severity_param = request.args.get('severity', '').strip()
    type_param = request.args.get('type', '').strip()
    store_filter = request.args.get('store', '').strip()
    sku_filter = request.args.get('sku', '').strip().lower()
    category_filter = request.args.get('category', '').strip()
    
    selected_severities = [s.strip().upper() for s in severity_param.split(',') if s.strip()]
    selected_types = [t.strip().upper() for t in type_param.split(',') if t.strip()]
    
    page = request.args.get('page', 1, type=int)
    per_page = 10
    
    stores = Store.query.order_by(Store.name).all()
    categories = db.session.query(Product.category).filter(
        Product.category.isnot(None)
    ).distinct().order_by(Product.category).all()
    categories = [c[0] for c in categories if c[0]]
    
    all_alerts_unfiltered = compute_alerts()
    
    count_by_severity = {'HIGH': 0, 'MEDIUM': 0, 'LOW': 0}
    count_by_type = {'PROJECTED_STOCKOUT': 0, 'OVERSTOCK': 0, 'SILENT_SKU': 0, 'BROKEN_STOCK': 0}
    for a in all_alerts_unfiltered:
        sev = a.get('severity', '')
        typ = a.get('type', '')
        if sev in count_by_severity:
            count_by_severity[sev] += 1
        if typ in count_by_type:
            count_by_type[typ] += 1
    
    all_alerts = all_alerts_unfiltered
    if selected_severities:
        all_alerts = [a for a in all_alerts if a['severity'] in selected_severities]
    if selected_types:
        all_alerts = [a for a in all_alerts if a['type'] in selected_types]
    if store_filter:
        all_alerts = [a for a in all_alerts if a['location'] == store_filter]
    if sku_filter:
        all_alerts = [a for a in all_alerts if sku_filter in a['sku'].lower() or sku_filter in a['product_name'].lower()]
    if category_filter:
        all_alerts = [a for a in all_alerts if a.get('category', '') == category_filter]
    
    total = len(all_alerts)
    start = (page - 1) * per_page
    end = start + per_page
    alerts = all_alerts[start:end]
    
    total_pages = (total + per_page - 1) // per_page if total > 0 else 1
    
    kpi_high = sum(1 for a in all_alerts if a['severity'] == 'HIGH')
    kpi_medium = sum(1 for a in all_alerts if a['severity'] == 'MEDIUM')
    kpi_low = sum(1 for a in all_alerts if a['severity'] == 'LOW')
    kpi_stockout = sum(1 for a in all_alerts if a['type'] == 'PROJECTED_STOCKOUT')
    kpi_overstock = sum(1 for a in all_alerts if a['type'] == 'OVERSTOCK')
    kpi_silent = sum(1 for a in all_alerts if a['type'] == 'SILENT_SKU')
    kpi_broken = sum(1 for a in all_alerts if a['type'] == 'BROKEN_STOCK')
    
    return render_template(
        'alerts.html',
        alerts=alerts,
        stores=stores,
        categories=categories,
        selected_severities=selected_severities,
        selected_types=selected_types,
        store_filter=store_filter,
        sku_filter=sku_filter,
        category_filter=category_filter,
        page=page,
        total_pages=total_pages,
        total=total,
        kpi_high=kpi_high,
        kpi_medium=kpi_medium,
        kpi_low=kpi_low,
        kpi_stockout=kpi_stockout,
        kpi_overstock=kpi_overstock,
        kpi_silent=kpi_silent,
        kpi_broken=kpi_broken,
        count_by_severity=count_by_severity,
        count_by_type=count_by_type
    )


# ------------------ FastPlanner Routes ------------------
@app.route('/planner')
@login_required
@require_permission('planner:view')
def planner():
    """FastPlanner Kanban board for warehouse execution."""
    status_filter = request.args.get('status', '')
    urgency_filter = request.args.get('urgency', '')
    folio_search = request.args.get('folio', '').strip()

    plans_by_status = {}
    for status in ['APPROVED', 'IN_PROGRESS', 'PACKED', 'DISPATCHED', 'CLOSED']:
        query = DistributionPlan.query.filter_by(status=status)
        if urgency_filter:
            query = query.filter_by(urgency=urgency_filter)
        if folio_search:
            query = query.filter(DistributionPlan.folio.ilike(f'%{folio_search}%'))
        plans_by_status[status] = query.order_by(DistributionPlan.created_at.desc()).all()

    can_operate = current_user.has_permission('planner:operate')

    return render_template(
        'planner.html',
        plans_by_status=plans_by_status,
        status_filter=status_filter,
        urgency_filter=urgency_filter,
        folio_search=folio_search,
        can_operate=can_operate,
        PLAN_URGENCIES=PLAN_URGENCIES
    )


@app.route('/planner/<int:plan_id>')
@login_required
@require_permission('planner:view')
def planner_detail(plan_id):
    """FastPlanner plan detail view."""
    plan = DistributionPlan.query.get_or_404(plan_id)
    
    page = request.args.get('page', 1, type=int)
    per_page = 10
    
    lines_query = plan.lines.join(Product).join(Store).order_by(Product.sku, Store.name)
    total_lines = lines_query.count()
    lines = lines_query.offset((page - 1) * per_page).limit(per_page).all()
    total_pages = max(1, (total_lines + per_page - 1) // per_page)
    
    activities = plan.activity_logs.limit(50).all()
    can_operate = current_user.has_permission('planner:operate')

    return render_template(
        'planner_detail.html',
        plan=plan,
        lines=lines,
        activities=activities,
        page=page,
        total_pages=total_pages,
        total_lines=total_lines,
        can_operate=can_operate
    )


@app.route('/planner/<int:plan_id>/take', methods=['POST'])
@login_required
@require_permission('planner:operate')
def planner_take(plan_id):
    """Take a plan: move from APPROVED to IN_PROGRESS."""
    plan = DistributionPlan.query.get_or_404(plan_id)
    
    if plan.status != 'APPROVED':
        flash('Solo se pueden tomar planes en estado APROBADO', 'warning')
        return redirect(url_for('planner'))
    
    plan.status = 'IN_PROGRESS'
    plan.assigned_to_user_id = current_user.id
    plan.started_at = datetime.utcnow()
    
    log = PlanActivityLog(
        plan_id=plan.id,
        action='TAKEN',
        user_id=current_user.id,
        comment=f'Plan tomado por {current_user.username}'
    )
    db.session.add(log)
    db.session.commit()
    
    flash(f'Plan {plan.folio} tomado exitosamente', 'success')
    return redirect(url_for('planner'))


@app.route('/planner/<int:plan_id>/pack', methods=['POST'])
@login_required
@require_permission('planner:operate')
def planner_pack(plan_id):
    """Pack a plan: move from IN_PROGRESS to PACKED."""
    plan = DistributionPlan.query.get_or_404(plan_id)
    
    if plan.status != 'IN_PROGRESS':
        flash('Solo se pueden empacar planes en estado EN PROGRESO', 'warning')
        return redirect(url_for('planner'))
    
    plan.status = 'PACKED'
    
    log = PlanActivityLog(
        plan_id=plan.id,
        action='PACKED',
        user_id=current_user.id,
        comment=f'Plan empacado por {current_user.username}'
    )
    db.session.add(log)
    db.session.commit()
    
    flash(f'Plan {plan.folio} marcado como EMPACADO', 'success')
    return redirect(url_for('planner'))


@app.route('/planner/<int:plan_id>/dispatch', methods=['POST'])
@login_required
@require_permission('planner:operate')
def planner_dispatch(plan_id):
    """Dispatch a plan: move from PACKED to DISPATCHED."""
    plan = DistributionPlan.query.get_or_404(plan_id)
    
    if plan.status != 'PACKED':
        flash('Solo se pueden despachar planes en estado EMPACADO', 'warning')
        return redirect(url_for('planner'))
    
    plan.status = 'DISPATCHED'
    
    log = PlanActivityLog(
        plan_id=plan.id,
        action='DISPATCHED',
        user_id=current_user.id,
        comment=f'Plan despachado por {current_user.username}'
    )
    db.session.add(log)
    db.session.commit()
    
    flash(f'Plan {plan.folio} marcado como DESPACHADO', 'success')
    return redirect(url_for('planner'))


@app.route('/planner/<int:plan_id>/close', methods=['POST'])
@login_required
@require_permission('planner:operate')
def planner_close(plan_id):
    """Close a plan: move from DISPATCHED to CLOSED."""
    plan = DistributionPlan.query.get_or_404(plan_id)
    
    if plan.status != 'DISPATCHED':
        flash('Solo se pueden cerrar planes en estado DESPACHADO', 'warning')
        return redirect(url_for('planner'))
    
    plan.status = 'CLOSED'
    plan.closed_at = datetime.utcnow()
    
    log = PlanActivityLog(
        plan_id=plan.id,
        action='CLOSED',
        user_id=current_user.id,
        comment=f'Plan cerrado por {current_user.username}'
    )
    db.session.add(log)
    db.session.commit()
    
    flash(f'Plan {plan.folio} cerrado exitosamente', 'success')
    return redirect(url_for('planner'))


@app.route('/planner/<int:plan_id>/notes', methods=['POST'])
@login_required
@require_permission('planner:operate')
def planner_notes(plan_id):
    """Update warehouse notes for a plan."""
    plan = DistributionPlan.query.get_or_404(plan_id)
    
    notes = request.form.get('notes_warehouse', '').strip()
    plan.notes_warehouse = notes
    
    log = PlanActivityLog(
        plan_id=plan.id,
        action='NOTES_UPDATED',
        user_id=current_user.id,
        comment='Notas de almacén actualizadas'
    )
    db.session.add(log)
    db.session.commit()
    
    flash('Notas actualizadas', 'success')
    return redirect(url_for('planner_detail', plan_id=plan_id))


@app.route('/planner/<int:plan_id>/export_picking')
@login_required
@require_permission('planner:view')
def planner_export_picking(plan_id):
    """Export picking list to Excel."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    plan = DistributionPlan.query.get_or_404(plan_id)
    lines = plan.lines.join(Product).join(Store).order_by(Product.sku, Store.name).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Picking List'
    
    header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ['SKU', 'Producto', 'Tienda', 'Cantidad', 'Notas']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    for row_idx, line in enumerate(lines, 2):
        ws.cell(row=row_idx, column=1, value=line.product.sku).border = thin_border
        ws.cell(row=row_idx, column=2, value=line.product.name).border = thin_border
        ws.cell(row=row_idx, column=3, value=line.store.name).border = thin_border
        ws.cell(row=row_idx, column=4, value=line.qty_planned).border = thin_border
        ws.cell(row=row_idx, column=5, value=line.line_notes or '').border = thin_border
        
        ws.cell(row=row_idx, column=1).number_format = '@'
    
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 30
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f'picking_{plan.folio}_{date.today().isoformat()}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


def init_database():
    """Initialize database with safe schema migration for RBAC and Audit columns."""
    from sqlalchemy import inspect, text
    
    db.create_all()
    
    inspector = inspect(db.engine)
    
    user_columns = [col['name'] for col in inspector.get_columns('user')]
    
    if 'role' not in user_columns:
        with db.engine.connect() as conn:
            conn.execute(text("ALTER TABLE user ADD COLUMN role VARCHAR(50) DEFAULT 'Admin'"))
            conn.commit()
        print("✅ Added 'role' column to user table")
    
    if 'is_active' not in user_columns:
        with db.engine.connect() as conn:
            conn.execute(text("ALTER TABLE user ADD COLUMN is_active BOOLEAN DEFAULT 1"))
            conn.commit()
        print("✅ Added 'is_active' column to user table")
    
    if 'run' in inspector.get_table_names():
        run_columns = [col['name'] for col in inspector.get_columns('run')]
        
        if 'user_id' not in run_columns:
            with db.engine.connect() as conn:
                conn.execute(text("ALTER TABLE run ADD COLUMN user_id INTEGER"))
                conn.commit()
            print("✅ Added 'user_id' column to run table")
        
        if 'rows_count' not in run_columns:
            with db.engine.connect() as conn:
                conn.execute(text("ALTER TABLE run ADD COLUMN rows_count INTEGER"))
                conn.commit()
            print("✅ Added 'rows_count' column to run table")
        
        if 'predictions_count' not in run_columns:
            with db.engine.connect() as conn:
                conn.execute(text("ALTER TABLE run ADD COLUMN predictions_count INTEGER"))
                conn.commit()
            print("✅ Added 'predictions_count' column to run table")
        
        if 'is_active' not in run_columns:
            with db.engine.connect() as conn:
                conn.execute(text("ALTER TABLE run ADD COLUMN is_active BOOLEAN DEFAULT 0"))
                conn.commit()
            print("✅ Added 'is_active' column to run table")
    
    admin = User.query.filter_by(username="admin").first()
    if not admin:
        admin = User(
            username="admin",
            password_hash=generate_password_hash("admin"),
            role='Admin',
            is_active=True
        )
        db.session.add(admin)
        db.session.commit()
        print("[INIT] Admin created")
    else:
        # Always ensure admin has correct role and is active (do NOT touch password)
        needs_update = False
        if admin.role != 'Admin':
            admin.role = 'Admin'
            needs_update = True
        if not admin.is_active:
            admin.is_active = True
            needs_update = True
        if needs_update:
            db.session.commit()
            print("[INIT] Admin promoted to admin role")
        else:
            print("[INIT] Admin OK")


if __name__ == "__main__":
    with app.app_context():
        init_database()

    app.run(host='0.0.0.0', port=5000, debug=True)